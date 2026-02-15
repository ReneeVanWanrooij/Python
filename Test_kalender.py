import tkinter as tk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from datetime import date, timedelta, datetime
import calendar
import holidays
import os
import requests

# ========================
# HELPERS
# ========================
def daterange(start_date, end_date):
    for n in range((end_date - start_date).days + 1):
        yield start_date + timedelta(n)

def hhmm_to_minutes(hhmm):
    if not hhmm or hhmm == "":
        return 0
    parts = str(hhmm).split(":")
    return int(parts[0]) * 60 + int(parts[1])

def minutes_to_hhmm(minutes):
    h = minutes // 60
    m = minutes % 60
    return f"{h:02d}:{m:02d}"

def get_cell_color(dt, val=None, day_number_row=False):
    today = date.today()
    if val == "" or val is None:
        return "white"
    if day_number_row:
        return "lightgrey"
    if dt == today:
        return "deepskyblue"
    if dt:
        str_date = dt.strftime("%Y-%m-%d")
        if str_date in vakantie_dagen:
            return "yellow"
        if str_date in school_vakanties:
            return "#FFDDDD"
        if dt in nl_holidays:
            return "red"
        if dt.weekday() >= 5:
            return "lightgrey"
    if val and "wk totaal" in str(val):
        return "lightgreen"
    if val and "maand totaal" in str(val):
        return "green"
    if val and "wk" in str(val) and "totaal" not in str(val):
        return "orange"
    return "white"

# ========================
# TOOLTIP
# ========================
class Tooltip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        widget.bind("<Enter>", self.show)
        widget.bind("<Leave>", self.hide)

    def show(self, event=None):
        if self.tipwindow:
            return
        x, y, cx, cy = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 20
        y = y + self.widget.winfo_rooty() + 20
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, justify='left',
                         background="lightyellow", relief='solid', borderwidth=1,
                         font=("Arial", 9))
        label.pack(ipadx=1)

    def hide(self, event=None):
        if self.tipwindow:
            self.tipwindow.destroy()
            self.tipwindow = None

# ========================
# CONFIG
# ========================
year = datetime.now().year
excel_path = f"Time_tabel_{year}.xlsx"

# Extra vrije dagen (mock)
vakantie_dagen = {"2026-02-16": "Verlof Ruben", "2026-05-05": "Vrije dag", "2026-08-10": "Verlof Jasper"}
nl_holidays = holidays.NL(years=[year])

months_nl = ["Januari", "Februari", "Maart", "April", "Mei", "Juni",
             "Juli", "Augustus", "September", "Oktober", "November", "December"]
weekdays_abbr = ["MA", "DI", "WO", "DO", "VR", "ZA", "ZO"]

# ========================
# MOCK DATA GENEREREN ALS EXCEL NOG NIET BESTAAT
# ========================
if not os.path.exists(excel_path):
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Data-tabblad
    ws_data = wb.create_sheet("vrije dagen")
    ws_data.append(["Datum", "Omschrijving"])
    for d, desc in vakantie_dagen.items():
        ws_data.append([d, desc])

    # Maand-tabbladen
    for month in range(1, 13):
        ws = wb.create_sheet(title=months_nl[month - 1])
        ws.append(["Week"] + weekdays_abbr)
        days_in_month = calendar.monthrange(year, month)[1]
        day = 1
        while day <= days_in_month:
            week_row = [f"wk{date(year, month, day).isocalendar()[1]}"]
            for wd in range(7):
                try:
                    dt = date(year, month, day)
                    week_row.append("00:00")  # mock uren
                    day += 1
                except ValueError:
                    week_row.append("")
            ws.append(week_row)

    wb.save(excel_path)

# ========================
# SCHOOLVAKANTIES LADEN EN OPSLAAN
# ========================
def load_school_holidays(year, region="zuid", excel_path=None):
    print("Schoolvakanties ophalen van overheid...")
    url = "https://opendata.rijksoverheid.nl/v1/infotypes/schoolholidays?output=json"
    try:
        resp = requests.get(url).json()
    except Exception as e:
        print("Ophalen mislukt:", e)
        # fallback: lees uit Excel indien beschikbaar
        if excel_path and os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            if "Schoolvakanties" in wb.sheetnames:
                ws = wb["Schoolvakanties"]
                holidays_dict = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2)}
                print(f"Schoolvakanties geladen uit Excel ({excel_path})")
                return holidays_dict
        return {}

    school_vakanties = {}
    found_year = False

    for item in resp:  # resp is een list
        for content in item.get("content", []):
            if str(year) not in content.get("schoolyear", ""):
                continue
            for vac in content.get("vacations", []):
                for reg in vac.get("regions", []):
                    if region.lower() in reg["region"].lower():
                        start = date.fromisoformat(reg["startdate"][:10])
                        end = date.fromisoformat(reg["enddate"][:10])
                        found_year = True
                        for d in daterange(start, end):
                            if d.year == year:
                                school_vakanties[d.strftime("%Y-%m-%d")] = vac.get("type", "").strip()

    if not found_year:
        print(f"Geen schoolvakanties gevonden voor {year}.")
    else:
        print(f"Schoolvakanties gevonden voor {year}: {len(school_vakanties)} dagen")

    # Opslaan in Excel voor offline gebruik
    if excel_path:
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()

        if "Schoolvakanties" in wb.sheetnames:
            ws = wb["Schoolvakanties"]
            wb.remove(ws)
        ws = wb.create_sheet(title="Schoolvakanties")
        ws.append(["Datum", "Type"])
        for k, v in sorted(school_vakanties.items()):
            ws.append([k, v])
        wb.save(excel_path)
        print(f"Schoolvakanties opgeslagen in Excel ({excel_path})")

    return school_vakanties
school_vakanties = load_school_holidays(year, region="zuid", excel_path=excel_path)

# ========================
# LOAD EXCEL
# ========================
wb = load_workbook(excel_path)

# ========================
# LAAD VRIJE DAGEN UIT EXCEL
# ========================
vakantie_dagen = {}
if "vrije dagen" in wb.sheetnames:
    ws_data = wb["vrije dagen"]
    for row in ws_data.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            # Zorg dat datum als string staat in 'YYYY-MM-DD'
            if isinstance(row[0], datetime):
                key = row[0].strftime("%Y-%m-%d")
            else:
                key = str(row[0])
            vakantie_dagen[key] = row[1]

# ========================
# HIGHLIGHT VANDAAG IN EXCEL VOOR HUIDIGE MAAND
# ========================
today = date.today()
today_month_name = months_nl[today.month - 1]

# Controleer of het tabblad bestaat
if today_month_name in wb.sheetnames:
    ws_today = wb[today_month_name]

    # Loop door alle rijen en kolommen van dit tabblad
    for r_idx, row in enumerate(ws_today.iter_rows(min_row=2)):
        for c_idx, cell in enumerate(row):
            if c_idx != 0 and cell.value and isinstance(cell.value, str):
                week_num = row[0].value
                try:
                    start_of_week = date.fromisocalendar(today.year, int(week_num[2:]), 1)
                    dt = start_of_week + timedelta(days=c_idx-1)
                except:
                    dt = None

                if dt == today:
                    # Lichtblauw invullen
                    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                    # Optioneel: waarde aanpassen
                    # cell.value = "08:00"

    if today_month_name in wb.sheetnames:
        ws_today = wb[today_month_name]
        # Tabbladkleur instellen
        ws_today.sheet_properties.tabColor = "ADD8E6"  # HEX zonder #

    # Sla Excel op met highlight
    wb.save(excel_path)


# ========================
# GUI
# ========================
root = tk.Tk()
root.title(f"Jaaroverzicht {year}")

canvas = tk.Canvas(root)
frame = tk.Frame(canvas)
canvas.create_window((0, 0), window=frame, anchor='nw')
canvas.pack(fill='both', expand=True)

month_col_width = 8
months_per_row = 4
current_row_gui = 0
current_col_gui = 0

for month in range(1, 13):
    month_name = months_nl[month - 1]
    ws = wb[month_name]

    # Maand label
    lbl_month = tk.Label(frame, text=month_name, bg="brown", fg="white",
                         font=("Arial", 12, "bold"))
    lbl_month.grid(row=current_row_gui, column=current_col_gui, columnspan=9, sticky="nsew", pady=(5,0))

    # Days header
    header = ["WK"] + weekdays_abbr + ["Wk totaal"]
    for c_idx, h in enumerate(header):
        lbl = tk.Label(frame, text=h, font=("Arial", 10, "bold"), borderwidth=1, relief="solid")
        lbl.grid(row=current_row_gui + 1, column=current_col_gui + c_idx, sticky="nsew")

    # Bereken alle dagen van de maand
    days_in_month = calendar.monthrange(year, month)[1]
    first_day = date(year, month, 1)
    week_rows = {}  # key=week_num, value=list van dt's per dag
    for day_offset in range(days_in_month):
        dt = first_day + timedelta(days=day_offset)
        week_num = dt.isocalendar()[1]
        weekday_idx = dt.weekday()  # 0=ma, 6=zo
        if week_num not in week_rows:
            week_rows[week_num] = [None]*7
        week_rows[week_num][weekday_idx] = dt

    # Teken de weken
    for r_idx, (week_num, week_days) in enumerate(sorted(week_rows.items())):
        week_total = 0

        # Weeknummer
        lbl_weeknum = tk.Label(frame, text=f"wk{week_num}", borderwidth=1, relief="solid",
                               width=month_col_width, height=3, bg="orange")
        lbl_weeknum.grid(row=current_row_gui + 2 + r_idx*3, column=current_col_gui, rowspan=2, sticky="nsew")

        # Dagnummer rij
        for c_idx, dt in enumerate(week_days):
            if dt:
                day_number = dt.day
                bg_color = "#8A2BE2" if dt.weekday() < 5 else "lightgrey"
                lbl_day = tk.Label(frame, text=str(day_number), borderwidth=1, relief="solid",
                                   width=month_col_width, height=1, bg=bg_color)
                lbl_day.grid(row=current_row_gui + 2 + r_idx*3, column=current_col_gui + 1 + c_idx, sticky="nsew")
            # lege dt = geen label (blijft blanco)

        # Week totaal (links naast dagnummer rij,zelfde hoogte als dagnummer)
        for c_idx, dt in enumerate(week_days):
            val = ""
            if dt:
                # Haal waarde uit Excel
                excel_week_row = None
                for er in ws.iter_rows(min_row=2, values_only=True):
                    if er[0] == f"wk{week_num}":
                        excel_week_row = er
                        break
                if excel_week_row and dt.weekday() < len(excel_week_row)-1:
                    val = excel_week_row[dt.weekday()+1]  # +1 want kolom 0=weeknummer
                    if val and ":" in str(val):
                        week_total += hhmm_to_minutes(val)

            # Alleen de Wk-totaal cel tekenen in laatste kolom
        lbl_total = tk.Label(frame, text=minutes_to_hhmm(week_total),
                             borderwidth=1, relief="solid",
                             width=month_col_width, height=3,
                             bg="lightgreen")
        lbl_total.grid(row=current_row_gui + 2 + r_idx * 3, column=current_col_gui + len(header) - 1, rowspan=2,sticky="nsew")

        # Uren rij
        for c_idx, dt in enumerate(week_days):
            val = ""
            if dt:
                # Excel waarde
                excel_week_row = None
                for er in ws.iter_rows(min_row=2, values_only=True):
                    if er[0] == f"wk{week_num}":
                        excel_week_row = er
                        break
                if excel_week_row and dt.weekday() < len(excel_week_row)-1:
                    val = excel_week_row[dt.weekday()+1]

            if val != "":
                lbl_val = tk.Label(frame, text=str(val), borderwidth=1, relief="solid",
                                   width=month_col_width, height=2, bg=get_cell_color(dt, val))
                lbl_val.grid(row=current_row_gui + 3 + r_idx*3, column=current_col_gui + 1 + c_idx, sticky="nsew")
                # Tooltips
                str_date = dt.strftime("%Y-%m-%d")
                if str_date in vakantie_dagen:
                    Tooltip(lbl_val, vakantie_dagen[str_date])
                elif dt in nl_holidays:
                    Tooltip(lbl_val, nl_holidays.get(dt))
                elif str_date in school_vakanties:
                    Tooltip(lbl_val, school_vakanties[str_date])

    # Verticale scheiding
    separator = tk.Label(frame, width=2, bg="white")
    separator.grid(row=current_row_gui, column=current_col_gui + len(header),
                   rowspan=len(week_rows)*3 + 2, sticky="nsew")

    current_col_gui += len(header) + 1
    if month % months_per_row == 0:
        current_col_gui = 0
        current_row_gui += len(week_rows)*3 + 4

frame.update_idletasks()
canvas.configure(scrollregion=canvas.bbox('all'))
root.geometry(f"{frame.winfo_reqwidth()+80}x{frame.winfo_reqheight()+50}")
root.mainloop()

