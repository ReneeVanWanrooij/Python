import os
import time
import tkinter as tk
from tkinter import messagebox
from datetime import datetime, date, timedelta
import ctypes
from ctypes import wintypes
import psutil
import pygetwindow as gw
import openpyxl
from openpyxl.styles import PatternFill
import subprocess
import sys
import pystray
from PIL import Image, ImageDraw
import threading


# ================================
# CONFIG
# ================================
YEAR = datetime.now().year
excel_file = os.path.join(os.path.dirname(__file__), f"Time_tabel_{YEAR}.xlsx")

IDLE_THRESHOLD_SEC = 60
IDLE_WARNING_SEC = 10

months_nl = [
    "Januari","Februari","Maart","April","Mei","Juni",
    "Juli","Augustus","September","Oktober","November","December"
]

weekdays_abbr = ["Ma","Di","Wo","Do","Vr","Za","Zo"]

SAVE_INTERVAL = 5
save_tick = 0

# ================================
# GLOBALS
# ================================
wb = None
ws_today = None
today_cell = None

work_seconds = 0
idle_seconds = 0
call_seconds = 0
in_pause = False

current_day = date.today()

idle_warning_visible = False
is_idle_state = False

# ================================
# HELPERS
# ================================
def sec_to_time(seconds):
    h = seconds//3600
    m = (seconds%3600)//60
    s = seconds%60
    return f"{h:02d}:{m:02d}:{s:02d}"

# ================================
# hh:mm:ss omzetten naar Secondes
# ================================
def time_to_sec(t):
    if not t:
        return 0
    if isinstance(t, int):
        return t
    if isinstance(t, str):
        parts = t.replace(".", ":").split(":")
        parts = [int(p) for p in parts]
        if len(parts) == 3:
            h, m, s = parts
        elif len(parts) == 2:
            h, m = parts
            s = 0
        else:
            return 0
        return h*3600 + m*60 + s
    return 0


# ================================
# Tray icoontje maken
# ================================

def create_tray_icon():
    # simpel vierkant icon
    img = Image.new("RGB", (64, 64), "black")
    d = ImageDraw.Draw(img)
    d.rectangle((16, 16, 48, 48), fill="lime")

    def on_open(icon, item):
        root.after(0, root.deiconify)

    def on_exit(icon, item):
        icon.stop()
        root.after(0, root.destroy)

    menu = pystray.Menu(
        pystray.MenuItem("Open", on_open),
        pystray.MenuItem("Exit", on_exit)
    )

    icon = pystray.Icon("timer", img, "Werk timer", menu)
    icon.run()

# ================================
# Grove data logger
# ================================


def log_data():
    global wb, work_seconds, idle_seconds, call_seconds

    today_str = datetime.now().strftime("%Y-%m-%d")
    now_time = datetime.now().strftime("%H:%M:%S")

    ws_log = wb["Data_log"]

    # check of vandaag al een regel heeft (vanaf rij 2)
    found = False
    for row in ws_log.iter_rows(min_row=2):
        if str(row[0].value) == today_str:
            # update bestaande regel
            row[2].value = sec_to_time(work_seconds)
            row[3].value = sec_to_time(idle_seconds)
            row[4].value = sec_to_time(call_seconds)
            found = True
            break

    if not found:
        # voeg nieuwe regel toe op rij 2 of lager
        ws_log.append([
            today_str,
            now_time,
            sec_to_time(work_seconds),
            sec_to_time(idle_seconds),
            sec_to_time(call_seconds)
        ])

    try:
        wb.save(excel_file)
    except PermissionError:
        print("Excel is open, kan niet opslaan. Probeer later opnieuw...")


# ================================
# EXCEL MAKEN
# ================================

import calendar
from openpyxl import Workbook

def ensure_excel_exists():
    if os.path.exists(excel_file):
        return

    print("Excel wordt aangemaakt:", excel_file)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    ws_vrij = wb.create_sheet("Data_log")
    ws_vrij.append(["Datum","Tijd","Werk_sec","Idle_sec","Call_sec"])

    ws_vrij = wb.create_sheet("vrije dagen")
    ws_vrij.append(["Datum", "Omschrijving"])

    for month in range(1, 13):
        ws = wb.create_sheet(title=months_nl[month-1])
        # Header
        ws.append(["Week"] + weekdays_abbr + ["Wk totaal"])  # optioneel weektotaal

        days_in_month = calendar.monthrange(YEAR, month)[1]
        first_day = date(YEAR, month, 1)

        # bouw dict: key=week_num, value=list van dt's per dag
        week_rows = {}
        for day_offset in range(days_in_month):
            dt = first_day + timedelta(days=day_offset)
            week_num = dt.isocalendar()[1]
            weekday_idx = dt.weekday()  # 0=ma, 6=zo
            if week_num not in week_rows:
                week_rows[week_num] = [None]*7
            week_rows[week_num][weekday_idx] = dt

        # vul werkblad per week
        for week_num, week_days in sorted(week_rows.items()):
            week_row = [f"wk{week_num}"]
            for dt in week_days:
                if dt:
                    week_row.append("00:00:00")
                else:
                    week_row.append("")
            week_row.append("")  # weektotaal cel
            ws.append(week_row)
        # ---------------------------
        # Schoolvakanties ophalen en opslaan
        # ---------------------------
    try:
        import requests
        url = "https://opendata.rijksoverheid.nl/v1/infotypes/schoolholidays?output=json"
        resp = requests.get(url).json()
        school_vakanties = {}

        for item in resp:
            for content in item.get("content", []):
                if str(YEAR) not in content.get("schoolyear", ""):
                    continue
                for vac in content.get("vacations", []):
                    for reg in vac.get("regions", []):
                        if "zuid" in reg["region"].lower():
                            start = date.fromisoformat(reg["startdate"][:10])
                            end = date.fromisoformat(reg["enddate"][:10])
                            for d in (start + timedelta(n) for n in range((end - start).days + 1)):
                                if d.year == YEAR:
                                    school_vakanties[d.strftime("%Y-%m-%d")] = vac.get("type", "").strip()

        ws_vac = wb.create_sheet("Schoolvakanties")
        ws_vac.append(["Datum", "Type"])
        for k, v in sorted(school_vakanties.items()):
            ws_vac.append([k, v])

        print(f"Schoolvakanties ({len(school_vakanties)} dagen) toegevoegd aan Excel.")

    except Exception as e:
        print("Kon schoolvakanties niet ophalen:", e)

    wb.save(excel_file)

# ================================
# EXCEL laden
# ================================
def init_excel():
    global wb, ws_today, today_cell, work_seconds, idle_seconds, call_seconds

    ensure_excel_exists()
    wb = openpyxl.load_workbook(excel_file)

    today = date.today()
    today_str = today.strftime("%Y-%m-%d")
    month_name = months_nl[today.month-1]
    ws_today = wb[month_name]

    # --- Wis oude highlights in maandtabblad ---
    for row in ws_today.iter_rows(min_row=2):
        for cell in row[1:]:
            cell.fill = PatternFill(fill_type=None)  # geen kleur

    # --- Zoek en highlight huidige dag ---
    for row in ws_today.iter_rows(min_row=2):
        week_label = row[0].value
        if not week_label:
            continue
        try:
            week_num = int(str(week_label).replace("wk", ""))
            start_of_week = date.fromisocalendar(today.year, week_num, 1)
        except:
            continue

        for col_index, cell in enumerate(row[1:], start=1):
            dt = start_of_week + timedelta(days=col_index-1)
            if dt == today:
                today_cell = cell

                # werk seconden instellen
                if cell.value and ":" in str(cell.value):
                    parts = list(map(int, str(cell.value).split(":")))
                    if len(parts) == 3:
                        h, m, s = parts
                    elif len(parts) == 2:
                        h, m = parts
                        s = 0
                    else:
                        h = m = s = 0
                    work_seconds = h * 3600 + m * 60 + s
                else:
                    work_seconds = 0

                # --- GROVE DATA ophalen ---
                if "Data_log" in wb.sheetnames:
                    ws_log = wb["Data_log"]
                    for row_log in ws_log.iter_rows(min_row=2):
                        if row_log[0].value == today_str:
                            idle_seconds = time_to_sec(row_log[3].value)
                            call_seconds = time_to_sec(row_log[4].value)
                            break

                # huidige dag kleur
                cell.fill = PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid")
                print("Timer start vanaf:", sec_to_time(work_seconds))
                break

    # --- Kleur de sheet-tab van de huidige maand ---
    ws_today.sheet_properties.tabColor = "00BFFF"  # lichtblauw, hex zonder #

    wb.save(excel_file)


# ================================
# SAVE
# ================================
def save_excel():
    global wb

    if not today_cell:
        return

    h = work_seconds // 3600
    m = (work_seconds % 3600)//60
    s = work_seconds % 60
    # today_cell.value = f"{h:02d}:{m:02d}:{s:02d}"
    today_cell.value = f"{h:02d}:{m:02d}:{s:02d}"

    today_cell.fill = PatternFill(
        start_color="ADD8E6",
        end_color="ADD8E6",
        fill_type="solid"
    )

    try:
        wb.save(excel_file)
    except PermissionError:
        # Excel staat open → later opnieuw proberen
        print("Excel is open, kan niet opslaan. Probeer later opnieuw...")
    except Exception as e:
        print("Onverwachte Excel save fout:", e)


# ================================
# IDLE DETECTIE
# ================================
class LASTINPUTINFO(ctypes.Structure):
    _fields_ = [("cbSize", wintypes.UINT),("dwTime",wintypes.DWORD)]

def get_idle_time():
    lii = LASTINPUTINFO()
    lii.cbSize = ctypes.sizeof(LASTINPUTINFO)
    ctypes.windll.user32.GetLastInputInfo(ctypes.byref(lii))
    return ctypes.windll.kernel32.GetTickCount() - lii.dwTime

# ================================
# CALL DETECTIE
# ================================
def detect_call():
    try:
        for proc in psutil.process_iter(['name']):
            name = proc.info['name'].lower()
            if 'teams' in name or 'chrome' in name or 'edge' in name:
                windows = gw.getWindowsWithTitle('Meeting') + gw.getWindowsWithTitle('Call') + gw.getWindowsWithTitle('Vergadering')
                if windows:
                    return True
    except:
        pass
    return False

# ================================
# GUI
# ================================
class TimerGUI:
    def __init__(self, root):
        self.root = root
        self.root.geometry("420x50")  # iets breder
        self.root.attributes('-topmost', True)
        self.root.configure(bg='black')
        self.root.overrideredirect(True)

        # frame container
        frame = tk.Frame(root, bg="black")
        frame.pack(fill=tk.BOTH, expand=True)

        # timer label
        self.label = tk.Label(
            frame,
            fg='lime',
            bg='black',
            font=('Consolas',12,'bold'),
            anchor="w"
        )
        self.label.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        # idle warning label (standaard verborgen)
        self.idle_label = tk.Label(
            root,
            text="⚠ IDLE binnenkort",
            fg="black",
            bg="yellow",
            font=('Consolas', 10, 'bold')
        )

        # smalle kalender knop
        self.btn = tk.Button(
            frame,
            text="📅",
            width=3,
            height=3,
            command=self.open_kalender,
            bg="#222",
            fg="white",
            relief="flat"
        )
        self.btn.pack(side=tk.RIGHT, padx=3, pady=3)

        # drag support
        self.label.bind('<Button-1>', self.start_drag)
        self.label.bind('<B1-Motion>', self.drag)

        self.drag_data = {"x":0,"y":0}

        self.warning_alpha = 1.0
        self.warning_direction = -0.1
        self.warning_job = None

    def start_drag(self,event):
        self.drag_data["x"]=event.x_root-self.root.winfo_x()
        self.drag_data["y"]=event.y_root-self.root.winfo_y()

    def drag(self,event):
        x = event.x_root-self.drag_data["x"]
        y = event.y_root-self.drag_data["y"]
        self.root.geometry(f"+{x}+{y}")

    def update_display(self):
        self.label.config(
            text=f"T: {sec_to_time(work_seconds)}  I: {sec_to_time(idle_seconds)}  C: {sec_to_time(call_seconds)}"
        )

    def start_warning_animation(self):
        if self.warning_job is None:
            self.animate_warning()

    def stop_warning_animation(self):
        if self.warning_job:
            self.root.after_cancel(self.warning_job)
            self.warning_job = None
            self.idle_label.config(bg="yellow")

    def animate_warning(self):
        # fade effect
        self.warning_alpha += self.warning_direction

        if self.warning_alpha <= 0.3 or self.warning_alpha >= 1:
            self.warning_direction *= -1

        # kleur berekenen
        intensity = int(255 * self.warning_alpha)
        color = f"#ffff{intensity:02x}"

        self.idle_label.config(bg=color)

        self.warning_job = self.root.after(100, self.animate_warning)

    # ================================
    # IDLE WARNING VISUAL
    # ================================
    def show_idle_warning(self):
        if not self.idle_label.winfo_ismapped():
            self.root.geometry("420x70")
            self.idle_label.pack(side=tk.BOTTOM, fill=tk.X)
            self.start_warning_animation()

    def hide_idle_warning(self):
        if self.idle_label.winfo_ismapped():
            self.stop_warning_animation()
            self.idle_label.pack_forget()
            self.root.geometry("420x50")

    # ================================
    # OPEN KALENDER SCRIPT
    # ================================
    def open_kalender(self):
        script_path = os.path.join(os.path.dirname(__file__), "test_kalender.py")

        try:
            subprocess.Popen([sys.executable, script_path])
        except Exception as e:
            print("Kon test_kalender.py niet starten:", e)

    # ================================
    # Check of het middernacht is.
    # ================================
def check_midnight():
    global current_day, today_cell, ws_today

    now = date.today()

    if now != current_day:
        print("Nieuwe dag gedetecteerd → reset timer")
        current_day = now

        # Excel opnieuw laden en juiste cel zoeken
        init_excel()

# ================================
# LOOP
# ================================
def update_loop():
    global work_seconds, idle_seconds, call_seconds
    global save_tick, idle_warning_visible, is_idle_state

    check_midnight()

    idle_ms = get_idle_time()
    idle_sec = idle_ms / 1000

    # -------------------------
    # WARNING PHASE
    # -------------------------
    if IDLE_THRESHOLD_SEC - IDLE_WARNING_SEC <= idle_sec < IDLE_THRESHOLD_SEC:
        gui.show_idle_warning()
    else:
        gui.hide_idle_warning()

    # -------------------------
    # IDLE PHASE
    # -------------------------
    if idle_sec >= IDLE_THRESHOLD_SEC:
        is_idle_state = True
    else:
        is_idle_state = False

    # -------------------------
    # TELLERS
    # -------------------------
    if is_idle_state:
        idle_seconds += 1
    else:
        work_seconds += 1

    if detect_call():
        call_seconds += 1

    gui.update_display()

    # -------------------------
    # SAVE
    # -------------------------
    save_tick += 1
    if save_tick >= SAVE_INTERVAL:
        save_excel()
        log_data()
        save_tick = 0

    root.after(1000, update_loop)



# ================================
# CLOSE
# ================================
def on_closing():
    save_excel()
    root.destroy()

# ================================
# MAIN
# ================================
if __name__ == "__main__":
    init_excel()

    root = tk.Tk()
    gui = TimerGUI(root)

    root.protocol("WM_DELETE_WINDOW", on_closing)

    # tray thread starten
    tray_thread = threading.Thread(target=create_tray_icon, daemon=True)
    tray_thread.start()

    root.after(1000, update_loop)
    root.mainloop()

