import calendar
import hashlib
import importlib
import json
import math
import os
import re
import shutil
import signal
import subprocess
import sys
import time
import tempfile
import ctypes
import urllib.request
import urllib.error
from ctypes import wintypes
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Protocol

try:
    import tkinter as tk
    from tkinter import ttk
except Exception:
    tk = None
    ttk = None

_MISSING_REQUIRED_DEPS: list[str] = []

try:
    import holidays
except Exception:
    holidays = None
    _MISSING_REQUIRED_DEPS.append("holidays")
try:
    import psutil
except Exception:
    psutil = None
try:
    import pygetwindow as gw
except Exception:
    gw = None
try:
    from pycaw.pycaw import AudioUtilities
except Exception:
    AudioUtilities = None
try:
    import win32com.client as win32_client
except Exception:
    win32_client = None
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill
except Exception:
    Workbook = None
    load_workbook = None
    PatternFill = None
    _MISSING_REQUIRED_DEPS.append("openpyxl")
try:
    from PySide6.QtCore import Qt, Signal, QEvent, QPoint, QObject, QTimer, QPropertyAnimation, QEasingCurve, QRect, QParallelAnimationGroup, Property, QLockFile, QAbstractNativeEventFilter, QDate
    from PySide6.QtGui import QAction, QColor, QFont, QBrush, QLinearGradient, QGradient, QPen, QPixmap, QPainter, QIcon, QPolygon
    from PySide6.QtWidgets import (
        QApplication,
        QComboBox,
        QDialog,
        QFormLayout,
        QGridLayout,
        QGroupBox,
        QHBoxLayout,
        QLabel,
        QLineEdit,
        QMainWindow,
        QMessageBox,
        QPushButton,
        QColorDialog,
        QScrollArea,
        QSystemTrayIcon,
        QMenu,
        QStatusBar,
        QStyledItemDelegate,
        QSizePolicy,
        QTabBar,
        QTabWidget,
        QTableWidget,
        QTableWidgetItem,
        QTextEdit,
        QDateEdit,
        QFrame,
        QHeaderView,
        QToolBar,
        QVBoxLayout,
        QWidget,
        QSpinBox,
        QSlider,
        QCheckBox,
        QAbstractItemView,
        QStyle,
    )
except Exception:
    _MISSING_REQUIRED_DEPS.append("PySide6")


# ============================================================================
# MODULE OVERVIEW
# Dit bestand bevat de volledige applicatie in één script: opslag, dialogs,
# kalender, dashboard en timer-overlay. De code is bewust in secties opgebouwd
# zodat je top-down kunt analyseren: eerst helpers, dan opslag, daarna UI shells.
# ============================================================================


def detect_imports(script_path: str) -> list[str]:
    """Detecteert externe imports in een script (top-level package names)."""
    imports = set()
    try:
        with open(script_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                match_import = re.match(r"import\s+([A-Za-z0-9_\.]+)", line)
                match_from = re.match(r"from\s+([A-Za-z0-9_\.]+)\s+import", line)
                if match_import:
                    imports.add(match_import.group(1).split(".")[0])
                elif match_from:
                    imports.add(match_from.group(1).split(".")[0])
    except Exception:
        return []
    stdlib = set(getattr(sys, "stdlib_module_names", set()))
    return sorted(m for m in imports if m and m not in stdlib and m not in {"tkinter"})


BOOTSTRAP_CHECKS: list[tuple[str, str, str]] = [
    ("PySide6", "PySide6", "UI framework"),
    ("openpyxl", "openpyxl", "Excel engine"),
    ("holidays", "holidays", "Feestdagen"),
    ("psutil", "psutil", "Procesdetectie"),
    ("pygetwindow", "pygetwindow", "Window detectie"),
    ("pycaw", "pycaw", "Audio sessie detectie"),
    ("win32com.client", "pywin32", "Outlook agenda hint"),
]
OPTIONAL_BOOTSTRAP_IMPORTS = {"pycaw", "win32com.client"}
CALL_DEBUG_ENABLED = "--call-debug" in sys.argv or os.environ.get("TP_CALL_DEBUG", "").strip() == "1"
CALL_DEBUG_LOG_ENABLED = "--call-debug-log" in sys.argv or os.environ.get("TP_CALL_DEBUG_LOG", "").strip() == "1"


def run_bootstrap(
    checks: list[tuple[str, str, str]],
    dry_run: bool = False,
    simulate_missing: bool = False,
    wait_for_close: bool = False,
) -> bool:
    """Checkt alle dependencies en installeert ontbrekende packages met splash/progress."""
    checks = [(imp, pip_pkg, label) for imp, pip_pkg, label in checks if imp and pip_pkg]
    if not checks:
        return True
    check_rows = [("__python__", "", "Python runtime")] + checks
    close_btn = None
    optional_install_failed: list[str] = []
    trusted_hosts = ("pypi.org", "files.pythonhosted.org")

    def _is_optional(import_name: str) -> bool:
        return import_name in OPTIONAL_BOOTSTRAP_IMPORTS

    def _bootstrap_failure_hint(import_name: str, pip_pkg: str, recent_output: list[str]) -> str:
        txt = " ".join((recent_output or [])).lower()
        if "no matching distribution found" in txt or "could not find a version that satisfies" in txt:
            return "Tip: update pip en controleer je Python-versie (bijv. 'python -m pip install --upgrade pip')."
        if "ssl" in txt or "certificate_verify_failed" in txt or "proxy" in txt:
            return "Tip: netwerk/proxy/certificaat blokkeert pip. Probeer via bedrijfs-VPN of configureer pip-proxy."
        if "access is denied" in txt or "permission denied" in txt:
            return "Tip: onvoldoende rechten. Sluit andere Python-processen en probeer opnieuw."
        if "microsoft visual c++" in txt or "error: microsoft visual c++" in txt:
            return "Tip: installeer Microsoft C++ Build Tools of gebruik een wheel-compatibele Python-versie."
        if import_name == "pycaw" or pip_pkg == "pycaw":
            return "Tip: pycaw is optioneel; app werkt door zonder audio-sessiedetectie."
        return "Tip: controleer de pip-log hierboven en probeer handmatig: python -m pip install <pakket> --user"

    def _ensure_pip_available(log_fn=None) -> bool:
        def _log(msg: str, color: str = "#dce4f0"):
            if callable(log_fn):
                try:
                    log_fn(msg, color)
                except Exception:
                    pass

        try:
            rc = subprocess.call([sys.executable, "-m", "pip", "--version"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            if rc == 0:
                return True
        except Exception:
            pass

        _log("[~] pip niet gevonden, probeer ensurepip...", "#f3b37a")
        try:
            rc = subprocess.call([sys.executable, "-m", "ensurepip", "--upgrade"])
            if rc != 0:
                _log("[X] ensurepip mislukt", "#ff9492")
                return False
        except Exception:
            _log("[X] ensurepip niet beschikbaar", "#ff9492")
            return False

        try:
            rc = subprocess.call([sys.executable, "-m", "pip", "--version"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            if rc == 0:
                _log("[V] pip hersteld via ensurepip", "#7ad18b")
                return True
        except Exception:
            pass
        _log("[X] pip nog steeds niet beschikbaar", "#ff9492")
        return False

    def _contains_ssl_problem(lines: list[str]) -> bool:
        txt = " ".join(lines or []).lower()
        needles = (
            "ssl",
            "certificate verify failed",
            "could not find a suitable tls ca certificate bundle",
            "invalid path",
            "tls ca certificate",
        )
        return any(n in txt for n in needles)

    def _run_pip_and_log(args: list[str], log_fn=None) -> tuple[int, list[str]]:
        recent: list[str] = []

        def _log(msg: str, color: str = "#dce4f0"):
            if callable(log_fn):
                try:
                    log_fn(msg, color)
                except Exception:
                    pass

        try:
            proc = subprocess.Popen(
                [sys.executable, "-m", "pip", *args],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
            )
        except Exception as exc:
            recent.append(str(exc))
            _log(f"[X] Pip starten mislukt: {exc}", "#ff9492")
            return 1, recent

        while True:
            line = proc.stdout.readline() if proc.stdout else ""
            if line == "" and proc.poll() is not None:
                break
            if line:
                clean = line.strip()
                if clean:
                    recent.append(clean)
                    if len(recent) > 25:
                        recent.pop(0)
                    _log(clean[:180], "#f3b37a")
        return int(proc.returncode or 0), recent

    def _pip_upgrade_healthcheck(log_fn=None):
        def _log(msg: str, color: str = "#dce4f0"):
            if callable(log_fn):
                try:
                    log_fn(msg, color)
                except Exception:
                    pass

        _log("[~] Pip health-check (upgrade test)...", "#9ecbff")
        rc, recent = _run_pip_and_log(["install", "--user", "--upgrade", "pip"], log_fn)
        if rc == 0:
            _log("[V] Pip upgrade-check geslaagd", "#7ad18b")
            return
        if _contains_ssl_problem(recent):
            _log("[!] SSL/proxy issue gedetecteerd; retry met trusted-host...", "#f3b37a")
            th_args: list[str] = []
            for host in trusted_hosts:
                th_args.extend(["--trusted-host", host])
            rc2, _recent2 = _run_pip_and_log(["install", "--user", "--upgrade", "pip", *th_args], log_fn)
            if rc2 == 0:
                _log("[V] Pip upgrade-check geslaagd via trusted-host", "#7ad18b")
                return
        _log("[!] Pip upgrade-check overgeslagen (niet-blokkerend).", "#f3b37a")

    def _pip_install_with_retry(import_name: str, pip_pkg: str, log_fn=None) -> tuple[int, list[str]]:
        rc, recent = _run_pip_and_log(["install", pip_pkg, "--user"], log_fn)
        if rc == 0:
            return rc, recent
        if not _contains_ssl_problem(recent):
            return rc, recent
        if callable(log_fn):
            log_fn("[!] SSL/cert probleem gedetecteerd; retry met trusted-host...", "#f3b37a")
        th_args: list[str] = []
        for host in trusted_hosts:
            th_args.extend(["--trusted-host", host])
        rc2, recent2 = _run_pip_and_log(["install", pip_pkg, "--user", *th_args], log_fn)
        return rc2, recent2

    if tk is None or ttk is None:
        py_ok = bool(sys.executable) and os.path.exists(sys.executable)
        if not py_ok:
            return False
        if not dry_run and not _ensure_pip_available():
            return False
        if not dry_run:
            _pip_upgrade_healthcheck()
        for import_name, pip_pkg, _label in checks:
            try:
                if simulate_missing:
                    raise ImportError("simulated missing")
                importlib.import_module(import_name)
                continue
            except Exception:
                pass
            if dry_run:
                continue
            rc, _recent = _pip_install_with_retry(import_name, pip_pkg)
            if rc != 0:
                if _is_optional(import_name):
                    continue
                return False
        return True

    root = tk.Tk()
    root.overrideredirect(not wait_for_close)
    root.geometry("760x560+500+220")
    root.configure(bg="#1c212b")

    frame = tk.Frame(root, bg="#1c212b", padx=12, pady=10)
    frame.pack(expand=True, fill="both")
    tk.Label(frame, text="Tijdplanner Bootstrapper", bg="#1c212b", fg="#e6edf7", font=("Segoe UI Semibold", 16)).pack(pady=(2, 2))
    tk.Label(
        frame,
        text="Controle op Python + vereiste imports. Ontbrekende modules worden automatisch geinstalleerd.",
        bg="#1c212b",
        fg="#9fb0c5",
        font=("Segoe UI", 10),
    ).pack(pady=(0, 6))

    status_label = tk.Label(frame, text="Start controles...", bg="#1c212b", fg="#dce4f0", font=("Segoe UI", 10))
    status_label.pack(pady=(0, 6))

    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except Exception:
        pass
    style.configure(
        "TP.Horizontal.TProgressbar",
        troughcolor="#222b38",
        background="#355071",
        bordercolor="#3b4759",
        lightcolor="#4e7fba",
        darkcolor="#2f4f75",
    )
    progress = ttk.Progressbar(frame, style="TP.Horizontal.TProgressbar", mode="determinate", length=720, maximum=100)
    progress.pack(pady=5)

    check_frame = tk.Frame(frame, bg="#263141", highlightbackground="#3b4759", highlightthickness=1, bd=0)
    check_frame.pack(fill="x", pady=(8, 8))
    tk.Label(check_frame, text="Checklist", bg="#263141", fg="#dce4f0", font=("Segoe UI Semibold", 10)).pack(anchor="w", padx=8, pady=(6, 2))

    check_labels: dict[str, tk.Label] = {}
    for import_name, _pip_pkg, label in check_rows:
        if import_name == "__python__":
            line_txt = "Python runtime (executable + versie)"
        else:
            line_txt = f"{import_name} ({label})"
        lbl = tk.Label(
            check_frame,
            text=f"[ ] {line_txt}",
            anchor="w",
            bg="#263141",
            fg="#9fb0c5",
            font=("Consolas", 10),
            padx=8,
        )
        lbl.pack(fill="x", pady=1)
        check_labels[import_name] = lbl

    list_frame = tk.Frame(frame, bg="#1c212b")
    list_frame.pack(expand=True, fill="both", pady=(0, 8))
    scrollbar = tk.Scrollbar(list_frame, bg="#222b38", troughcolor="#1a2230", activebackground="#355071")
    scrollbar.pack(side="right", fill="y")
    listbox = tk.Listbox(
        list_frame,
        font=("Consolas", 10),
        yscrollcommand=scrollbar.set,
        bg="#1f2835",
        fg="#dce4f0",
        selectbackground="#355071",
        selectforeground="#ffffff",
        borderwidth=1,
        highlightthickness=1,
        highlightbackground="#3b4759",
    )
    listbox.pack(expand=True, fill="both")
    scrollbar.config(command=listbox.yview)
    root.update()

    def show_manual_close():
        nonlocal close_btn
        if close_btn is not None:
            return
        close_btn = tk.Button(
            frame,
            text="Sluiten",
            bg="#355071",
            fg="#ffffff",
            activebackground="#4e7fba",
            activeforeground="#ffffff",
            relief="flat",
            padx=14,
            pady=6,
            command=root.destroy,
        )
        close_btn.pack(pady=(0, 4))

    def hold_on_error(status_text: str) -> bool:
        status_label.config(text=status_text, fg="#ff9492")
        progress["value"] = 100
        root.update()
        show_manual_close()
        root.mainloop()
        return False

    def set_check_state(key: str, state: str, suffix: str = ""):
        lbl = check_labels.get(key)
        if not lbl:
            return
        if state == "running":
            prefix = "[~]"
            color = "#9ecbff"
        elif state == "ok":
            prefix = "[V]"
            color = "#7ad18b"
        elif state == "fail":
            prefix = "[X]"
            color = "#ff9492"
        else:
            prefix = "[ ]"
            color = "#9fb0c5"
        text = lbl.cget("text")
        base = text[text.find("]") + 2 :] if "]" in text else text
        if suffix:
            base = f"{base.split(' - ')[0]} - {suffix}"
        lbl.config(text=f"{prefix} {base}", fg=color)

    def log_line(text: str, color: str = "#dce4f0"):
        listbox.insert(tk.END, text)
        listbox.itemconfig(tk.END, foreground=color)
        listbox.yview_moveto(1.0)
        root.update()

    total = max(1, len(check_rows))
    done = 0

    set_check_state("__python__", "running")
    status_label.config(text="Check: Python runtime")
    root.update()
    py_ok = bool(sys.executable) and os.path.exists(sys.executable)
    if py_ok:
        py_ver = f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"
        set_check_state("__python__", "ok", f"Python {py_ver}")
        log_line(f"[V] Python runtime aanwezig ({py_ver})", "#7ad18b")
    else:
        set_check_state("__python__", "fail", "niet gevonden")
        log_line("[X] Python runtime ontbreekt", "#ff9492")
        return hold_on_error("Fout: Python niet gevonden")
    done += 1
    progress["value"] = (done / total) * 100
    root.update()

    if not dry_run:
        status_label.config(text="Check: pip")
        root.update()
        if not _ensure_pip_available(log_line):
            log_line("[!] Tip: controleer of Python volledig is geinstalleerd en of ensurepip beschikbaar is.", "#f3b37a")
            return hold_on_error("Fout: pip niet beschikbaar")
        _pip_upgrade_healthcheck(log_line)

    for i, (import_name, pip_pkg, label) in enumerate(checks, start=1):
        set_check_state(import_name, "running")
        status_label.config(text=f"Check: {import_name} ({label})")
        root.update()
        try:
            if simulate_missing:
                raise ImportError("simulated missing")
            importlib.import_module(import_name)
            set_check_state(import_name, "ok", "aanwezig")
            log_line(f"[V] {import_name} aanwezig", "#7ad18b")
            time.sleep(0.04)
            done += 1
            progress["value"] = (done / total) * 100
            root.update()
            continue
        except Exception:
            set_check_state(import_name, "fail", "ontbreekt")
            log_line(f"[X] {import_name} ontbreekt -> installeren ({pip_pkg})", "#ff9492")
            root.update()

        if dry_run:
            time.sleep(0.1)
            set_check_state(import_name, "ok", "dry-run")
            done += 1
            progress["value"] = (done / total) * 100
            log_line(f"[V] {import_name} dry-run install (gesimuleerd)", "#7ad18b")
            root.update()
            continue

        rc, pip_recent_lines = _pip_install_with_retry(import_name, pip_pkg, log_line)
        if rc != 0:
            hint = _bootstrap_failure_hint(import_name, pip_pkg, pip_recent_lines)
            if _is_optional(import_name):
                set_check_state(import_name, "fail", "optioneel: installatie fout")
                optional_install_failed.append(import_name)
                log_line(
                    f"[!] {import_name} kon niet worden geinstalleerd; app start zonder deze optionele detectie.",
                    "#f3b37a",
                )
                log_line(f"[!] {hint}", "#f3b37a")
                done += 1
                progress["value"] = (done / total) * 100
                root.update()
                continue
            set_check_state(import_name, "fail", "installatie fout")
            log_line(f"[X] Installatie van {import_name} is mislukt. Bekijk de logregels hierboven.", "#ff9492")
            log_line(f"[!] {hint}", "#f3b37a")
            return hold_on_error(f"Fout bij {import_name}")
        set_check_state(import_name, "ok", "geinstalleerd")
        done += 1
        progress["value"] = (done / total) * 100
        log_line(f"[V] {import_name} geinstalleerd", "#7ad18b")
        root.update()
        time.sleep(0.08)

    if optional_install_failed:
        status_label.config(
            text=f"Gereed met waarschuwing: optioneel mislukt ({', '.join(optional_install_failed)})",
            fg="#f3b37a",
        )
    else:
        status_label.config(text="Alle modules klaar")
    progress["value"] = 100
    root.update()
    if wait_for_close:
        show_manual_close()
        status_label.config(text="Preview klaar - sluit het venster handmatig")
        root.update()
        root.mainloop()
    else:
        time.sleep(0.3)
        root.destroy()
    return True


def _bootstrap_and_restart_if_needed():
    # We doen de dependency-check vroeg in de opstartflow zodat alle imports die later gebruikt worden
    # daadwerkelijk beschikbaar zijn. Als we hier pas halverwege de app zouden falen, krijg je
    # lastige runtime-errors op willekeurige plekken in plaats van een gecontroleerde startup-fout.
    checks = list(BOOTSTRAP_CHECKS)
    if "--bootstrap-demo" in sys.argv:
        ok = run_bootstrap(checks, dry_run=True)
        if not ok:
            raise SystemExit(1)
        raise SystemExit(0)
    if "--bootstrap-preview" in sys.argv:
        ok = run_bootstrap(checks, dry_run=True, simulate_missing=True, wait_for_close=True)
        if not ok:
            raise SystemExit(1)
        raise SystemExit(0)
    missing_required: list[tuple[str, str, str]] = []
    for import_name, _pip_pkg, _label in checks:
        try:
            importlib.import_module(import_name)
        except Exception:
            if import_name in OPTIONAL_BOOTSTRAP_IMPORTS:
                continue
            missing_required.append((import_name, _pip_pkg, _label))
    if not missing_required:
        return
    if os.environ.get("TP_BOOTSTRAP_DONE") == "1":
        print("Ontbrekende modules na bootstrap:", ", ".join(m[0] for m in missing_required))
        raise SystemExit(1)
    # Start bootstrap alleen voor écht ontbrekende vereiste modules.
    # Dit maakt startup sneller in omgevingen waar vrijwel alles al aanwezig is.
    ok = run_bootstrap(missing_required, dry_run=False)
    if not ok:
        print("Bootstrap mislukt. Installeer handmatig:", ", ".join(m[0] for m in missing_required))
        raise SystemExit(1)
    # Na succesvolle install herstarten we het proces expres hard met hetzelfde script + args.
    # Zo laden we direct de nieuw geïnstalleerde modules in een schone interpreter-state.
    os.environ["TP_BOOTSTRAP_DONE"] = "1"
    os.execv(sys.executable, [sys.executable, os.path.abspath(__file__), *sys.argv[1:]])


_bootstrap_and_restart_if_needed()


# ============================================================================
# BASISCONSTANTEN EN PURE HELPERS
# In dit deel staan parse/format helpers en vaste waarden die op meerdere
# plekken gebruikt worden. Door dit centraal te houden voorkom je afwijkende
# interpretaties van tijd, data en labels tussen schermen.
# ============================================================================

WEEKDAYS = ["MA", "DI", "WO", "DO", "VR", "ZA", "ZO"]
WEEKDAY_NAMES_NL = ["maandag", "dinsdag", "woensdag", "donderdag", "vrijdag", "zaterdag", "zondag"]
MONTHS_NL = [
    "Januari",
    "Februari",
    "Maart",
    "April",
    "Mei",
    "Juni",
    "Juli",
    "Augustus",
    "September",
    "Oktober",
    "November",
    "December",
]

DEFAULT_EXTRA_INFO_OPTIONS = [
    "cursus",
    "bijzonder verlof",
    "zwangerschap",
    "ziekte",
    "tijd voor tijd",
]

COLOR_DEFAULTS = {
    "bg_today_dark": "#1f2a3a",
    "bg_vacation_dark": "#7a6a44",
    "bg_holiday_dark": "#630000",
    "bg_school_dark": "#4F78A7",
    "bg_weekend_dark": "#232a34",
    "bg_default_dark": "#161b22",
    "daynum_today_dark": "#79c0ff",
    "daynum_holiday_dark": "#ff9492",
    "daynum_vacation_dark": "#ffb757",
    "daynum_school_dark": "#dbeafe",
    "daynum_weekend_dark": "#8b949e",
    "daynum_default_dark": "#e6edf3",
    "weeknum_bg_dark": "#334155",
    "weeknum_fg_dark": "#c9d1d9",
    "weektotal_bg_dark": "#223047",
    "weektotal_fg_dark": "#9ecbff",
    "empty_bg_dark": "#1b2230",
    "hours_fg_dark": "#c9d1d9",
    "planned_work_bg_dark": "#3f5f87",
    "planned_free_bg_dark": "#7a6a44",
    "timer_bg_dark": "#1f2835",
    "timer_text_dark": "#e6edf7",
    "timer_btn_dark": "#355071",
}

PLANNED_COLOR_DEFAULTS = dict(COLOR_DEFAULTS)
WORKED_COLOR_DEFAULTS = dict(PLANNED_COLOR_DEFAULTS)

GLASS_OPACITY_MIN = 0.15
GLASS_OPACITY_MAX = 0.95

APP_VERSION = "1.7.4"
APP_AUTHOR = "Rvwan"
APP_LAST_UPDATE = "2026-02-24"
_APP_INSTANCE_LOCK = None


def set_windows_app_user_model_id(app_id: str = "Rvwan.TijdplannerPro") -> None:
    """Zet expliciet AppUserModelID zodat taskbar-icoon correct wordt getoond op Windows."""
    if sys.platform != "win32":
        return
    try:
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(str(app_id))
    except Exception:
        pass


def acquire_single_instance_lock() -> bool:
    """Forceer 1 actieve instantie van de app."""
    global _APP_INSTANCE_LOCK
    lock_path = os.path.join(tempfile.gettempdir(), "tijdplanner_pro.instance.lock")
    lock = QLockFile(lock_path)
    lock.setStaleLockTime(120 * 1000)
    if not lock.tryLock(0):
        return False
    _APP_INSTANCE_LOCK = lock
    return True


def _terminate_process_by_pid(pid: int) -> bool:
    if not pid or pid <= 0:
        return False
    if pid == os.getpid():
        return False
    try:
        if sys.platform == "win32":
            os.kill(int(pid), signal.SIGTERM)
        else:
            os.kill(int(pid), signal.SIGTERM)
        return True
    except Exception:
        return False


def try_takeover_existing_instance() -> bool:
    """Probeer bestaande lockhouder te sluiten en lock over te nemen."""
    lock_path = os.path.join(tempfile.gettempdir(), "tijdplanner_pro.instance.lock")
    lock = QLockFile(lock_path)
    lock.setStaleLockTime(120 * 1000)
    pid, _host, _app = lock.getLockInfo()
    if not _terminate_process_by_pid(int(pid or 0)):
        return False
    for _ in range(25):
        if lock.tryLock(100):
            global _APP_INSTANCE_LOCK
            _APP_INSTANCE_LOCK = lock
            return True
        time.sleep(0.06)
    return False


def build_status_icon(running: bool = True, work_seconds: int = 0, idle_seconds: int = 0) -> QIcon:
    """Genereert één consistente app-/tray-icoonstijl."""
    pm = QPixmap(64, 64)
    pm.fill(Qt.transparent)
    p = QPainter(pm)
    p.setRenderHint(QPainter.Antialiasing, True)

    grad = QLinearGradient(0, 0, 64, 64)
    if running:
        grad.setColorAt(0.0, QColor("#1ea672"))
        grad.setColorAt(1.0, QColor("#2563eb"))
    else:
        grad.setColorAt(0.0, QColor("#6b7280"))
        grad.setColorAt(1.0, QColor("#374151"))
    p.setBrush(grad)
    p.setPen(QPen(QColor("#dff7ff"), 2))
    p.drawRoundedRect(6, 6, 52, 52, 14, 14)

    p.setBrush(Qt.NoBrush)
    ring_color = QColor("#facc15") if idle_seconds > 0 and running else QColor("#ecfeff")
    p.setPen(QPen(ring_color, 3))
    p.drawEllipse(12, 12, 40, 40)

    sec = int(work_seconds) % 60
    arc_span = int((sec / 60.0) * 360 * 16)
    p.setPen(QPen(QColor("#ffffff"), 3))
    p.drawArc(12, 12, 40, 40, 90 * 16, -arc_span)

    mins = max(0, int(work_seconds) // 60)
    if mins < 60:
        txt = f"{mins:02d}"
    else:
        hours = mins // 60
        txt = f"{hours}h" if hours < 100 else "99h+"
    p.setPen(QPen(QColor("#f8fafc"), 1))
    p.setFont(QFont("Segoe UI", 12 if mins >= 60 else 14, QFont.Bold))
    p.drawText(pm.rect(), Qt.AlignCenter, txt)

    p.setBrush(QColor("#f8fafc"))
    p.setPen(QPen(QColor("#f8fafc"), 1))
    if running:
        tri = QPolygon([QPoint(45, 50), QPoint(45, 58), QPoint(52, 54)])
        p.drawPolygon(tri)
    else:
        p.drawRect(45, 50, 2, 8)
        p.drawRect(50, 50, 2, 8)
    p.end()
    return QIcon(pm)


def normalize_hhmm(value: str | None, default: str = "00:00") -> str:
    if value is None:
        return default
    s = str(value).strip()
    if not s:
        return default
    if ":" in s:
        parts = s.split(":")
        if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
            h = min(23, max(0, int(parts[0])))
            m = min(59, max(0, int(parts[1])))
            return f"{h:02d}:{m:02d}"
    digits = "".join(ch for ch in s if ch.isdigit())
    if not digits:
        return default
    if len(digits) <= 2:
        h = 0
        m = int(digits)
    else:
        h = int(digits[:-2])
        m = int(digits[-2:])
    h = min(23, max(0, h))
    m = min(59, max(0, m))
    return f"{h:02d}:{m:02d}"


def parse_hhmm_strict(value: str | None) -> bool:
    if value is None:
        return False
    s = str(value).strip()
    if len(s) != 5 or s[2] != ":":
        return False
    hh, mm = s.split(":")
    if not (hh.isdigit() and mm.isdigit()):
        return False
    h, m = int(hh), int(mm)
    return 0 <= h <= 23 and 0 <= m <= 59


def hhmm_to_minutes(value: str | None) -> int:
    if not value:
        return 0
    hhmm = normalize_hhmm(value, "00:00")
    hh, mm = hhmm.split(":")
    return int(hh) * 60 + int(mm)


def minutes_to_hhmm(minutes: int) -> str:
    m = max(0, int(minutes))
    return f"{m // 60:02d}:{m % 60:02d}"


def seconds_to_hhmmss(seconds: int) -> str:
    s = max(0, int(seconds))
    hh = s // 3600
    mm = (s % 3600) // 60
    ss = s % 60
    return f"{hh:02d}:{mm:02d}:{ss:02d}"


def parse_duration_hhmm(value: str | None, default: int = 0, allow_signed: bool = False) -> int:
    if value is None:
        return int(default)
    s = str(value).strip().replace(".", ":")
    if not s:
        return int(default)
    sign = 1
    if allow_signed and s[0] in "+-":
        if s[0] == "-":
            sign = -1
        s = s[1:].strip()
    if ":" not in s:
        if s.isdigit():
            return sign * (int(s) * 60)
        return int(default)
    h_txt, m_txt = s.split(":", 1)
    if not (h_txt.isdigit() and m_txt.isdigit()):
        return int(default)
    mins = int(m_txt)
    if mins < 0 or mins > 59:
        return int(default)
    return sign * (int(h_txt) * 60 + mins)


def format_duration_hhmm(minutes: int, signed: bool = False) -> str:
    m = int(minutes or 0)
    sign = ""
    if signed and m < 0:
        sign = "-"
    m_abs = abs(m)
    return f"{sign}{m_abs // 60}:{m_abs % 60:02d}"


def format_hours_int(minutes: int, signed: bool = False) -> str:
    hrs = int(round((minutes or 0) / 60.0))
    if signed:
        return f"{hrs:+d}"
    return str(hrs)


def hhmmss_to_seconds(value) -> int:
    if value is None:
        return 0
    if isinstance(value, int):
        return max(0, value)
    txt = str(value).strip().replace(".", ":")
    if not txt:
        return 0
    parts = txt.split(":")
    try:
        parts_i = [int(p) for p in parts]
    except ValueError:
        return 0
    if len(parts_i) == 3:
        h, m, s = parts_i
    elif len(parts_i) == 2:
        h, m = parts_i
        s = 0
    else:
        return 0
    return max(0, h * 3600 + m * 60 + s)


def normalize_to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass
    return None


def daterange(start_date: date, end_date: date):
    for n in range((end_date - start_date).days + 1):
        yield start_date + timedelta(n)


def save_workbook_atomic(workbook, path: str, keep_backup: bool = True):
    """Sla workbook atomisch op om corruptie bij mislukte write te beperken.

    Eerst wordt een tijdelijk bestand geschreven in dezelfde map en daarna pas
    het doelbestand vervangen. Dat voorkomt dat een halve write het originele
    bestand direct beschadigt.
    """
    dir_path = os.path.abspath(os.path.dirname(path) or ".")
    base_name = os.path.splitext(os.path.basename(path))[0] or "workbook"
    tmp_path = ""
    bak_path = f"{path}.bak"
    # Rotating snapshots beperken schade bij late detectie van bestandscorruptie.
    # We schrijven deze snapshots bewust beperkt in frequentie om save-latency laag te houden.
    snapshots_dir = os.path.join(os.path.dirname(path), "_backups")
    snapshot_keep = 12
    snapshot_interval_sec = 30 * 60  # 30 min
    if not hasattr(save_workbook_atomic, "_last_snapshot_ts"):
        save_workbook_atomic._last_snapshot_ts = {}

    try:
        # OneDrive/nieuwe werkplekken kunnen een pad doorgeven waarvan de map
        # nog niet fysiek bestaat; borg dat eerst om FileNotFound te voorkomen.
        os.makedirs(dir_path, exist_ok=True)
        # Probeer eerst een temp-bestand in de doelmap (atomisch replace-pad).
        # Sommige enterprise OneDrive-profielen blokkeren dit onverwacht met
        # FileNotFound; dan vallen we terug op %TEMP% en doen we daarna move/replace.
        try:
            fd, created_tmp = tempfile.mkstemp(prefix=f"{base_name}_", suffix=".tmp.xlsx", dir=dir_path)
            os.close(fd)
            tmp_path = created_tmp
        except FileNotFoundError:
            fd, created_tmp = tempfile.mkstemp(prefix=f"{base_name}_", suffix=".tmp.xlsx", dir=tempfile.gettempdir())
            os.close(fd)
            tmp_path = created_tmp
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        workbook.save(tmp_path)
        if keep_backup and os.path.exists(path):
            try:
                shutil.copy2(path, bak_path)
            except Exception:
                pass
            try:
                now_ts = time.time()
                key = os.path.abspath(path).casefold()
                last_ts = float(save_workbook_atomic._last_snapshot_ts.get(key, 0.0))
                if (now_ts - last_ts) >= snapshot_interval_sec:
                    os.makedirs(snapshots_dir, exist_ok=True)
                    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    base = os.path.splitext(os.path.basename(path))[0]
                    snap_path = os.path.join(snapshots_dir, f"{base}_{stamp}.xlsx")
                    shutil.copy2(path, snap_path)
                    save_workbook_atomic._last_snapshot_ts[key] = now_ts
                    snaps = sorted(
                        [p for p in os.listdir(snapshots_dir) if p.lower().startswith(base.lower() + "_") and p.lower().endswith(".xlsx")]
                    )
                    while len(snaps) > snapshot_keep:
                        old = snaps.pop(0)
                        try:
                            os.remove(os.path.join(snapshots_dir, old))
                        except Exception:
                            break
            except Exception:
                pass
        try:
            os.replace(tmp_path, path)
        except OSError:
            # Fallback voor omgevingen waar cross-dir replace policy-technisch
            # faalt: forceer via move.
            shutil.move(tmp_path, path)
    except PermissionError as exc:
        raise RuntimeError(
            "Opslaan mislukt: het Excel-bestand lijkt in gebruik. Sluit het bestand en probeer opnieuw."
        ) from exc
    except OSError as exc:
        msg = (
            f"Opslaan mislukt door bestandsfout: {exc}\n"
            f"Pad: {path}\n"
            f"Map bestaat: {os.path.isdir(dir_path)}"
        )
        raise RuntimeError(msg) from exc
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


@dataclass
class DayData:
    w: str = "00:00"
    v: str = "00:00"
    z: str = "00:00"
    worked: str = ""


class StorageBackend(Protocol):
    """Kleine contractlaag tussen UI en opslagimplementatie.

    Nu is ExcelStore de concrete backend. Later kan een andere backend deze
    interface implementeren zonder de UI-code fundamenteel te herschrijven.
    """

    year: int
    wb: object
    path: str

    def safe_save(self) -> None:
        ...

    def get_day(self, dt: date) -> DayData:
        ...

    def set_day(self, dt: date, data: DayData, reason: str = "", extra_info: str = ""):
        ...

    def get_timer_log(self, dt: date) -> dict[str, int]:
        ...

    def save_timer_log(self, dt: date, work_s: int, idle_s: int, call_s: int):
        ...

    def get_day_limit(self, dt: date) -> int:
        ...

    def day_reason(self, dt: date) -> str:
        ...

    def get_extra_info(self, dt: date) -> str:
        ...

    def get_saldo_text(self) -> str:
        ...


def split_reason_and_type(raw_text: str) -> tuple[str, str]:
    def _looks_like_free_prefix(value: str) -> bool:
        s = value.strip().lower()
        if s == "vrij":
            return True
        if s.endswith(" vrij") and len(s) >= 10 and s[2] == ":" and s[:2].isdigit() and s[3:5].isdigit():
            return True
        return False

    if not raw_text:
        return "", "Vrij"
    txt = str(raw_text).strip()
    if " - " in txt:
        left, right = txt.split(" - ", 1)
        if _looks_like_free_prefix(left):
            return "", right.strip()
        kind = left.split("(", 1)[0].strip() or "Vrij"
        return kind, right.strip()
    if _looks_like_free_prefix(txt):
        return "", ""
    kind = txt.split("(", 1)[0].strip() or "Vrij"
    return kind, ""


def parse_nl_date(value: str) -> date | None:
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def normalize_extra_info_options(options: list[str] | None) -> list[str]:
    out: list[str] = []
    seen = set()

    for d in DEFAULT_EXTRA_INFO_OPTIONS:
        k = d.casefold()
        if k not in seen:
            seen.add(k)
            out.append(d)

    for raw in options or []:
        v = str(raw).strip()
        if not v:
            continue
        if v.casefold() == "optioneel":
            continue
        k = v.casefold()
        if k in seen:
            continue
        seen.add(k)
        out.append(v)

    return out


def normalize_extra_info_enabled(options: list[str] | None, enabled_options: list[str] | None) -> list[str]:
    opts = normalize_extra_info_options(options)
    enabled = {str(v).strip().casefold() for v in (enabled_options or []) if str(v).strip()}
    if not enabled:
        return opts
    return [opt for opt in opts if opt.casefold() in enabled]


def default_extra_info_color(option: str) -> str:
    preset = {
        "cursus": "#5b6f8f",
        "bijzonder verlof": "#7b5f3e",
        "zwangerschap": "#7a4f79",
        "ziekte": "#8b3f3f",
        "tijd voor tijd": "#2f7d4a",
    }
    k = option.strip().casefold()
    if k in preset:
        return preset[k]
    h = hashlib.md5(k.encode("utf-8")).hexdigest()
    r = 72 + (int(h[0:2], 16) % 92)
    g = 72 + (int(h[2:4], 16) % 92)
    b = 72 + (int(h[4:6], 16) % 92)
    return f"#{r:02x}{g:02x}{b:02x}"


# ============================================================================
# UI INPUT PRIMITIVES
# Deze sectie bevat kleine bouwstenen voor invoer en toggles.
# Het doel is om invoerfouten vroeg te blokkeren zodat latere dialoglogica
# uit kan gaan van valide waarden en minder defensieve code nodig heeft.
# ============================================================================
class HhmmEntryFilter(QObject):
    """Dwingt hh:mm gedrag af op een QLineEdit.

    We filteren toets-events direct zodat de gebruiker alleen geldige karakters
    op de juiste positie kan typen. Daardoor hoeven formulieren achteraf
    minder te repareren en blijft tijdsinvoer overal consistent.
    """
    DIGIT_POS = (0, 1, 3, 4)

    def __init__(self, line_edit: QLineEdit, default: str = "00:00"):
        super().__init__(line_edit)
        self.line_edit = line_edit
        self.default = default

    def _ensure_mask(self):
        cur = self.line_edit.text()
        if len(cur) != 5 or (len(cur) > 2 and cur[2] != ":"):
            self.line_edit.setText(normalize_hhmm(cur if cur else self.default, self.default))

    def _set_char(self, idx: int, ch: str):
        s = self.line_edit.text()
        if len(s) != 5:
            s = self.default
        out = s[:idx] + ch + s[idx + 1 :]
        self.line_edit.setText(out)

    def _next_digit_pos(self, pos: int) -> int:
        for p in self.DIGIT_POS:
            if p >= pos:
                return p
        return self.DIGIT_POS[0]

    def _prev_digit_pos(self, pos: int) -> int:
        for p in reversed(self.DIGIT_POS):
            if p <= pos:
                return p
        return self.DIGIT_POS[0]

    def eventFilter(self, obj, event):
        if obj is not self.line_edit:
            return False

        if event.type() == QEvent.FocusIn:
            self._ensure_mask()
            self.line_edit.setCursorPosition(0)
            return False

        if event.type() == QEvent.FocusOut:
            self.line_edit.setText(normalize_hhmm(self.line_edit.text() or self.default, self.default))
            return False

        if event.type() != QEvent.KeyPress:
            return False

        if event.modifiers() & (Qt.ControlModifier | Qt.AltModifier | Qt.MetaModifier):
            return False

        self._ensure_mask()
        key = event.key()
        txt = event.text()
        pos = self.line_edit.cursorPosition()

        if txt and txt.isdigit():
            target = self._next_digit_pos(pos if pos <= 4 else 5)
            self._set_char(target, txt)
            next_pos = self.DIGIT_POS[0]
            for p in self.DIGIT_POS:
                if p > target:
                    next_pos = p
                    break
            if target == self.DIGIT_POS[-1]:
                next_pos = self.DIGIT_POS[0]
            self.line_edit.setCursorPosition(next_pos)
            return True

        if key == Qt.Key_Backspace:
            if pos > 0:
                target = self._prev_digit_pos(pos - 1)
                self._set_char(target, "0")
                self.line_edit.setCursorPosition(target)
            return True

        if key == Qt.Key_Delete:
            target = self._next_digit_pos(pos)
            self._set_char(target, "0")
            self.line_edit.setCursorPosition(target)
            return True

        if key == Qt.Key_Left:
            target = self._prev_digit_pos(max(0, pos - 1))
            self.line_edit.setCursorPosition(target)
            return True

        if key == Qt.Key_Right:
            target = self._next_digit_pos(min(4, pos + 1))
            self.line_edit.setCursorPosition(target)
            return True

        if key == Qt.Key_Home:
            self.line_edit.setCursorPosition(0)
            return True

        if key == Qt.Key_End:
            self.line_edit.setCursorPosition(4)
            return True

        if key in (Qt.Key_Tab, Qt.Key_Backtab, Qt.Key_Up, Qt.Key_Down, Qt.Key_Return, Qt.Key_Enter):
            return False

        if txt == ":":
            return True

        if txt and txt.isprintable():
            return True

        return False


def force_hhmm_line_edit(line_edit: QLineEdit, default: str = "00:00"):
    line_edit.setText(normalize_hhmm(line_edit.text() or default, default))
    filt = HhmmEntryFilter(line_edit, default=default)
    line_edit.installEventFilter(filt)
    line_edit._hhmm_filter = filt


class ModeSwitch(QCheckBox):
    """Visuele mode-switch voor Uren versus Planning.

    De widget animeert de knoppositie en blendt de track-kleur zodat de
    context direct zichtbaar is. Dit is bewust custom getekend omdat standaard
    Qt-checkbox styling te weinig controle gaf over herkenbaarheid.
    """
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.setObjectName("modeSwitch")
        self.setCursor(Qt.PointingHandCursor)
        self.setText("")
        self.setFixedSize(58, 28)
        self._offset = 0.0
        self._anim = QPropertyAnimation(self, b"offset", self)
        self._anim.setDuration(130)
        self._anim.setEasingCurve(QEasingCurve.OutCubic)
        self.toggled.connect(self._on_toggled)

    def _on_toggled(self, checked: bool):
        self._anim.stop()
        self._anim.setStartValue(self._offset)
        self._anim.setEndValue(1.0 if checked else 0.0)
        self._anim.start()

    def set_visual_checked(self, checked: bool):
        self._anim.stop()
        super().setChecked(bool(checked))
        self._offset = 1.0 if checked else 0.0
        self.update()

    def get_offset(self) -> float:
        return float(self._offset)

    def set_offset(self, value: float):
        try:
            v = float(value)
        except Exception:
            v = 0.0
        self._offset = max(0.0, min(1.0, v))
        self.update()

    offset = Property(float, get_offset, set_offset)

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        r = self.rect().adjusted(1, 1, -1, -1)
        k = self._offset

        # Track: green (uren) -> blue (planning).
        tr = int(58 + ((71 - 58) * k))
        tg = int(111 + ((103 - 111) * k))
        tb = int(82 + ((138 - 82) * k))
        track = QColor(tr, tg, tb)

        g = QLinearGradient(r.left(), r.top(), r.left(), r.bottom())
        g.setColorAt(0.0, track.lighter(118))
        g.setColorAt(1.0, track.darker(110))
        p.setBrush(g)
        p.setPen(QPen(QColor("#2a394a"), 1))
        p.drawRoundedRect(r, 13, 13)

        d = r.height() - 6
        x = r.left() + 3 + int((r.width() - d - 6) * k)
        y = r.top() + 3
        knob = QRect(x, y, d, d)

        kg = QLinearGradient(knob.left(), knob.top(), knob.left(), knob.bottom())
        kg.setColorAt(0.0, QColor("#f8fcff"))
        kg.setColorAt(1.0, QColor("#d4e2f0"))
        p.setBrush(kg)
        p.setPen(QPen(QColor("#8aa0b6"), 1))
        p.drawEllipse(knob)

        inner = knob.adjusted(3, 3, -3, -3)
        p.setBrush(QColor(255, 255, 255, 52))
        p.setPen(Qt.NoPen)
        p.drawEllipse(inner)


class FramelessDialog(QDialog):
    """Basis voor frameless dialogs met drag-gedrag.

    Alle popups die dezelfde look-and-feel delen erven hiervan.
    Zo hoeven we vensterdragging niet in elke dialog afzonderlijk te dupliceren.
    """
    INACTIVE_OPACITY = 0.84

    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowFlag(Qt.FramelessWindowHint, True)
        self._drag_active = False
        self._drag_offset = QPoint()
        app = QApplication.instance()
        if app is not None and hasattr(app, "applicationStateChanged"):
            app.applicationStateChanged.connect(self._on_app_state_changed)
        self._apply_focus_opacity()

    def _on_app_state_changed(self, _state):
        self._apply_focus_opacity()

    def _apply_focus_opacity(self):
        app = QApplication.instance()
        is_active = bool(app is not None and app.applicationState() == Qt.ApplicationActive)
        self.setWindowOpacity(1.0 if is_active else self.INACTIVE_OPACITY)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_active = True
            self._drag_offset = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._drag_active:
            self.move(event.globalPosition().toPoint() - self._drag_offset)
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_active = False
        super().mouseReleaseEvent(event)

    def showEvent(self, event):
        super().showEvent(event)
        self._apply_focus_opacity()

    def focusInEvent(self, event):
        super().focusInEvent(event)
        self._apply_focus_opacity()

    def focusOutEvent(self, event):
        super().focusOutEvent(event)
        self._apply_focus_opacity()


class CalendarCellDelegate(QStyledItemDelegate):
    def __init__(self, owner):
        super().__init__(owner.table)
        self.owner = owner

    def paint(self, painter, option, index):
        super().paint(painter, option, index)
        if bool(index.data(Qt.UserRole)):
            painter.save()
            painter.setPen(QPen(QColor("#2e3948"), 1))
            painter.drawRect(option.rect.adjusted(0, 0, -1, -1))
            painter.restore()


# ============================================================================
# STORAGE LAYER (EXCEL)
# Deze klasse is de enige plek waar direct met openpyxl wordt gewerkt.
# Door IO en migraties hier te centreren, blijft de UI-laag eenvoudiger en
# kunnen toekomstige opslagmigraties (bijv. SQLite) gecontroleerd gebeuren.
# ============================================================================
class ExcelStore(StorageBackend):
    """Persistente opslaglaag (Excel).

    Alle workbook-migraties, normalisatie en IO lopen via deze klasse.
    UI mag hierop lezen/schrijven, maar niet direct aan openpyxl-sheets zitten.
    """
    def __init__(self, year: int, base_dir: str):
        self.year = year
        self.base_dir = os.path.abspath(base_dir)
        os.makedirs(self.base_dir, exist_ok=True)
        self.path = os.path.join(self.base_dir, f"Time_tabel_{year}.xlsx")
        self.nl_holidays = holidays.NL(years=[year])
        self.planned_data: dict[str, dict[str, str]] = {}
        self.worked_data: dict[str, str] = {}
        self.vakantie_dagen: dict[str, str] = {}
        self.school_vakanties: dict[str, str] = {}
        self.extra_info_data: dict[str, str] = {}
        self.weekday_pattern: dict[int, dict[str, str]] = {}
        self.extra_info_options: list[str] = list(DEFAULT_EXTRA_INFO_OPTIONS)
        self.extra_info_enabled: list[str] = list(DEFAULT_EXTRA_INFO_OPTIONS)
        self.data_log: dict[str, dict[str, int]] = {}
        self.day_max_minutes = 480
        self.employment_pct = 100.0
        self.fulltime_week_minutes = parse_duration_hhmm("38:00", default=2280)
        self.contract_week_minutes = parse_duration_hhmm("38:00", default=2280)
        self.contract_year_minutes = parse_duration_hhmm("1938:00", default=116280)
        self.vacation_stat_minutes = parse_duration_hhmm("152:00", default=9120)
        self.vacation_extra_minutes = parse_duration_hhmm("56:00", default=3360)
        self.carry_hours_prev_minutes = 0
        self.carry_vac_prev_minutes = 0
        self.workdays_mask = [True, True, True, True, True, False, False]
        self.school_region = "zuid"
        self._load_or_init()
        self.load_all()

    def _load_or_init(self):
        created_new = False
        if not os.path.exists(self.path):
            created_new = True
            wb = Workbook()
            wb.remove(wb.active)
            for month_name in MONTHS_NL:
                ws = wb.create_sheet(month_name)
                ws.append(["WK"] + WEEKDAYS + ["Wk totaal"])
            ws_vrij = wb.create_sheet("vrije dagen")
            ws_vrij.append(["datum", "type"])
            ws_plan = wb.create_sheet("Planning")
            ws_plan.append(["datum", "W"])
            ws_sv = wb.create_sheet("Schoolvakanties")
            ws_sv.append(["datum", "vakantie"])
            ws_wp = wb.create_sheet("Werkpatroon")
            ws_wp.append(["weekdag", "W", "MAX"])
            defaults = {
                "MA": ("08:00", "00:00", "00:00", "08:00"),
                "DI": ("08:00", "00:00", "00:00", "08:00"),
                "WO": ("08:00", "00:00", "00:00", "08:00"),
                "DO": ("08:00", "00:00", "00:00", "08:00"),
                "VR": ("08:00", "00:00", "00:00", "08:00"),
                "ZA": ("00:00", "00:00", "00:00", "00:00"),
                "ZO": ("00:00", "00:00", "00:00", "00:00"),
            }
            for wd, vals in defaults.items():
                ws_wp.append([wd, vals[0], vals[3]])
            ws_set = wb.create_sheet("Instellingen")
            ws_set.append(["key", "value"])
            ws_set.append(["dag_max", "08:00"])
            ws_set.append(["employment_pct", "100"])
            ws_set.append(["fulltime_week", "38:00"])
            ws_set.append(["contract_week", "38:00"])
            ws_set.append(["contract_year", "1938:00"])
            ws_set.append(["vacation_stat", "152:00"])
            ws_set.append(["vacation_extra", "56:00"])
            ws_set.append(["carry_hours_prev", "0:00"])
            ws_set.append(["carry_vac_prev", "0:00"])
            ws_set.append(["workdays_mask", "1111100"])
            ws_ai = wb.create_sheet("AanvullendeInfo")
            ws_ai.append(["optie", "actief"])
            for opt in DEFAULT_EXTRA_INFO_OPTIONS:
                ws_ai.append([opt, 1])
            ws_aid = wb.create_sheet("AanvullendeInfoData")
            ws_aid.append(["datum", "info"])
            ws_log = wb.create_sheet("Data_log")
            ws_log.append(["Datum", "Tijd", "Werk_sec", "Idle_sec", "Call_sec"])
            save_workbook_atomic(wb, self.path, keep_backup=False)
        self.wb = load_workbook(self.path)
        self.ensure_sheets()
        if created_new:
            self.seed_public_holidays_to_free_days()
            self.seed_school_holidays_from_api(region_filter=None)
            self.safe_save()

    def safe_save(self) -> None:
        save_workbook_atomic(self.wb, self.path, keep_backup=True)

    def ensure_sheets(self):
        if "Planning" not in self.wb.sheetnames:
            ws = self.wb.create_sheet("Planning")
            ws.append(["datum", "W"])
        if "vrije dagen" not in self.wb.sheetnames:
            ws = self.wb.create_sheet("vrije dagen")
            ws.append(["datum", "type"])
        if "Schoolvakanties" not in self.wb.sheetnames:
            ws = self.wb.create_sheet("Schoolvakanties")
            ws.append(["datum", "vakantie", "regio"])
        else:
            ws = self.wb["Schoolvakanties"]
            h1 = str(ws.cell(1, 1).value or "").strip().casefold()
            h2 = str(ws.cell(1, 2).value or "").strip().casefold()
            h3 = str(ws.cell(1, 3).value or "").strip().casefold()
            if h1 != "datum" or h2 != "vakantie":
                ws.cell(1, 1, "datum")
                ws.cell(1, 2, "vakantie")
            if h3 != "regio":
                ws.cell(1, 3, "regio")
        if "Werkpatroon" not in self.wb.sheetnames:
            ws = self.wb.create_sheet("Werkpatroon")
            ws.append(["weekdag", "W", "MAX"])
        if "Instellingen" not in self.wb.sheetnames:
            ws = self.wb.create_sheet("Instellingen")
            ws.append(["key", "value"])
            ws.append(["dag_max", "08:00"])
            ws.append(["employment_pct", "100"])
            ws.append(["fulltime_week", "38:00"])
            ws.append(["contract_week", "38:00"])
            ws.append(["contract_year", "1938:00"])
            ws.append(["vacation_stat", "152:00"])
            ws.append(["vacation_extra", "56:00"])
            ws.append(["carry_hours_prev", "0:00"])
            ws.append(["carry_vac_prev", "0:00"])
            ws.append(["workdays_mask", "1111100"])
        if "AanvullendeInfo" not in self.wb.sheetnames:
            ws = self.wb.create_sheet("AanvullendeInfo")
            ws.append(["optie", "actief"])
            for opt in DEFAULT_EXTRA_INFO_OPTIONS:
                ws.append([opt, 1])
        if "AanvullendeInfoData" not in self.wb.sheetnames:
            ws = self.wb.create_sheet("AanvullendeInfoData")
            ws.append(["datum", "info"])
        if "Data_log" not in self.wb.sheetnames:
            ws = self.wb.create_sheet("Data_log")
            ws.append(["Datum", "Tijd", "Werk_sec", "Idle_sec", "Call_sec"])
        for m in MONTHS_NL:
            if m not in self.wb.sheetnames:
                ws = self.wb.create_sheet(m)
                ws.append(["WK"] + WEEKDAYS + ["Wk totaal"])
        self._normalize_planning_sheet()
        self._normalize_workpattern_sheet()
        self.safe_save()

    def seed_public_holidays_to_free_days(self):
        ws = self.wb["vrije dagen"]
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                d = normalize_to_date(row[0])
                if d:
                    existing.add(d)

        for d in sorted(self.nl_holidays.keys()):
            if d.year != self.year:
                continue
            if d in existing:
                continue
            ws.append([d, str(self.nl_holidays.get(d) or "Feestdag")])
            existing.add(d)

    def fetch_school_holidays_from_api(self, region_filter: str | None = None) -> list[tuple[str, str, str]]:
        url = "https://opendata.rijksoverheid.nl/v1/infotypes/schoolholidays?output=json"
        req = urllib.request.Request(url, headers={"User-Agent": "TijdplannerPro/1.0"})
        out: dict[tuple[str, str], str] = {}

        try:
            with urllib.request.urlopen(req, timeout=12) as resp:
                payload = json.loads(resp.read().decode("utf-8"))
        except Exception:
            return []

        items = payload if isinstance(payload, list) else payload.get("result", []) if isinstance(payload, dict) else []
        year_str = str(self.year)
        rf = (region_filter or "").strip().casefold()

        for item in items:
            content_list = item.get("content", []) if isinstance(item, dict) else []
            for content in content_list:
                schoolyear = str(content.get("schoolyear", ""))
                if year_str not in schoolyear:
                    continue
                vacations = content.get("vacations", [])
                for vac in vacations:
                    vac_type = str(vac.get("type", "")).strip() or "Schoolvakantie"
                    for reg in vac.get("regions", []):
                        region_name = str(reg.get("region", "")).strip().casefold()
                        if not region_name:
                            continue
                        if rf and rf not in region_name:
                            continue
                        if "noord" in region_name:
                            region_key = "noord"
                        elif "midden" in region_name:
                            region_key = "midden"
                        elif "zuid" in region_name:
                            region_key = "zuid"
                        else:
                            continue
                        try:
                            start = date.fromisoformat(str(reg.get("startdate", ""))[:10])
                            end = date.fromisoformat(str(reg.get("enddate", ""))[:10])
                        except Exception:
                            continue
                        if end < start:
                            continue
                        cur = start
                        while cur <= end:
                            if cur.year == self.year:
                                out[(cur.strftime("%Y-%m-%d"), region_key)] = vac_type
                            cur += timedelta(days=1)

        rows = [(d, t, r) for (d, r), t in out.items()]
        rows.sort(key=lambda x: (x[0], x[2]))
        return rows

    def seed_school_holidays_from_api(self, region_filter: str | None = None) -> int:
        ws = self.wb["Schoolvakanties"]
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and row[1]:
                d = normalize_to_date(row[0])
                if d:
                    reg = str(row[2] if len(row) > 2 and row[2] else "").strip().casefold()
                    if reg not in {"noord", "midden", "zuid"}:
                        reg = ""
                    existing.add((d.strftime("%Y-%m-%d"), reg))

        school_rows = self.fetch_school_holidays_from_api(region_filter=region_filter)
        added = 0
        for d_key, vac_type, reg in school_rows:
            pair = (d_key, reg)
            if pair in existing:
                continue
            d = normalize_to_date(d_key)
            if not d:
                continue
            ws.append([d, vac_type, reg])
            existing.add(pair)
            added += 1
        return added

    def _normalize_planning_sheet(self):
        ws = self.wb["Planning"]
        rows: list[tuple[date, str]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            d = normalize_to_date(row[0])
            if not d:
                continue
            w = normalize_hhmm((row[1] if len(row) > 1 else "00:00") or "00:00")
            rows.append((d, w))
        rows.sort(key=lambda x: x[0])
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
        ws.append(["datum", "W"])
        for d, w in rows:
            ws.append([d, w])

    def _normalize_workpattern_sheet(self):
        ws = self.wb["Werkpatroon"]
        idx = {"MA": 0, "DI": 1, "WO": 2, "DO": 3, "VR": 4, "ZA": 5, "ZO": 6}
        vals: dict[str, tuple[str, str]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            wd = str(row[0]).strip().upper()
            if wd not in idx:
                continue
            w = normalize_hhmm((row[1] if len(row) > 1 else "00:00") or "00:00")
            if len(row) > 4:
                m_raw = row[4]
            else:
                m_raw = row[2] if len(row) > 2 else ("08:00" if idx[wd] < 5 else "00:00")
            m = normalize_hhmm(m_raw or ("08:00" if idx[wd] < 5 else "00:00"))
            vals[wd] = (w, m)
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
        ws.append(["weekdag", "W", "MAX"])
        for wd in ("MA", "DI", "WO", "DO", "VR", "ZA", "ZO"):
            i = idx[wd]
            default_w = "08:00" if i < 5 else "00:00"
            default_m = "08:00" if i < 5 else "00:00"
            w, m = vals.get(wd, (default_w, default_m))
            ws.append([wd, w, m])

    def load_all(self):
        self.load_patterns()
        self.load_settings()
        self.load_free_days()
        self.load_school_holidays()
        if not self.school_vakanties:
            added = self.seed_school_holidays_from_api(region_filter=None)
            if added:
                self.safe_save()
                self.load_school_holidays()
        self.load_extra_info_options()
        self.load_extra_info_data()
        self.load_data_log()
        self.load_planning()
        self.load_worked()
        self.highlight_today_excel()

    def load_extra_info_options(self):
        ws = self.wb["AanvullendeInfo"]
        raw_options = []
        raw_enabled = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            opt = str(row[0]).strip()
            if not opt:
                continue
            raw_options.append(opt)
            is_active = True
            if len(row) > 1:
                marker = str(row[1]).strip().casefold()
                if marker in {"0", "false", "nee", "no", "n"}:
                    is_active = False
            if is_active:
                raw_enabled.append(opt)

        opts = normalize_extra_info_options(raw_options)
        if not raw_options:
            self.save_extra_info_options(opts, opts)
            return

        enabled = normalize_extra_info_enabled(opts, raw_enabled)
        self.extra_info_options = opts
        self.extra_info_enabled = enabled

    def save_extra_info_options(self, options: list[str], enabled_options: list[str] | None = None):
        opts = normalize_extra_info_options(options)
        enabled = normalize_extra_info_enabled(opts, enabled_options if enabled_options is not None else opts)
        enabled_set = {o.casefold() for o in enabled}

        ws = self.wb["AanvullendeInfo"]
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
        ws.append(["optie", "actief"])
        for opt in opts:
            ws.append([opt, 1 if opt.casefold() in enabled_set else 0])

        self.extra_info_options = opts
        self.extra_info_enabled = enabled
        self.safe_save()

    def load_extra_info_data(self):
        self.extra_info_data.clear()
        ws = self.wb["AanvullendeInfoData"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and row[1]:
                d = normalize_to_date(row[0])
                if d:
                    self.extra_info_data[d.strftime("%Y-%m-%d")] = str(row[1]).strip()

    def load_data_log(self):
        self.data_log.clear()
        ws = self.wb["Data_log"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            d = normalize_to_date(row[0])
            if not d:
                continue
            key = d.strftime("%Y-%m-%d")
            self.data_log[key] = {
                "work": hhmmss_to_seconds(row[2] if len(row) > 2 else 0),
                "idle": hhmmss_to_seconds(row[3] if len(row) > 3 else 0),
                "call": hhmmss_to_seconds(row[4] if len(row) > 4 else 0),
            }

    def _save_extra_info_cell(self, dt: date, info: str):
        ws = self.wb["AanvullendeInfoData"]
        found = None
        for row in ws.iter_rows(min_row=2):
            if normalize_to_date(row[0].value) == dt:
                found = row
                break
        if found:
            found[1].value = info
        else:
            ws.append([dt, info])

    def _delete_extra_info_cell(self, dt: date):
        ws = self.wb["AanvullendeInfoData"]
        rows = []
        for row in ws.iter_rows(min_row=2):
            if normalize_to_date(row[0].value) == dt:
                rows.append(row[0].row)
        for r in sorted(rows, reverse=True):
            ws.delete_rows(r, 1)

    def get_extra_info(self, dt: date) -> str:
        return self.extra_info_data.get(dt.strftime("%Y-%m-%d"), "")

    def load_patterns(self):
        ws = self.wb["Werkpatroon"]
        idx = {"MA": 0, "DI": 1, "WO": 2, "DO": 3, "VR": 4, "ZA": 5, "ZO": 6}
        self.weekday_pattern.clear()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            wd = str(row[0]).strip().upper()
            if wd in idx:
                i = idx[wd]
                max_raw = row[4] if len(row) > 4 else (row[2] if len(row) > 2 else ("08:00" if i < 5 else "00:00"))
                self.weekday_pattern[i] = {
                    "W": normalize_hhmm((row[1] if len(row) > 1 else "00:00") or "00:00"),
                    "V": "00:00",
                    "Z": "00:00",
                    "M": normalize_hhmm(max_raw or ("08:00" if i < 5 else "00:00")),
                }
        for i in range(7):
            if i not in self.weekday_pattern:
                self.weekday_pattern[i] = {"W": "00:00", "V": "00:00", "Z": "00:00", "M": "08:00" if i < 5 else "00:00"}

    def load_settings(self):
        ws = self.wb["Instellingen"]
        found = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            key = str(row[0]).strip()
            raw = str(row[1] or "").strip()
            if key == "dag_max":
                val = normalize_hhmm(raw or "08:00", "08:00")
                self.day_max_minutes = hhmm_to_minutes(val) or 480
                found.add(key)
            elif key == "employment_pct":
                try:
                    self.employment_pct = max(0.0, min(200.0, float(raw or "100")))
                except Exception:
                    self.employment_pct = 100.0
                found.add(key)
            elif key == "fulltime_week":
                self.fulltime_week_minutes = max(0, parse_duration_hhmm(raw, default=self.fulltime_week_minutes))
                found.add(key)
            elif key == "contract_week":
                self.contract_week_minutes = max(0, parse_duration_hhmm(raw, default=self.contract_week_minutes))
                found.add(key)
            elif key == "contract_year":
                self.contract_year_minutes = max(0, parse_duration_hhmm(raw, default=self.contract_year_minutes))
                found.add(key)
            elif key == "vacation_stat":
                self.vacation_stat_minutes = max(0, parse_duration_hhmm(raw, default=self.vacation_stat_minutes))
                found.add(key)
            elif key == "vacation_extra":
                self.vacation_extra_minutes = max(0, parse_duration_hhmm(raw, default=self.vacation_extra_minutes))
                found.add(key)
            elif key == "carry_hours_prev":
                self.carry_hours_prev_minutes = parse_duration_hhmm(raw, default=0, allow_signed=True)
                found.add(key)
            elif key == "carry_vac_prev":
                self.carry_vac_prev_minutes = parse_duration_hhmm(raw, default=0, allow_signed=True)
                found.add(key)
            elif key == "workdays_mask":
                bits = "".join(ch for ch in raw if ch in "01")
                if len(bits) >= 7:
                    self.workdays_mask = [b == "1" for b in bits[:7]]
                found.add(key)
        # Migratie/defaults voor oudere bestanden.
        if "dag_max" not in found:
            self._set_setting("dag_max", minutes_to_hhmm(self.day_max_minutes or 480))
        if "employment_pct" not in found:
            self._set_setting("employment_pct", f"{self.employment_pct:.0f}")
        if "fulltime_week" not in found:
            self._set_setting("fulltime_week", format_duration_hhmm(self.fulltime_week_minutes))
        if "contract_week" not in found:
            self._set_setting("contract_week", format_duration_hhmm(self.contract_week_minutes))
        if "contract_year" not in found:
            self._set_setting("contract_year", format_duration_hhmm(self.contract_year_minutes))
        if "vacation_stat" not in found:
            self._set_setting("vacation_stat", format_duration_hhmm(self.vacation_stat_minutes))
        if "vacation_extra" not in found:
            self._set_setting("vacation_extra", format_duration_hhmm(self.vacation_extra_minutes))
        if "carry_hours_prev" not in found:
            self._set_setting("carry_hours_prev", format_duration_hhmm(self.carry_hours_prev_minutes, signed=True))
        if "carry_vac_prev" not in found:
            self._set_setting("carry_vac_prev", format_duration_hhmm(self.carry_vac_prev_minutes, signed=True))
        if "workdays_mask" not in found:
            self._set_setting("workdays_mask", "".join("1" if x else "0" for x in self.workdays_mask))
        self.safe_save()

    def _set_setting(self, key: str, value: str):
        ws = self.wb["Instellingen"]
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value or "").strip() == key:
                row[1].value = value
                return
        ws.append([key, value])

    def save_contract_budget_settings(
        self,
        employment_pct: float,
        fulltime_week: str,
        contract_year: str,
        vacation_stat: str,
        vacation_extra: str,
        carry_hours_prev: str,
        carry_vac_prev: str,
        workdays_mask: list[bool],
    ):
        self.employment_pct = max(0.0, min(200.0, float(employment_pct)))
        self.fulltime_week_minutes = max(0, parse_duration_hhmm(fulltime_week, default=self.fulltime_week_minutes))
        self.contract_year_minutes = max(0, parse_duration_hhmm(contract_year, default=self.contract_year_minutes))
        self.contract_week_minutes = int(round(self.fulltime_week_minutes * (self.employment_pct / 100.0)))
        self.vacation_stat_minutes = max(0, parse_duration_hhmm(vacation_stat, default=self.vacation_stat_minutes))
        self.vacation_extra_minutes = max(0, parse_duration_hhmm(vacation_extra, default=self.vacation_extra_minutes))
        self.carry_hours_prev_minutes = parse_duration_hhmm(carry_hours_prev, default=0, allow_signed=True)
        self.carry_vac_prev_minutes = parse_duration_hhmm(carry_vac_prev, default=0, allow_signed=True)
        if workdays_mask and len(workdays_mask) >= 7:
            self.workdays_mask = [bool(x) for x in workdays_mask[:7]]

        self._set_setting("employment_pct", f"{self.employment_pct:.0f}")
        self._set_setting("fulltime_week", format_duration_hhmm(self.fulltime_week_minutes))
        self._set_setting("contract_week", format_duration_hhmm(self.contract_week_minutes))
        self._set_setting("contract_year", format_duration_hhmm(self.contract_year_minutes))
        self._set_setting("vacation_stat", format_duration_hhmm(self.vacation_stat_minutes))
        self._set_setting("vacation_extra", format_duration_hhmm(self.vacation_extra_minutes))
        self._set_setting("carry_hours_prev", format_duration_hhmm(self.carry_hours_prev_minutes, signed=True))
        self._set_setting("carry_vac_prev", format_duration_hhmm(self.carry_vac_prev_minutes, signed=True))
        self._set_setting("workdays_mask", "".join("1" if x else "0" for x in self.workdays_mask))
        self._apply_workdays_mask_to_pattern()
        self.safe_save()

    def _apply_workdays_mask_to_pattern(self):
        # Werkdagmasker is leidend voor basisplanning: uitgevinkte dagen worden echt 00:00.
        # Dit zorgt ervoor dat niet-werkdagen niet per ongeluk uren blijven bevatten uit oude patronen.
        # Actieve dagen krijgen een evenredige baseline op basis van contract/week.
        active = [i for i, on in enumerate(self.workdays_mask) if on]
        if not active:
            active = [0, 1, 2, 3, 4]
            self.workdays_mask = [i in active for i in range(7)]
        per_day = int(round(self.contract_week_minutes / max(1, len(active))))
        ws = self.wb["Werkpatroon"]
        idx_to_day = {0: "MA", 1: "DI", 2: "WO", 3: "DO", 4: "VR", 5: "ZA", 6: "ZO"}
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
        ws.append(["weekdag", "W", "MAX"])
        for i in range(7):
            cur = self.weekday_pattern.get(i, {"W": "00:00", "M": "00:00"})
            if self.workdays_mask[i]:
                m = max(per_day, hhmm_to_minutes(normalize_hhmm(cur.get("M", "00:00"), "00:00")))
                w = min(m, max(hhmm_to_minutes(normalize_hhmm(cur.get("W", "00:00"), "00:00")), m))
                m_txt = minutes_to_hhmm(m)
                w_txt = minutes_to_hhmm(w)
            else:
                w_txt = "00:00"
                m_txt = "00:00"
            self.weekday_pattern[i] = {"W": w_txt, "V": "00:00", "Z": "00:00", "M": m_txt}
            ws.append([idx_to_day[i], w_txt, m_txt])

    def planned_work_minutes_between(self, start: date, end: date) -> int:
        total = 0
        for dt in daterange(start, end):
            if dt.year != self.year:
                continue
            d = self.get_day(dt)
            total += hhmm_to_minutes(d.w)
        return total

    def planned_free_minutes_between(self, start: date, end: date) -> int:
        total = 0
        for dt in daterange(start, end):
            if dt.year != self.year:
                continue
            d = self.get_day(dt)
            total += hhmm_to_minutes(d.v) + hhmm_to_minutes(d.z)
        return total

    def annual_required_minutes(self) -> int:
        return max(0, int(self.contract_year_minutes))

    def illness_reduction_minutes_between(self, start: date, end: date) -> int:
        """Aantal minuten dat van plan-noodzaak afgaat door ziekte-markeringen."""
        total = 0
        for dt in daterange(start, end):
            key = dt.strftime("%Y-%m-%d")
            info = str(self.extra_info_data.get(key, "") or "").strip().casefold()
            if info != "ziekte":
                continue
            day_limit = self.get_day_limit(dt)
            if day_limit <= 0:
                continue
            total += day_limit
        return total

    def get_budget_overview(self) -> dict[str, int]:
        # We rekenen jaarbudget en verlofbudget vanuit geplande data, niet vanuit gewerkte data.
        # Daarmee blijft dit blok een planning-instrument: "wat heb ik nog te plannen / over".
        # Carry-over wordt expliciet meegenomen zodat de jaarstartpositie klopt.
        start = date(self.year, 1, 1)
        end = date(self.year, 12, 31)
        planned_work = self.planned_work_minutes_between(start, end)
        planned_free = self.planned_free_minutes_between(start, end)
        illness_reduction = self.illness_reduction_minutes_between(start, end)
        required = self.annual_required_minutes()
        target = max(0, required + self.carry_hours_prev_minutes - illness_reduction)
        to_plan = max(0, target - planned_work)

        stat_total = max(0, self.vacation_stat_minutes)
        extra_total_raw = self.vacation_extra_minutes + self.carry_vac_prev_minutes
        if extra_total_raw < 0:
            stat_total = max(0, stat_total + extra_total_raw)
            extra_total = 0
        else:
            extra_total = extra_total_raw
        stat_used = min(stat_total, planned_free)
        extra_used = min(extra_total, max(0, planned_free - stat_total))
        return {
            "employment_pct": int(round(self.employment_pct)),
            "fulltime_week": self.fulltime_week_minutes,
            "contract_week": self.contract_week_minutes,
            "contract_year": required,
            "carry_hours_prev": self.carry_hours_prev_minutes,
            "carry_vac_prev": self.carry_vac_prev_minutes,
            "illness_reduction_year": illness_reduction,
            "required_year": required,
            "planned_year": planned_work,
            "target_year": target,
            "to_plan_year": to_plan,
            "vacation_planned": planned_free,
            "vacation_stat_total": stat_total,
            "vacation_extra_total": extra_total,
            "vacation_stat_left": max(0, stat_total - stat_used),
            "vacation_extra_left": max(0, extra_total - extra_used),
        }

    def get_vacation_budget_text(self) -> str:
        b = self.get_budget_overview()
        return (
            f"Vrij-budget: wettelijk over {minutes_to_hhmm(b['vacation_stat_left'])} / "
            f"bovenwettelijk over {minutes_to_hhmm(b['vacation_extra_left'])} "
            f"(ingepland vrij {minutes_to_hhmm(b['vacation_planned'])})"
        )

    def load_free_days(self):
        self.vakantie_dagen.clear()
        ws = self.wb["vrije dagen"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and row[1]:
                d = normalize_to_date(row[0])
                if d:
                    self.vakantie_dagen[d.strftime("%Y-%m-%d")] = str(row[1])

    def load_school_holidays(self):
        self.school_vakanties.clear()
        ws = self.wb["Schoolvakanties"]
        region = (self.school_region or "zuid").strip().casefold()
        if region not in {"noord", "midden", "zuid"}:
            region = "zuid"
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[1]:
                continue
            d = normalize_to_date(row[0])
            if not d:
                continue
            row_region = str(row[2] if len(row) > 2 and row[2] else "").strip().casefold()
            if row_region and row_region != region:
                continue
            self.school_vakanties[d.strftime("%Y-%m-%d")] = str(row[1])

    def load_planning(self):
        self.planned_data.clear()
        ws = self.wb["Planning"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            d = normalize_to_date(row[0])
            if not d:
                continue
            key = d.strftime("%Y-%m-%d")
            free_txt = self.vakantie_dagen.get(key, "")
            free_h = "00:00"
            if free_txt:
                left = free_txt.split(" - ", 1)[0].strip().lower()
                if left.endswith(" vrij") and len(left) >= 10 and left[2] == ":":
                    free_h = normalize_hhmm(left[:5], "00:00")
            self.planned_data[key] = {
                "W": normalize_hhmm(row[1] or "00:00"),
                "V": free_h,
                "Z": "00:00",
            }

    def load_worked(self):
        self.worked_data.clear()
        for month_name in MONTHS_NL:
            ws = self.wb[month_name]
            month = MONTHS_NL.index(month_name) + 1
            for row in ws.iter_rows(min_row=2, values_only=True):
                wk = row[0] if row else None
                if not isinstance(wk, str) or not wk.startswith("wk"):
                    continue
                try:
                    week = int(wk[2:])
                except ValueError:
                    continue
                for wd in range(7):
                    try:
                        dt = date.fromisocalendar(self.year, week, wd + 1)
                    except ValueError:
                        continue
                    if dt.month != month:
                        continue
                    val = row[wd + 1] if wd + 1 < len(row) else ""
                    self.worked_data[dt.strftime("%Y-%m-%d")] = normalize_hhmm(val or "", "")

    def highlight_today_excel(self):
        today = date.today()
        if today.year != self.year:
            return
        ws = self.wb[MONTHS_NL[today.month - 1]]
        changed = False
        for row in ws.iter_rows(min_row=2):
            wk = row[0].value
            if not isinstance(wk, str) or not wk.startswith("wk"):
                continue
            try:
                week = int(wk[2:])
            except ValueError:
                continue
            for c in range(7):
                try:
                    dt = date.fromisocalendar(self.year, week, c + 1)
                except ValueError:
                    continue
                if dt == today:
                    cell = row[c + 1]
                    cur = str(getattr(cell.fill, "start_color", None).rgb or "").upper()
                    if "ADD8E6" not in cur:
                        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                        changed = True
                    break
        cur_tab = str(getattr(ws.sheet_properties, "tabColor", None) or "")
        if "ADD8E6" not in cur_tab.upper():
            ws.sheet_properties.tabColor = "ADD8E6"
            changed = True
        if changed:
            self.safe_save()

    def get_day(self, dt: date) -> DayData:
        key = dt.strftime("%Y-%m-%d")
        p = self.planned_data.get(key, {"W": "00:00", "V": "00:00", "Z": "00:00"})
        return DayData(
            w=normalize_hhmm(p.get("W", "00:00")),
            v=normalize_hhmm(p.get("V", "00:00")),
            z=normalize_hhmm(p.get("Z", "00:00")),
            worked=normalize_hhmm(self.worked_data.get(key, ""), ""),
        )

    def get_timer_log(self, dt: date) -> dict[str, int]:
        key = dt.strftime("%Y-%m-%d")
        row = self.data_log.get(key)
        if not row:
            return {"work": 0, "idle": 0, "call": 0}
        return {
            "work": int(row.get("work", 0)),
            "idle": int(row.get("idle", 0)),
            "call": int(row.get("call", 0)),
        }

    def set_worked_seconds(self, dt: date, seconds: int):
        key = dt.strftime("%Y-%m-%d")
        worked = seconds_to_hhmmss(seconds)
        self.worked_data[key] = worked
        self._save_worked_cell(dt, worked)
        self.safe_save()

    def save_timer_log(self, dt: date, work_s: int, idle_s: int, call_s: int):
        key = dt.strftime("%Y-%m-%d")
        self.data_log[key] = {"work": max(0, int(work_s)), "idle": max(0, int(idle_s)), "call": max(0, int(call_s))}
        ws = self.wb["Data_log"]
        found = None
        for row in ws.iter_rows(min_row=2):
            if normalize_to_date(row[0].value) == dt:
                found = row
                break
        now_txt = datetime.now().strftime("%H:%M:%S")
        if found:
            found[1].value = now_txt
            found[2].value = seconds_to_hhmmss(work_s)
            found[3].value = seconds_to_hhmmss(idle_s)
            found[4].value = seconds_to_hhmmss(call_s)
        else:
            ws.append([dt, now_txt, seconds_to_hhmmss(work_s), seconds_to_hhmmss(idle_s), seconds_to_hhmmss(call_s)])
        self.set_worked_seconds(dt, work_s)

    def set_day(self, dt: date, data: DayData, reason: str = "", extra_info: str = ""):
        key = dt.strftime("%Y-%m-%d")
        self.planned_data[key] = {"W": data.w, "V": data.v, "Z": data.z}
        self.worked_data[key] = data.worked
        self._save_planning_row(dt, data)
        self._save_worked_cell(dt, data.worked)
        info = (extra_info or "").strip()
        if info:
            self.extra_info_data[key] = info
            self._save_extra_info_cell(dt, info)
        else:
            self.extra_info_data.pop(key, None)
            self._delete_extra_info_cell(dt)
        if hhmm_to_minutes(data.v) > 0 or hhmm_to_minutes(data.z) > 0:
            abs_val = data.v if hhmm_to_minutes(data.v) > 0 else data.z
            text = f"{abs_val} vrij"
            details = (reason or "").strip()
            # Op officiële feestdagen houden we de feestdagnaam vast als reden
            # zodra er vrij gepland is en de gebruiker geen expliciete reden invult.
            if not details and dt in self.nl_holidays:
                details = str(self.nl_holidays.get(dt) or "").strip()
            if details:
                text = f"{text} - {details}"
            self._save_free_day(dt, text)
        else:
            self._delete_free_day(dt)
        self.safe_save()

    def _save_planning_row(self, dt: date, d: DayData):
        ws = self.wb["Planning"]
        found = None
        for row in ws.iter_rows(min_row=2, max_col=2):
            if normalize_to_date(row[0].value) == dt:
                found = row
                break
        if found:
            found[1].value = d.w
        else:
            ws.append([dt, d.w])
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
        rows.sort(key=lambda x: normalize_to_date(x[0]) or date.min)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        for i, row in enumerate(rows, start=2):
            ws.cell(i, 1, row[0]).number_format = "DD-MM-YYYY"
            ws.cell(i, 2, normalize_hhmm(row[1] or "00:00"))

    def _save_worked_cell(self, dt: date, worked: str):
        ws = self.wb[MONTHS_NL[dt.month - 1]]
        week_key = f"wk{dt.isocalendar()[1]}"
        row_obj = None
        for row in ws.iter_rows(min_row=2):
            if row[0].value == week_key:
                row_obj = row
                break
        if row_obj is None:
            ws.append([week_key, "", "", "", "", "", "", "", ""])
            row_obj = next(ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row))
        row_obj[dt.weekday() + 1].value = worked

    def _save_free_day(self, dt: date, text: str):
        ws = self.wb["vrije dagen"]
        exists = False
        for row in ws.iter_rows(min_row=2):
            if normalize_to_date(row[0].value) == dt:
                row[1].value = text
                exists = True
                break
        if not exists:
            ws.append([dt, text])
        self.vakantie_dagen[dt.strftime("%Y-%m-%d")] = text

    def _delete_free_day(self, dt: date):
        ws = self.wb["vrije dagen"]
        rows = []
        for row in ws.iter_rows(min_row=2):
            if normalize_to_date(row[0].value) == dt:
                rows.append(row[0].row)
        for r in sorted(rows, reverse=True):
            ws.delete_rows(r, 1)
        self.vakantie_dagen.pop(dt.strftime("%Y-%m-%d"), None)

    def get_day_limit(self, dt: date) -> int:
        raw = self.weekday_pattern.get(dt.weekday(), {}).get("M", "08:00")
        return hhmm_to_minutes(normalize_hhmm(raw, "08:00"))

    def day_reason(self, dt: date) -> str:
        key = dt.strftime("%Y-%m-%d")
        if key in self.vakantie_dagen:
            return self.vakantie_dagen[key]
        if key in self.school_vakanties:
            return self.school_vakanties[key]
        if dt in self.nl_holidays:
            return str(self.nl_holidays.get(dt) or "")
        return ""

    def get_saldo_minutes(self) -> int:
        norm = 0
        free = 0
        worked = 0
        daily = int((38 * 60) / 5)
        start = date(self.year, 1, 1)
        end = min(date.today(), date(self.year, 12, 31))
        for dt in daterange(start, end):
            if dt.weekday() < 5 and dt not in self.nl_holidays:
                norm += daily
            d = self.get_day(dt)
            free += hhmm_to_minutes(d.v)
            worked += hhmm_to_minutes(d.worked)
        expected = max(0, norm - free)
        return worked - expected

    def get_saldo_text(self) -> str:
        diff = self.get_saldo_minutes()
        norm = 0
        free = 0
        worked = 0
        daily = int((38 * 60) / 5)
        start = date(self.year, 1, 1)
        end = min(date.today(), date(self.year, 12, 31))
        for dt in daterange(start, end):
            if dt.weekday() < 5 and dt not in self.nl_holidays:
                norm += daily
            d = self.get_day(dt)
            free += hhmm_to_minutes(d.v)
            worked += hhmm_to_minutes(d.worked)
        expected = max(0, norm - free)
        sign = "+" if diff >= 0 else "-"
        return f"Saldo {sign}{minutes_to_hhmm(abs(diff))} (gewerkt {minutes_to_hhmm(worked)} / norm {minutes_to_hhmm(expected)})"

    def export_copy(self) -> str:
        self.safe_save()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        dst = os.path.join(self.base_dir, f"Time_tabel_{self.year}_export_{stamp}.xlsx")
        shutil.copy2(self.path, dst)
        return dst


# ============================================================================
# DIALOGS EN POPUPS
# Alle modale invoervensters staan hieronder gegroepeerd.
# Deze dialogs zijn bewust klein gehouden: één duidelijk doel per popup, met
# lokale validatie, zodat de hoofdlogica in MainWindow overzichtelijk blijft.
# ============================================================================
class DayEditDialog(FramelessDialog):
    """Dialog voor één kalenderdag (uren, vrij, gewerkt en context-info).

    Dit is het primaire handmatige correctiepunt van de app. De dialog doet
    daarom zowel validatie als budgetzicht, zodat een gebruiker direct ziet
    wat de impact is voordat de wijziging wordt opgeslagen.
    """
    def __init__(
        self,
        parent: QWidget,
        dt: date,
        day_data: DayData,
        day_limit: int,
        current_reason: str = "",
        current_extra_info: str = "",
        extra_info_options: list[str] | None = None,
        extra_info_colors: dict[str, str] | None = None,
        free_budget_text: str = "",
        allow_tvt: bool = True,
    ):
        super().__init__(parent)
        self.setWindowTitle(f"Dag bewerken - {dt.strftime('%d-%m-%Y')}")
        self.setModal(True)
        self.data_out: DayData | None = None
        self.reason_out = ""
        self.extra_info_out = ""
        self.day_limit = day_limit
        self.existing_worked = day_data.worked

        self.e_w = QLineEdit(day_data.w)
        self.e_v = QLineEdit(day_data.v)
        self.e_worked = QLineEdit(normalize_hhmm(day_data.worked, "00:00"))
        self.e_reason = QLineEdit()

        self.extra_options = normalize_extra_info_options(extra_info_options)
        if not allow_tvt:
            self.extra_options = [o for o in self.extra_options if o.strip().casefold() != "tijd voor tijd"]
        self.extra_info_colors = dict(extra_info_colors or {})
        existing_type, existing_reason = split_reason_and_type(current_reason)
        selected_extra = current_extra_info.strip()
        if not selected_extra and existing_type and existing_type != "Vrij":
            selected_extra = existing_type
        if selected_extra and selected_extra.casefold() not in {o.casefold() for o in self.extra_options}:
            self.extra_options.append(selected_extra)

        self.extra_buttons: dict[str, QPushButton] = {}
        self.selected_extra: str | None = None
        self.extra_host = QWidget()
        chip_grid = QGridLayout(self.extra_host)
        chip_grid.setContentsMargins(0, 0, 0, 0)
        chip_grid.setHorizontalSpacing(6)
        chip_grid.setVerticalSpacing(6)
        cols = 3
        for idx, opt in enumerate(self.extra_options):
            btn = QPushButton(opt)
            btn.setCheckable(True)
            btn.clicked.connect(lambda checked, o=opt: self._on_extra_clicked(o, checked))
            btn.setStyleSheet(self._extra_button_style(self.extra_info_colors.get(opt, default_extra_info_color(opt))))
            self.extra_buttons[opt] = btn
            chip_grid.addWidget(btn, idx // cols, idx % cols)

        self.e_reason.setText(existing_reason)
        for w in (self.e_w, self.e_v, self.e_worked):
            w.setPlaceholderText("hh:mm")
            force_hhmm_line_edit(w)
        self._syncing = False

        if selected_extra:
            self._set_extra_selection(selected_extra)
        else:
            self._set_extra_selection(None)

        form = QFormLayout()
        form.addRow("Werk (W)", self.e_w)
        form.addRow("Vrij (V)", self.e_v)
        form.addRow("Gewerkt (handmatig)", self.e_worked)
        form.addRow("Aanvullende informatie", self.extra_host)
        form.addRow("Reden vrij", self.e_reason)

        btn_cancel = QPushButton("Annuleren")
        btn_save = QPushButton("Opslaan")
        btn_cancel.clicked.connect(self.reject)
        btn_save.clicked.connect(self.on_save)

        row = QHBoxLayout()
        row.addStretch(1)
        row.addWidget(btn_cancel)
        row.addWidget(btn_save)

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 10)
        root.setSpacing(8)
        root.addLayout(form)
        root.addWidget(QLabel(f"Daglimiet: {minutes_to_hhmm(day_limit)}"))
        self.lbl_total = QLabel("")
        root.addWidget(self.lbl_total)
        if free_budget_text.strip():
            lbl_budget = QLabel(free_budget_text.strip())
            lbl_budget.setWordWrap(True)
            lbl_budget.setStyleSheet("color:#b8c7d8;")
            root.addWidget(lbl_budget)
        root.addLayout(row)
        self.resize(560, 320)
        self.e_w.editingFinished.connect(lambda: self._rebalance("w"))
        self.e_v.editingFinished.connect(lambda: self._rebalance("v"))
        self._rebalance(None)

    def _extra_button_style(self, color: str) -> str:
        return (
            f"QPushButton {{ background:{color}; color:#f3f8ff; border:1px solid #5b6b80; border-radius:7px; padding:6px 10px; }}"
            "QPushButton:hover { border:1px solid #8aa2bf; }"
            "QPushButton:checked { border:2px solid #e6edf3; }"
        )

    def _on_extra_clicked(self, option: str, checked: bool):
        if checked:
            self._set_extra_selection(option)
        else:
            self._set_extra_selection(None)

    def _set_extra_selection(self, option: str | None):
        found = None
        if option:
            for key in self.extra_buttons.keys():
                if key.casefold() == option.strip().casefold():
                    found = key
                    break
        self.selected_extra = found
        for key, btn in self.extra_buttons.items():
            btn.blockSignals(True)
            btn.setChecked(key == found)
            btn.blockSignals(False)

    def _set_minutes(self, edit: QLineEdit, minutes: int):
        edit.setText(minutes_to_hhmm(max(0, minutes)))

    def _rebalance(self, changed: str | None):
        if self._syncing:
            return
        self._syncing = True
        w_min = min(hhmm_to_minutes(normalize_hhmm(self.e_w.text())), self.day_limit)
        v_min = min(hhmm_to_minutes(normalize_hhmm(self.e_v.text())), self.day_limit)
        total = w_min + v_min
        if total > self.day_limit:
            overflow = total - self.day_limit
            if changed == "w":
                reduce_v = min(v_min, overflow)
                v_min -= reduce_v
                overflow -= reduce_v
                if overflow > 0:
                    w_min = max(0, w_min - overflow)
            elif changed == "v":
                reduce_w = min(w_min, overflow)
                w_min -= reduce_w
                overflow -= reduce_w
                if overflow > 0:
                    v_min = max(0, v_min - overflow)
            else:
                reduce_v = min(v_min, overflow)
                v_min -= reduce_v
                overflow -= reduce_v
                if overflow > 0:
                    w_min = max(0, w_min - overflow)
        self._set_minutes(self.e_w, w_min)
        self._set_minutes(self.e_v, v_min)
        total = w_min + v_min
        self.lbl_total.setText(f"Totaal: {minutes_to_hhmm(total)} / {minutes_to_hhmm(self.day_limit)}")
        self._syncing = False

    def on_save(self):
        self._rebalance(None)
        w = normalize_hhmm(self.e_w.text())
        v = normalize_hhmm(self.e_v.text())
        z = "00:00"
        worked = normalize_hhmm(self.e_worked.text(), "00:00")
        for x in (w, v, worked):
            if not parse_hhmm_strict(x):
                QMessageBox.warning(self, "Fout", "Gebruik hh:mm (00:00 t/m 23:59).")
                return
        if hhmm_to_minutes(w) + hhmm_to_minutes(v) + hhmm_to_minutes(z) > self.day_limit:
            QMessageBox.warning(self, "Fout", f"Planning mag niet boven {minutes_to_hhmm(self.day_limit)}.")
            return
        self.data_out = DayData(w=w, v=v, z=z, worked=worked)
        self.reason_out = self.e_reason.text().strip()
        picked = (self.selected_extra or "").strip()
        self.extra_info_out = picked
        self.accept()


class MonthCard(QGroupBox):
    day_double_clicked = Signal(date)
    day_range_selected = Signal(date, date)
    day_clicked = Signal(date, bool)
    day_shift_double_clicked = Signal(date)

    def __init__(
        self,
        store: StorageBackend,
        year: int,
        month: int,
        mode: str,
        focus_mode: bool = False,
        dark_mode: bool = False,
        colors: dict[str, str] | None = None,
        extra_info_colors: dict[str, str] | None = None,
    ):
        q = (month - 1) // 3 + 1
        super().__init__(f"{MONTHS_NL[month - 1]}  (Q{q})")
        self.store = store
        self.year = year
        self.month = month
        self.mode = mode
        self.focus_mode = focus_mode
        self.dark_mode = dark_mode
        self.colors = colors or dict(PLANNED_COLOR_DEFAULTS)
        self.extra_info_colors = dict(extra_info_colors or {})
        self.cell_map: dict[tuple[int, int], date] = {}
        self.table = QTableWidget(self)
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels(["WK"] + WEEKDAYS + ["Wk totaal"])
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table.setAlternatingRowColors(True)
        self.table.setMouseTracking(True)
        self.table.setShowGrid(False)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table.setItemDelegate(CalendarCellDelegate(self))
        self.table.cellClicked.connect(self._on_click)
        self.table.cellDoubleClicked.connect(self._on_double)
        self.table.setMinimumHeight(560 if focus_mode else 220)
        self.table.horizontalHeader().setMinimumSectionSize(70)
        if not self.focus_mode:
            self.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            self.setMinimumWidth(680)
            self.setMaximumWidth(680)
        else:
            self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            self.setMinimumWidth(0)
            self.setMaximumWidth(16777215)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(8, 10, 8, 8)
        lay.addWidget(self.table)
        self.refresh()

    def _weeks(self):
        first = date(self.year, self.month, 1)
        days = calendar.monthrange(self.year, self.month)[1]
        week_rows = {}
        for offset in range(days):
            dt = first + timedelta(days=offset)
            wn = dt.isocalendar()[1]
            wd = dt.weekday()
            if wn not in week_rows:
                week_rows[wn] = [None] * 7
            week_rows[wn][wd] = dt
        return sorted(week_rows.items())

    def _bg(self, dt: date) -> QColor:
        key = dt.strftime("%Y-%m-%d")
        if dt == date.today():
            return QColor(self.colors["bg_today_dark"])
        mode = getattr(self, "mode", "planned")
        if mode == "worked":
            # In Uren-weergave tonen we geen "vrij"-achtergrond voor geplande vrije dagen.
            # Zo blijft deze view primair een gewerkt-uren overzicht.
            if dt in self.store.nl_holidays:
                return QColor(self.colors["bg_holiday_dark"])
            if key in self.store.school_vakanties:
                return QColor(self.colors["bg_school_dark"])
            if dt.weekday() >= 5:
                return QColor(self.colors["bg_weekend_dark"])
            return QColor(self.colors["bg_default_dark"])
        if dt in self.store.nl_holidays:
            return QColor(self.colors["bg_holiday_dark"])
        if key in self.store.school_vakanties:
            return QColor(self.colors["bg_school_dark"])
        if dt.weekday() >= 5:
            return QColor(self.colors["bg_weekend_dark"])
        return QColor(self.colors["bg_default_dark"])

    def _extra_info_color(self, info: str) -> str | None:
        if not info:
            return None
        info_key = info.strip().casefold()
        for k, v in self.extra_info_colors.items():
            if k.strip().casefold() == info_key:
                return v
        return None

    def _worked_hours_base_bg(self, dt: date) -> QColor:
        """Basis-achtergrond voor urenblok in Uren-tab (zonder geplande-vrij override)."""
        key = dt.strftime("%Y-%m-%d")
        if dt == date.today():
            return QColor(self.colors["bg_today_dark"])
        if dt in self.store.nl_holidays:
            return QColor(self.colors["bg_holiday_dark"])
        if key in self.store.school_vakanties:
            return QColor(self.colors["bg_school_dark"])
        if dt.weekday() >= 5:
            return QColor(self.colors["bg_weekend_dark"])
        return QColor(self.colors["bg_default_dark"])

    def _planned_hours_brush(self, dt: date, fallback: QColor, day_data: DayData | None = None, info: str | None = None) -> QBrush:
        info = (info if info is not None else self.store.get_extra_info(dt)).strip()
        override = self._extra_info_color(info)
        if override:
            return QBrush(QColor(override))
        d = day_data or self.store.get_day(dt)
        w_m = hhmm_to_minutes(d.w)
        v_m = hhmm_to_minutes(d.v)
        total = w_m + v_m + hhmm_to_minutes(d.z)
        if total <= 0:
            return QBrush(fallback)
        work_clr = QColor(self.colors["planned_work_bg_dark"])
        free_clr = QColor(self.colors["planned_free_bg_dark"])
        if w_m > 0 and v_m > 0:
            ratio = w_m / max(1, (w_m + v_m))
            grad = QLinearGradient(0.0, 0.0, 1.0, 0.0)
            grad.setCoordinateMode(QGradient.ObjectBoundingMode)
            stop = max(0.05, min(0.95, ratio))
            grad.setColorAt(0.0, work_clr)
            grad.setColorAt(stop, work_clr)
            grad.setColorAt(min(1.0, stop + 0.001), free_clr)
            grad.setColorAt(1.0, free_clr)
            return QBrush(grad)
        if w_m > 0:
            return QBrush(work_clr)
        if v_m > 0:
            return QBrush(free_clr)
        return QBrush(fallback)

    def _worked_hours_brush(self, dt: date, fallback: QColor, day_data: DayData | None = None) -> QBrush:
        """Uren-tab: toon vrij-aandeel als blend op de standaard urenachtergrond."""
        d = day_data or self.store.get_day(dt)
        w_m = hhmm_to_minutes(d.w)
        v_m = hhmm_to_minutes(d.v) + hhmm_to_minutes(d.z)
        total = w_m + v_m
        if total <= 0 or v_m <= 0:
            return QBrush(fallback)
        base_clr = QColor(fallback)
        free_clr = QColor(self.colors["planned_free_bg_dark"])
        if v_m >= total:
            return QBrush(free_clr)

        free_ratio = v_m / max(1, total)
        stop = max(0.05, min(0.95, free_ratio))
        grad = QLinearGradient(0.0, 0.0, 1.0, 0.0)
        grad.setCoordinateMode(QGradient.ObjectBoundingMode)
        grad.setColorAt(0.0, free_clr)
        grad.setColorAt(stop, free_clr)
        grad.setColorAt(min(1.0, stop + 0.001), base_clr)
        grad.setColorAt(1.0, base_clr)
        return QBrush(grad)

    def _hours_text(self, dt: date, day_data: DayData | None = None) -> str:
        d = day_data or self.store.get_day(dt)
        if self.mode == "planned":
            total = hhmm_to_minutes(d.w) + hhmm_to_minutes(d.v) + hhmm_to_minutes(d.z)
            if total == 0:
                reason = self.store.day_reason(dt)
                return reason[:18] if reason else "00:00"
            return minutes_to_hhmm(total)
        if hhmm_to_minutes(d.w) == 0 and hhmm_to_minutes(d.v) > 0:
            return d.v
        return d.worked or "00:00"

    @staticmethod
    def _seconds_to_hhmm_short(seconds: int) -> str:
        return minutes_to_hhmm(int(max(0, int(seconds))) // 60)

    def _worked_wpc_text(self, dt: date, day_data: DayData | None = None) -> str:
        d = day_data or self.store.get_day(dt)
        row = self.store.get_timer_log(dt)
        w_txt = normalize_hhmm(d.worked or "00:00", "00:00")
        p_txt = self._seconds_to_hhmm_short(int(row.get("idle", 0)))
        c_txt = self._seconds_to_hhmm_short(int(row.get("call", 0)))
        return f"W {w_txt}\nP {p_txt}\nC {c_txt}"

    def _day_number_color(self, dt: date, day_data: DayData | None = None) -> QColor:
        key = dt.strftime("%Y-%m-%d")
        if dt == date.today():
            return QColor(self.colors["daynum_today_dark"])
        if self.mode == "planned":
            d = day_data or self.store.get_day(dt)
            has_planning = (hhmm_to_minutes(d.w) + hhmm_to_minutes(d.v) + hhmm_to_minutes(d.z)) > 0
            if has_planning:
                if dt in self.store.nl_holidays:
                    return QColor(self.colors["daynum_holiday_dark"])
                if key in self.store.school_vakanties:
                    return QColor(self.colors["daynum_school_dark"])
                if key in self.store.vakantie_dagen:
                    return QColor(self.colors["daynum_vacation_dark"])
                return QColor(self.colors["daynum_default_dark"])
            if dt.weekday() >= 5:
                return QColor(self.colors["daynum_weekend_dark"])
            return QColor(self.colors["daynum_default_dark"])
        if dt in self.store.nl_holidays:
            return QColor(self.colors["daynum_holiday_dark"])
        if key in self.store.school_vakanties:
            return QColor(self.colors["daynum_school_dark"])
        if dt.weekday() >= 5:
            return QColor(self.colors["daynum_weekend_dark"])
        return QColor(self.colors["daynum_default_dark"])

    def refresh(self):
        self.table.setUpdatesEnabled(False)
        self.table.viewport().setUpdatesEnabled(False)
        self.table.blockSignals(True)
        try:
            total_label = "Gepland" if self.mode == "planned" else "Gewerkt"
            self.table.setHorizontalHeaderLabels(["WK"] + WEEKDAYS + [total_label])
            weeks = self._weeks()
            target_rows = len(weeks) * 2
            structure_changed = self.table.rowCount() != target_rows
            if structure_changed:
                self.table.clearSpans()
                self.table.setRowCount(target_rows)
            self.cell_map.clear()
            day_h = 38 if self.focus_mode else 18
            hour_h = 100 if self.focus_mode else 24
            for week_i, (week_no, week_days) in enumerate(weeks):
                day_row = week_i * 2
                hour_row = day_row + 1
                if structure_changed:
                    self.table.setRowHeight(day_row, day_h)
                    self.table.setRowHeight(hour_row, hour_h)

                wk = self.table.item(day_row, 0)
                if wk is None:
                    wk = QTableWidgetItem("")
                    self.table.setItem(day_row, 0, wk)
                wk.setText(f"wk{week_no}" if week_no is not None else "")
                wk.setTextAlignment(Qt.AlignCenter)
                wk.setBackground(QColor(self.colors["weeknum_bg_dark"]))
                wk.setForeground(QColor(self.colors["weeknum_fg_dark"]))
                wk.setData(Qt.UserRole, True)
                if structure_changed:
                    self.table.setSpan(day_row, 0, 2, 1)

                total_min = 0
                iso_year = None
                iso_week = None
                seed_dt = next((x for x in week_days if x is not None), None)
                if seed_dt:
                    iso = seed_dt.isocalendar()
                    iso_year, iso_week = iso[0], iso[1]
                for wd, dt in enumerate(week_days):
                    c = wd + 1
                    day_item = self.table.item(day_row, c)
                    if day_item is None:
                        day_item = QTableWidgetItem("")
                        self.table.setItem(day_row, c, day_item)
                    hours_item = self.table.item(hour_row, c)
                    if hours_item is None:
                        hours_item = QTableWidgetItem("")
                        self.table.setItem(hour_row, c, hours_item)
                    day_item.setTextAlignment(Qt.AlignCenter)
                    hours_item.setTextAlignment(Qt.AlignCenter)
                    if dt:
                        day_data = self.store.get_day(dt)
                        info = self.store.get_extra_info(dt)
                        bg = self._bg(dt)
                        day_item.setText(str(dt.day))
                        day_item.setBackground(bg)
                        day_item.setForeground(self._day_number_color(dt, day_data))
                        if self.mode == "worked" and self.focus_mode:
                            hours_item.setText(self._worked_wpc_text(dt, day_data))
                        else:
                            hours_item.setText(self._hours_text(dt, day_data))
                        if self.mode == "planned":
                            hours_item.setBackground(self._planned_hours_brush(dt, bg, day_data, info))
                        else:
                            base_bg = self._worked_hours_base_bg(dt)
                            hours_item.setBackground(self._worked_hours_brush(dt, base_bg, day_data))
                        hours_item.setForeground(QColor(self.colors["hours_fg_dark"]))
                        if self.mode == "worked" and self.focus_mode:
                            hours_item.setTextAlignment(Qt.AlignCenter)
                            hours_item.setFont(QFont("Consolas", 12, QFont.Bold))
                        day_item.setData(Qt.UserRole, True)
                        hours_item.setData(Qt.UserRole, True)
                        reason = self.store.day_reason(dt)
                        if not reason and dt.strftime("%Y-%m-%d") in self.store.school_vakanties:
                            reason = self.store.school_vakanties.get(dt.strftime("%Y-%m-%d"), "")
                        if reason:
                            day_item.setToolTip(reason)
                            hours_item.setToolTip(reason)
                        else:
                            day_item.setToolTip("")
                            hours_item.setToolTip("")
                        self.cell_map[(day_row, c)] = dt
                        self.cell_map[(hour_row, c)] = dt
                        if self.mode == "planned":
                            total_min += hhmm_to_minutes(day_data.w) + hhmm_to_minutes(day_data.v) + hhmm_to_minutes(day_data.z)
                        else:
                            total_min += hhmm_to_minutes(day_data.worked)
                    else:
                        blank = QColor(self.colors["empty_bg_dark"])
                        day_item.setText("")
                        hours_item.setText("")
                        day_item.setBackground(blank)
                        hours_item.setBackground(blank)
                        day_item.setData(Qt.UserRole, False)
                        hours_item.setData(Qt.UserRole, False)
                        day_item.setToolTip("")
                        hours_item.setToolTip("")

                if iso_year is not None and iso_week is not None:
                    week_total = 0
                    planned_work_week = 0
                    planned_free_week = 0
                    idle_week_s = 0
                    call_week_s = 0
                    for d in range(1, 8):
                        try:
                            wdt = date.fromisocalendar(iso_year, iso_week, d)
                        except ValueError:
                            continue
                        day = self.store.get_day(wdt)
                        planned_work_week += hhmm_to_minutes(day.w)
                        planned_free_week += hhmm_to_minutes(day.v) + hhmm_to_minutes(day.z)
                        row = self.store.get_timer_log(wdt)
                        idle_week_s += int(row.get("idle", 0))
                        call_week_s += int(row.get("call", 0))
                        if self.mode == "planned":
                            week_total += hhmm_to_minutes(day.w) + hhmm_to_minutes(day.v) + hhmm_to_minutes(day.z)
                        else:
                            week_total += hhmm_to_minutes(day.worked)
                    if self.mode == "worked" and self.focus_mode:
                        if week_total == 0 and planned_work_week == 0 and planned_free_week > 0:
                            tot_txt = "Vrij\nW 00:00\nP 00:00\nC 00:00"
                        else:
                            tot_txt = (
                                f"W {minutes_to_hhmm(week_total)}\n"
                                f"P {self._seconds_to_hhmm_short(idle_week_s)}\n"
                                f"C {self._seconds_to_hhmm_short(call_week_s)}"
                            )
                    elif self.mode == "worked" and week_total == 0 and planned_work_week == 0 and planned_free_week > 0:
                        tot_txt = "vrij"
                    else:
                        tot_txt = minutes_to_hhmm(week_total)
                else:
                    tot_txt = minutes_to_hhmm(total_min)
                tot = self.table.item(day_row, 8)
                if tot is None:
                    tot = QTableWidgetItem("")
                    self.table.setItem(day_row, 8, tot)
                tot.setText(tot_txt)
                tot.setTextAlignment(Qt.AlignCenter)
                tot.setBackground(QColor(self.colors["weektotal_bg_dark"]))
                tot.setForeground(QColor(self.colors["weektotal_fg_dark"]))
                tot.setFont(QFont("Segoe UI", 9, QFont.Bold))
                if self.mode == "worked" and self.focus_mode:
                    tot.setTextAlignment(Qt.AlignCenter)
                    tot.setFont(QFont("Consolas", 12, QFont.Bold))
                tot.setData(Qt.UserRole, True)
                if structure_changed:
                    self.table.setSpan(day_row, 8, 2, 1)

            if structure_changed:
                if self.focus_mode:
                    self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
                    self.table.setColumnWidth(0, 92)
                    for c in range(1, 8):
                        self.table.horizontalHeader().setSectionResizeMode(c, QHeaderView.Stretch)
                    self.table.horizontalHeader().setSectionResizeMode(8, QHeaderView.Fixed)
                    self.table.setColumnWidth(8, 148)
                else:
                    self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
                    self.table.setColumnWidth(0, 56)
                    for c in range(1, 8):
                        self.table.horizontalHeader().setSectionResizeMode(c, QHeaderView.Fixed)
                        self.table.setColumnWidth(c, 66)
                    self.table.horizontalHeader().setSectionResizeMode(8, QHeaderView.Fixed)
                    self.table.setColumnWidth(8, 106)

                total_h = self.table.horizontalHeader().height()
                for r in range(self.table.rowCount()):
                    total_h += self.table.rowHeight(r)
                total_h += 8
                self.table.setMinimumHeight(total_h)
                self.table.setMaximumHeight(total_h)
                if not self.focus_mode:
                    # Gebruik vaste hoogte op basis van 6 weekrijen zonder extra lege rasterrijen.
                    header_h = self.table.horizontalHeader().height()
                    fixed_h = header_h + (6 * (18 + 24)) + 8
                    self.table.setMinimumHeight(fixed_h)
                    self.table.setMaximumHeight(fixed_h)
                    self.setMinimumHeight(fixed_h + 56)
                    self.setMaximumHeight(fixed_h + 56)
        finally:
            self.table.blockSignals(False)
            self.table.viewport().setUpdatesEnabled(True)
            self.table.setUpdatesEnabled(True)
            self.table.viewport().update()

    def _on_double(self, row: int, col: int):
        dt = self.cell_map.get((row, col))
        if dt:
            mods = QApplication.keyboardModifiers()
            if mods & Qt.ShiftModifier:
                self.day_shift_double_clicked.emit(dt)
                return
            self.day_double_clicked.emit(dt)

    def _on_click(self, row: int, col: int):
        dt = self.cell_map.get((row, col))
        if not dt:
            return
        mods = QApplication.keyboardModifiers()
        self.day_clicked.emit(dt, bool(mods & Qt.ShiftModifier))

    def select_date_range(self, start_dt: date, end_dt: date):
        a, b = (start_dt, end_dt) if start_dt <= end_dt else (end_dt, start_dt)
        self.table.clearSelection()
        for (row, col), dt in self.cell_map.items():
            if a <= dt <= b:
                item = self.table.item(row, col)
                if item is not None:
                    item.setSelected(True)


class CalendarBoard(QWidget):
    day_double_clicked = Signal(date)
    day_range_selected = Signal(date, date)

    def __init__(
        self,
        store: StorageBackend,
        year: int,
        mode: str,
        months: list[int],
        per_row: int,
        focus_mode: bool = False,
        dark_mode: bool = False,
        colors: dict[str, str] | None = None,
        extra_info_colors: dict[str, str] | None = None,
    ):
        super().__init__()
        self.store = store
        self.year = year
        self.mode = mode
        self.months = months
        self.per_row = per_row
        self._active_per_row = per_row
        self.focus_mode = focus_mode
        self.dark_mode = dark_mode
        self.colors = colors or dict(PLANNED_COLOR_DEFAULTS)
        self.extra_info_colors = dict(extra_info_colors or {})
        self.cards: list[MonthCard] = []
        self._bulk_anchor_date: date | None = None
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 10)
        root.setSpacing(8)
        root.setContentsMargins(0, 0, 0, 0)
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setObjectName("calendarScroll")
        self.host = QWidget()
        self.host.setObjectName("calendarHost")
        self.grid = QGridLayout(self.host)
        self.grid.setContentsMargins(4, 4, 4, 4)
        self.grid.setSpacing(4)
        self.scroll.setWidget(self.host)
        root.addWidget(self.scroll)
        self.rebuild()

    def rebuild(self):
        cols = 1 if self.focus_mode else self.per_row
        self._active_per_row = cols

        while self.grid.count():
            item = self.grid.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()
        self.cards.clear()
        for i, month in enumerate(self.months):
            card = MonthCard(
                self.store,
                self.year,
                month,
                self.mode,
                focus_mode=self.focus_mode,
                dark_mode=self.dark_mode,
                colors=self.colors,
                extra_info_colors=self.extra_info_colors,
            )
            card.day_double_clicked.connect(self.day_double_clicked.emit)
            card.day_clicked.connect(self._on_card_day_clicked)
            card.day_shift_double_clicked.connect(self._on_card_day_shift_double)
            self.cards.append(card)
            self.grid.addWidget(card, i // cols, i % cols)
        for c in range(cols):
            self.grid.setColumnStretch(c, 1)

    def set_mode(self, mode: str):
        self.mode = mode
        for c in self.cards:
            c.mode = mode

    def set_month(self, month: int):
        self.months = [month]
        self.per_row = 1
        self.rebuild()

    def refresh(self):
        self.setUpdatesEnabled(False)
        try:
            for c in self.cards:
                c.refresh()
        finally:
            self.setUpdatesEnabled(True)

    def _select_range_all_cards(self, start_dt: date, end_dt: date):
        for c in self.cards:
            c.select_date_range(start_dt, end_dt)

    def _on_card_day_clicked(self, dt: date, shift: bool):
        if shift and self._bulk_anchor_date is not None:
            a = self._bulk_anchor_date
            self._select_range_all_cards(min(a, dt), max(a, dt))
            return
        self._bulk_anchor_date = dt
        self._select_range_all_cards(dt, dt)

    def _on_card_day_shift_double(self, dt: date):
        if self._bulk_anchor_date is None:
            self._bulk_anchor_date = dt
            self._select_range_all_cards(dt, dt)
            return
        a = self._bulk_anchor_date
        s, e = (min(a, dt), max(a, dt))
        self._select_range_all_cards(s, e)
        self.day_range_selected.emit(s, e)

    def set_dark_mode(self, enabled: bool):
        self.dark_mode = enabled
        for c in self.cards:
            c.dark_mode = enabled

    def set_colors(self, colors: dict[str, str]):
        self.colors = colors
        for c in self.cards:
            c.colors = colors

    def set_extra_info_colors(self, extra_info_colors: dict[str, str]):
        self.extra_info_colors = dict(extra_info_colors or {})
        for c in self.cards:
            c.extra_info_colors = dict(self.extra_info_colors)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        return


class WorkPatternDialog(FramelessDialog):
    """Instellingen voor weekpatroon en dagmax op datumrange.

    Dit venster beheert vaste basiswaarden waarmee planning en normberekening
    gevoed worden. Het opsplitsen in een aparte dialog voorkomt dat de hoofd-UI
    te druk wordt met zeldzaam gebruikte configuratievelden.
    """
    def __init__(self, parent: QWidget, store: StorageBackend, year: int, apply_callback=None):
        super().__init__(parent)
        self.store = store
        self.year = year
        self.apply_callback = apply_callback
        self.setWindowTitle("Werkpatroon en Dagmax")
        self.setModal(True)
        self.resize(640, 440)

        self.inputs_w: dict[int, QLineEdit] = {}
        self.inputs_m: dict[int, QLineEdit] = {}
        day_names = ["MA", "DI", "WO", "DO", "VR", "ZA", "ZO"]

        form_box = QWidget()
        grid = QGridLayout(form_box)
        title = QLabel("Weekpatroon (hh:mm)")
        title.setStyleSheet("font: 11pt 'Segoe UI Semibold'; color:#e6edf7; padding-bottom:4px;")
        grid.addWidget(title, 0, 0, 1, 3)
        grid.addWidget(QLabel("Dag"), 1, 0)
        grid.addWidget(QLabel("Werk"), 1, 1)
        grid.addWidget(QLabel("Max dag"), 1, 2)

        for i, dn in enumerate(day_names):
            row = i + 2
            grid.addWidget(QLabel(dn), row, 0)
            e_w = QLineEdit(self.store.weekday_pattern[i]["W"])
            e_m = QLineEdit(self.store.weekday_pattern[i]["M"])
            e_w.setPlaceholderText("hh:mm")
            e_m.setPlaceholderText("hh:mm")
            force_hhmm_line_edit(e_w)
            force_hhmm_line_edit(e_m, "00:00" if i > 4 else "08:00")
            self.inputs_w[i] = e_w
            self.inputs_m[i] = e_m
            grid.addWidget(e_w, row, 1)
            grid.addWidget(e_m, row, 2)

        btn_cancel = QPushButton("Annuleren")
        btn_save = QPushButton("Opslaan")
        btn_apply = QPushButton("Toepassen")
        btn_cancel.clicked.connect(self.reject)
        btn_save.clicked.connect(self._save_only)
        btn_apply.clicked.connect(self._apply)

        self.e_from = QLineEdit(f"01-01-{self.year}")
        self.e_to = QLineEdit(f"31-12-{self.year}")
        self.e_from.setPlaceholderText("dd-mm-jjjj")
        self.e_to.setPlaceholderText("dd-mm-jjjj")

        range_row = QHBoxLayout()
        range_row.addWidget(QLabel("Van"))
        range_row.addWidget(self.e_from)
        range_row.addWidget(QLabel("Tot en met"))
        range_row.addWidget(self.e_to)
        btn_row = QHBoxLayout()
        btn_row.addStretch(1)
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_save)
        btn_row.addWidget(btn_apply)

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 10)
        root.setSpacing(8)
        root.addWidget(form_box)
        root.addLayout(range_row)
        root.addLayout(btn_row)

    def _apply(self):
        if not self._save_silent():
            return
        start_date = parse_nl_date(self.e_from.text())
        end_date = parse_nl_date(self.e_to.text())
        if not start_date or not end_date:
            QMessageBox.warning(self, "Fout", "Gebruik datumformaat dd-mm-jjjj voor van en tot.")
            return
        if start_date > end_date:
            QMessageBox.warning(self, "Fout", "Van-datum mag niet na tot-datum liggen.")
            return
        if self.apply_callback:
            self.apply_callback(start_date, end_date)
        self.accept()

    def _save_only(self):
        if self._save_silent():
            self.accept()

    def _save_silent(self):
        ws = self.store.wb["Werkpatroon"]
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
        ws.append(["weekdag", "W", "MAX"])
        idx_to_day = {0: "MA", 1: "DI", 2: "WO", 3: "DO", 4: "VR", 5: "ZA", 6: "ZO"}
        for i in range(7):
            w = normalize_hhmm(self.inputs_w[i].text())
            m = normalize_hhmm(self.inputs_m[i].text(), "00:00" if i > 4 else "08:00")
            if not (parse_hhmm_strict(w) and parse_hhmm_strict(m)):
                QMessageBox.warning(self, "Fout", "Gebruik hh:mm voor werk en max.")
                return False
            if hhmm_to_minutes(w) > hhmm_to_minutes(m):
                QMessageBox.warning(self, "Fout", f"{idx_to_day[i]}: werk hoger dan max.")
                return False
            self.store.weekday_pattern[i] = {"W": w, "V": "00:00", "Z": "00:00", "M": m}
            ws.append([idx_to_day[i], w, m])
        self.store.safe_save()
        return True


class ColorSettingsDialog(FramelessDialog):
    LABELS = {
        "bg_today_dark": "Vandaag achtergrond",
        "bg_vacation_dark": "Vrij-dag achtergrond",
        "bg_holiday_dark": "Feestdag achtergrond",
        "bg_school_dark": "Schoolvakantie achtergrond",
        "bg_weekend_dark": "Weekend achtergrond",
        "bg_default_dark": "Werkdag achtergrond",
        "daynum_today_dark": "Dagnummer vandaag",
        "daynum_holiday_dark": "Dagnummer feestdag",
        "daynum_vacation_dark": "Dagnummer vrij",
        "daynum_school_dark": "Dagnummer schoolvakantie",
        "daynum_weekend_dark": "Dagnummer weekend",
        "daynum_default_dark": "Dagnummer werkdag",
        "weeknum_bg_dark": "Weeknummer achtergrond",
        "weeknum_fg_dark": "Weeknummer tekst",
        "weektotal_bg_dark": "Weektotaal achtergrond",
        "weektotal_fg_dark": "Weektotaal tekst",
        "empty_bg_dark": "Lege cel achtergrond",
        "hours_fg_dark": "Uren tekst",
        "planned_work_bg_dark": "Planning werk achtergrond",
        "planned_free_bg_dark": "Planning vrij achtergrond",
        "timer_bg_dark": "Timer achtergrond",
        "timer_text_dark": "Timer tekst",
        "timer_btn_dark": "Timer knop",
    }
    GROUPS = [
        ("Kalender dagen", ["bg_today_dark", "bg_vacation_dark", "bg_holiday_dark", "bg_school_dark", "bg_weekend_dark", "bg_default_dark", "empty_bg_dark"]),
        ("Dagnummering", ["daynum_today_dark", "daynum_holiday_dark", "daynum_vacation_dark", "daynum_school_dark", "daynum_weekend_dark", "daynum_default_dark"]),
        ("Weekoverzicht en uren", ["weeknum_bg_dark", "weeknum_fg_dark", "weektotal_bg_dark", "weektotal_fg_dark", "hours_fg_dark"]),
        ("Planning en timer", ["planned_work_bg_dark", "planned_free_bg_dark", "timer_bg_dark", "timer_text_dark", "timer_btn_dark"]),
    ]

    def __init__(
        self,
        parent: QWidget,
        colors: dict[str, str],
        extra_info_options: list[str],
        extra_info_colors: dict[str, str],
    ):
        super().__init__(parent)
        self.setWindowTitle("Kleuren kalender")
        self.setModal(True)
        self.resize(760, 860)
        self.colors = dict(PLANNED_COLOR_DEFAULTS)
        self.colors.update(dict(colors or {}))
        self.extra_info_options = [o for o in normalize_extra_info_options(extra_info_options) if o.casefold() != "optioneel"]
        self.extra_info_colors = dict(extra_info_colors or {})
        self.preview_buttons: dict[str, QPushButton] = {}
        self.preview_extra_buttons: dict[str, QPushButton] = {}
        self.preview_bg_targets: dict[str, list[QWidget]] = {}
        self.preview_fg_targets: dict[str, list[QWidget]] = {}
        self.preview_timer_panel: QFrame | None = None
        self.preview_timer_text: QLabel | None = None
        self.preview_timer_btn: QPushButton | None = None
        for opt in self.extra_info_options:
            if opt not in self.extra_info_colors:
                self.extra_info_colors[opt] = default_extra_info_color(opt)

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 10)
        root.setSpacing(8)
        color_content = QWidget()
        color_content.setObjectName("colorContent")
        color_content_lay = QVBoxLayout(color_content)
        color_content_lay.setContentsMargins(0, 0, 0, 0)
        color_content_lay.setSpacing(8)
        for title, keys in self.GROUPS:
            color_content_lay.addWidget(self._build_color_group(title, keys))
        color_content_lay.addWidget(self._build_extra_group())
        color_content_lay.addStretch(1)

        color_scroll = QScrollArea()
        color_scroll.setObjectName("colorScroll")
        color_scroll.setWidgetResizable(True)
        color_scroll.setWidget(color_content)
        color_scroll.setMinimumHeight(440)
        color_scroll.viewport().setObjectName("colorScrollViewport")
        color_scroll.setStyleSheet(
            """
            QScrollArea#colorScroll {
                background: #263141;
                border: 1px solid #3b4759;
                border-radius: 10px;
            }
            QWidget#colorScrollViewport { background: #263141; }
            QWidget#colorContent { background: #263141; }
            """
        )

        preview_box = self._build_live_preview()
        self._refresh_previews()

        btn_reset = QPushButton("Reset standaard")
        btn_reset.clicked.connect(self.reset_defaults)
        btn_cancel = QPushButton("Annuleren")
        btn_cancel.clicked.connect(self.reject)
        btn_ok = QPushButton("Opslaan")
        btn_ok.clicked.connect(self.accept)

        row_btn = QHBoxLayout()
        row_btn.addWidget(btn_reset)
        row_btn.addStretch(1)
        row_btn.addWidget(btn_cancel)
        row_btn.addWidget(btn_ok)

        root.addWidget(preview_box)
        root.addWidget(color_scroll, 1)
        root.addLayout(row_btn)

    def _add_preview_target(self, widget: QWidget, bg_key: str | None = None, fg_key: str | None = None):
        if bg_key:
            self.preview_bg_targets.setdefault(bg_key, []).append(widget)
        if fg_key:
            self.preview_fg_targets.setdefault(fg_key, []).append(widget)

    def _build_live_preview(self) -> QGroupBox:
        box = QGroupBox("Live voorbeeld")
        grid = QGridLayout(box)
        grid.setHorizontalSpacing(8)
        grid.setVerticalSpacing(8)

        wk = QLabel("wk09")
        wk.setAlignment(Qt.AlignCenter)
        wk.setMinimumSize(62, 42)
        self._add_preview_target(wk, "weeknum_bg_dark", "weeknum_fg_dark")
        grid.addWidget(wk, 0, 0)

        c_default = QLabel("13\n08:00")
        c_default.setAlignment(Qt.AlignCenter)
        c_default.setMinimumSize(74, 52)
        self._add_preview_target(c_default, "bg_default_dark", "daynum_default_dark")
        grid.addWidget(c_default, 0, 1)

        c_today = QLabel("NU\n07:30")
        c_today.setAlignment(Qt.AlignCenter)
        c_today.setMinimumSize(74, 52)
        self._add_preview_target(c_today, "bg_today_dark", "daynum_today_dark")
        grid.addWidget(c_today, 0, 2)

        c_vac = QLabel("15\nVrij")
        c_vac.setAlignment(Qt.AlignCenter)
        c_vac.setMinimumSize(74, 52)
        self._add_preview_target(c_vac, "bg_vacation_dark", "daynum_vacation_dark")
        grid.addWidget(c_vac, 0, 3)

        c_holiday = QLabel("16\nFeest")
        c_holiday.setAlignment(Qt.AlignCenter)
        c_holiday.setMinimumSize(74, 52)
        self._add_preview_target(c_holiday, "bg_holiday_dark", "daynum_holiday_dark")
        grid.addWidget(c_holiday, 0, 4)

        c_school = QLabel("17\nSchool")
        c_school.setAlignment(Qt.AlignCenter)
        c_school.setMinimumSize(74, 52)
        self._add_preview_target(c_school, "bg_school_dark", "daynum_school_dark")
        grid.addWidget(c_school, 0, 5)

        c_weekend = QLabel("ZA\n00:00")
        c_weekend.setAlignment(Qt.AlignCenter)
        c_weekend.setMinimumSize(74, 52)
        self._add_preview_target(c_weekend, "bg_weekend_dark", "daynum_weekend_dark")
        grid.addWidget(c_weekend, 0, 6)

        c_empty = QLabel("")
        c_empty.setAlignment(Qt.AlignCenter)
        c_empty.setMinimumSize(74, 52)
        self._add_preview_target(c_empty, "empty_bg_dark")
        grid.addWidget(c_empty, 0, 7)

        planned_work = QLabel("Planning werk 08:00")
        planned_work.setAlignment(Qt.AlignCenter)
        planned_work.setMinimumHeight(34)
        self._add_preview_target(planned_work, "planned_work_bg_dark", "hours_fg_dark")
        grid.addWidget(planned_work, 1, 1, 1, 3)

        planned_free = QLabel("Planning vrij 08:00")
        planned_free.setAlignment(Qt.AlignCenter)
        planned_free.setMinimumHeight(34)
        self._add_preview_target(planned_free, "planned_free_bg_dark", "hours_fg_dark")
        grid.addWidget(planned_free, 1, 4, 1, 3)

        total = QLabel("38:00")
        total.setAlignment(Qt.AlignCenter)
        total.setMinimumSize(74, 34)
        self._add_preview_target(total, "weektotal_bg_dark", "weektotal_fg_dark")
        grid.addWidget(total, 1, 7)

        self.preview_timer_panel = QFrame()
        timer_lay = QHBoxLayout(self.preview_timer_panel)
        timer_lay.setContentsMargins(6, 4, 6, 4)
        timer_lay.setSpacing(8)
        self.preview_timer_text = QLabel("T: 01:23:45  P: 00:04:00  C: 00:12:10")
        self.preview_timer_btn = QPushButton("Open")
        self.preview_timer_btn.setEnabled(False)
        timer_lay.addWidget(self.preview_timer_text, 1)
        timer_lay.addWidget(self.preview_timer_btn)
        grid.addWidget(self.preview_timer_panel, 2, 0, 1, 8)

        return box

    def _refresh_live_preview(self):
        for key, widgets in self.preview_bg_targets.items():
            bg = self.colors.get(key, PLANNED_COLOR_DEFAULTS.get(key, "#263141"))
            for w in widgets:
                fg = self.colors.get("hours_fg_dark", "#e6edf7")
                if key.startswith("daynum_"):
                    fg = self.colors.get(key, fg)
                style = (
                    f"background:{bg}; color:{fg}; border:1px solid #344156; "
                    "border-radius:6px; padding:2px 4px;"
                )
                w.setStyleSheet(style)

        for key, widgets in self.preview_fg_targets.items():
            fg = self.colors.get(key, PLANNED_COLOR_DEFAULTS.get(key, "#e6edf7"))
            for w in widgets:
                cur = w.styleSheet()
                if "color:" in cur:
                    cur = re.sub(r"color\\s*:\\s*[^;]+;", f"color:{fg};", cur)
                    w.setStyleSheet(cur)
                else:
                    w.setStyleSheet(cur + f" color:{fg};")

        if self.preview_timer_panel and self.preview_timer_text and self.preview_timer_btn:
            tbg = self.colors.get("timer_bg_dark", PLANNED_COLOR_DEFAULTS["timer_bg_dark"])
            tfg = self.colors.get("timer_text_dark", PLANNED_COLOR_DEFAULTS["timer_text_dark"])
            tbtn = self.colors.get("timer_btn_dark", PLANNED_COLOR_DEFAULTS["timer_btn_dark"])
            self.preview_timer_panel.setStyleSheet(
                f"QFrame {{ background:{tbg}; border:1px solid #344156; border-radius:8px; }}"
            )
            self.preview_timer_text.setStyleSheet(f"color:{tfg}; font: 9pt 'Consolas';")
            self.preview_timer_btn.setStyleSheet(
                f"QPushButton {{ background:{tbtn}; color:#ffffff; border:1px solid #344156; border-radius:6px; padding:4px 10px; }}"
            )

    def _build_color_group(self, title: str, keys: list[str]) -> QGroupBox:
        box = QGroupBox(title)
        grid = QGridLayout(box)
        row = 0
        for key in keys:
            if key not in self.LABELS:
                continue
            grid.addWidget(QLabel(self.LABELS[key]), row, 0)
            btn = QPushButton("Kies kleur")
            btn.clicked.connect(lambda _, k=key: self.pick_color(k))
            self.preview_buttons[key] = btn
            grid.addWidget(btn, row, 1)
            row += 1
        return box

    def _build_extra_group(self) -> QGroupBox:
        box = QGroupBox("Aanvullende info kleuren")
        grid = QGridLayout(box)
        row = 0
        for opt in self.extra_info_options:
            grid.addWidget(QLabel(opt), row, 0)
            btn = QPushButton("Kies kleur")
            btn.clicked.connect(lambda _, o=opt: self.pick_extra_color(o))
            self.preview_extra_buttons[opt] = btn
            grid.addWidget(btn, row, 1)
            row += 1
        return box

    def _refresh_previews(self):
        for key, btn in self.preview_buttons.items():
            clr = self.colors.get(key, PLANNED_COLOR_DEFAULTS[key])
            btn.setText(clr.upper())
            btn.setStyleSheet(
                f"QPushButton {{ background:{clr}; color:#ffffff; border:1px solid #334155; border-radius:6px; padding:4px 8px; }}"
            )
        for opt, btn in self.preview_extra_buttons.items():
            clr = self.extra_info_colors.get(opt, default_extra_info_color(opt))
            btn.setText(clr.upper())
            btn.setStyleSheet(
                f"QPushButton {{ background:{clr}; color:#ffffff; border:1px solid #334155; border-radius:6px; padding:4px 8px; }}"
            )
        self._refresh_live_preview()

    def pick_color(self, key: str):
        current = QColor(self.colors.get(key, PLANNED_COLOR_DEFAULTS[key]))
        c = QColorDialog.getColor(current, self, f"Kies kleur: {self.LABELS.get(key, key)}")
        if c.isValid():
            self.colors[key] = c.name()
            self._refresh_previews()

    def pick_extra_color(self, option: str):
        current = QColor(self.extra_info_colors.get(option, default_extra_info_color(option)))
        c = QColorDialog.getColor(current, self, f"Kies kleur: {option}")
        if c.isValid():
            self.extra_info_colors[option] = c.name()
            self._refresh_previews()

    def reset_defaults(self):
        self.colors = dict(PLANNED_COLOR_DEFAULTS)
        self.extra_info_colors = {opt: default_extra_info_color(opt) for opt in self.extra_info_options}
        self._refresh_previews()


class TimerSettingsDialog(FramelessDialog):
    def __init__(self, parent: QWidget, idle_threshold_sec: int, include_lockscreen_idle: bool, include_sleep_idle: bool):
        super().__init__(parent)
        self.setWindowTitle("Timer idle/pauze instellingen")
        self.setModal(True)
        self.resize(460, 250)
        self.idle_threshold_sec = max(15, min(3600, int(idle_threshold_sec or 60)))
        self.include_lockscreen_idle = bool(include_lockscreen_idle)
        self.include_sleep_idle = bool(include_sleep_idle)

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 10)
        root.setSpacing(6)

        mins = max(1, round(self.idle_threshold_sec / 60))
        self.lbl_minutes = QLabel("")
        root.addWidget(self.lbl_minutes)
        self.slider_min = QSlider(Qt.Horizontal)
        self.slider_min.setRange(1, 60)
        self.slider_min.setValue(mins)
        self.slider_min.valueChanged.connect(self._update_minutes_label)
        root.addWidget(self.slider_min)
        self._update_minutes_label(self.slider_min.value())

        self.chk_lockscreen = QCheckBox("Lockscreen meetellen als pauze-kandidaat")
        self.chk_lockscreen.setChecked(self.include_lockscreen_idle)
        root.addWidget(self.chk_lockscreen)

        self.chk_sleep = QCheckBox("Sleep/hibernate meetellen als pauze-kandidaat")
        self.chk_sleep.setChecked(self.include_sleep_idle)
        root.addWidget(self.chk_sleep)

        hint = QLabel(
            "Advies: lockscreen aan laten voor pauzeregistratie, sleep/hibernate meestal uit om systeemevents te negeren."
        )
        hint.setWordWrap(True)
        hint.setStyleSheet("color:#b8c7d8;")
        root.addWidget(hint)

        btn_cancel = QPushButton("Annuleren")
        btn_save = QPushButton("Opslaan")
        btn_cancel.clicked.connect(self.reject)
        btn_save.clicked.connect(self._save)

        row = QHBoxLayout()
        row.addStretch(1)
        row.addWidget(btn_cancel)
        row.addWidget(btn_save)
        root.addLayout(row)

    def _update_minutes_label(self, value: int):
        self.lbl_minutes.setText(f"Pauze-drempel (idle): {int(value)} minuut/minuten")

    def _save(self):
        total = int(self.slider_min.value()) * 60
        self.idle_threshold_sec = max(60, min(3600, total))
        self.include_lockscreen_idle = bool(self.chk_lockscreen.isChecked())
        self.include_sleep_idle = bool(self.chk_sleep.isChecked())
        self.accept()


class GlassSettingsDialog(FramelessDialog):
    def __init__(self, parent: QWidget, inactive_glass_opacity: float):
        super().__init__(parent)
        self.setWindowTitle("Doorzichtigheid")
        self.setModal(True)
        self.resize(430, 180)
        try:
            op = float(inactive_glass_opacity)
        except Exception:
            op = 0.72
        self.inactive_glass_opacity = max(GLASS_OPACITY_MIN, min(GLASS_OPACITY_MAX, op))

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 10)
        root.setSpacing(6)

        self.lbl_glass = QLabel("")
        root.addWidget(self.lbl_glass)
        self.slider_glass = QSlider(Qt.Horizontal)
        self.slider_glass.setRange(int(GLASS_OPACITY_MIN * 100), int(GLASS_OPACITY_MAX * 100))
        self.slider_glass.setValue(int(round(self.inactive_glass_opacity * 100)))
        self.slider_glass.valueChanged.connect(self._update_glass_label)
        root.addWidget(self.slider_glass)
        self._update_glass_label(self.slider_glass.value())

        hint = QLabel(
            f"Wordt toegepast als de app niet actief is (Alt-Tab naar andere app).\n"
            f"Vaste grenzen: min {int(GLASS_OPACITY_MIN * 100)}% / max {int(GLASS_OPACITY_MAX * 100)}%."
        )
        hint.setWordWrap(True)
        root.addWidget(hint)

        btn_cancel = QPushButton("Annuleren")
        btn_save = QPushButton("Opslaan")
        btn_cancel.clicked.connect(self.reject)
        btn_save.clicked.connect(self._save)
        row = QHBoxLayout()
        row.addStretch(1)
        row.addWidget(btn_cancel)
        row.addWidget(btn_save)
        root.addLayout(row)

    def _update_glass_label(self, value: int):
        self.lbl_glass.setText(f"Doorzichtigheid bij focusverlies: {int(value)}%")

    def _save(self):
        self.inactive_glass_opacity = max(GLASS_OPACITY_MIN, min(GLASS_OPACITY_MAX, float(self.slider_glass.value()) / 100.0))
        self.accept()


class BulkPlanningDialog(FramelessDialog):
    """Bulk planner voor Werk/Vrij over een datumrange met weekdagfilter."""

    def __init__(
        self,
        parent: QWidget,
        year: int,
        extra_info_options: list[str] | None = None,
        allow_tvt: bool = True,
        start_date: date | None = None,
        end_date: date | None = None,
        weekday_limits: list[int] | None = None,
    ):
        super().__init__(parent)
        self.setWindowTitle("Bulk plannen")
        self.setModal(True)
        self.resize(760, 520)
        today = date.today()
        start_default = date(year, today.month, 1) if today.year == year else date(year, 1, 1)
        end_default = date(year, today.month, calendar.monthrange(year, today.month)[1]) if today.year == year else date(year, 12, 31)

        self.result_data: dict[str, object] | None = None
        self.weekday_limits = list(weekday_limits or [8 * 60] * 7)
        if len(self.weekday_limits) < 7:
            self.weekday_limits.extend([8 * 60] * (7 - len(self.weekday_limits)))
        self.weekday_limits = [max(0, int(v)) for v in self.weekday_limits[:7]]
        start_default = start_date or start_default
        end_default = end_date or end_default
        if start_default > end_default:
            start_default, end_default = end_default, start_default
        self.e_from = QDateEdit()
        self.e_from.setDisplayFormat("dd-MM-yyyy")
        self.e_from.setCalendarPopup(True)
        self.e_from.setDate(QDate(start_default.year, start_default.month, start_default.day))
        self.e_to = QDateEdit()
        self.e_to.setDisplayFormat("dd-MM-yyyy")
        self.e_to.setCalendarPopup(True)
        self.e_to.setDate(QDate(end_default.year, end_default.month, end_default.day))
        self.chk_skip_weekend = QCheckBox("Weekend overslaan")
        self.chk_skip_weekend.setChecked(True)
        self.chk_skip_holiday = QCheckBox("Feestdagen overslaan")
        self.chk_skip_holiday.setChecked(True)
        self.extra_options = ["(ongewijzigd)", "(leegmaken)"]
        opts = normalize_extra_info_options(extra_info_options)
        if not allow_tvt:
            opts = [o for o in opts if o.strip().casefold() != "tijd voor tijd"]
        self.extra_options.extend(opts)

        self.day_rows: list[dict[str, object]] = []
        weekday_box = QGroupBox("Per dag invullen (MA t/m ZO)")
        weekday_grid = QGridLayout(weekday_box)
        weekday_grid.setContentsMargins(8, 8, 8, 8)
        weekday_grid.setHorizontalSpacing(8)
        weekday_grid.setVerticalSpacing(6)
        weekday_grid.addWidget(QLabel("Dag"), 0, 0)
        weekday_grid.addWidget(QLabel("Actief"), 0, 1)
        weekday_grid.addWidget(QLabel("Werk"), 0, 2)
        weekday_grid.addWidget(QLabel("Vrij"), 0, 3)
        weekday_grid.addWidget(QLabel("Aanvullende info"), 0, 4)
        weekday_grid.addWidget(QLabel("Vrij-toelichting"), 0, 5)
        for i, wd in enumerate(WEEKDAYS):
            row = i + 1
            lbl = QLabel(wd)
            chk = QCheckBox()
            chk.setChecked(i < 5)
            e_w = QLineEdit("08:00" if i < 5 else "00:00")
            e_v = QLineEdit("00:00")
            force_hhmm_line_edit(e_w)
            force_hhmm_line_edit(e_v)
            cmb = QComboBox()
            for item in self.extra_options:
                cmb.addItem(item)
            cmb.setCurrentText("(ongewijzigd)")
            e_reason = QLineEdit("Vakantie {weekdag}" if i < 5 else "")
            e_reason.setPlaceholderText("bijv. Vakantie {datum}")
            limit_txt = minutes_to_hhmm(self.weekday_limits[i])
            e_w.setToolTip(f"Dagmax {wd}: {limit_txt}")
            e_v.setToolTip(f"Dagmax {wd}: {limit_txt}")
            e_w.editingFinished.connect(lambda idx=i: self._rebalance_day_row(idx, "w"))
            e_v.editingFinished.connect(lambda idx=i: self._rebalance_day_row(idx, "v"))
            weekday_grid.addWidget(lbl, row, 0)
            weekday_grid.addWidget(chk, row, 1)
            weekday_grid.addWidget(e_w, row, 2)
            weekday_grid.addWidget(e_v, row, 3)
            weekday_grid.addWidget(cmb, row, 4)
            weekday_grid.addWidget(e_reason, row, 5)
            self.day_rows.append({"check": chk, "work": e_w, "free": e_v, "extra": cmb, "reason": e_reason})
            self._rebalance_day_row(i, None)

        form = QFormLayout()
        form.addRow("Van", self.e_from)
        form.addRow("Tot en met", self.e_to)
        form.addRow("", self.chk_skip_weekend)
        form.addRow("", self.chk_skip_holiday)

        hint = QLabel("Tokens in toelichting: {datum} en {weekdag}. Dagmax uit werkpatroon blijft leidend.")
        hint.setWordWrap(True)
        hint.setStyleSheet("color:#b8c7d8;")

        btn_cancel = QPushButton("Annuleren")
        btn_apply = QPushButton("Toepassen")
        btn_cancel.clicked.connect(self.reject)
        btn_apply.clicked.connect(self._save)
        row_btn = QHBoxLayout()
        row_btn.addStretch(1)
        row_btn.addWidget(btn_cancel)
        row_btn.addWidget(btn_apply)

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 10)
        root.setSpacing(8)
        root.addLayout(form)
        root.addWidget(weekday_box, 1)
        root.addWidget(hint)
        root.addLayout(row_btn)

    def _rebalance_day_row(self, weekday_idx: int, changed: str | None):
        if weekday_idx < 0 or weekday_idx >= len(self.day_rows):
            return
        cfg = self.day_rows[weekday_idx]
        e_w = cfg["work"]
        e_v = cfg["free"]
        day_limit = self.weekday_limits[weekday_idx]
        w_min = min(hhmm_to_minutes(normalize_hhmm(e_w.text())), day_limit)
        v_min = min(hhmm_to_minutes(normalize_hhmm(e_v.text())), day_limit)
        total = w_min + v_min
        if total > day_limit:
            overflow = total - day_limit
            if changed == "w":
                reduce_v = min(v_min, overflow)
                v_min -= reduce_v
                overflow -= reduce_v
                if overflow > 0:
                    w_min = max(0, w_min - overflow)
            elif changed == "v":
                reduce_w = min(w_min, overflow)
                w_min -= reduce_w
                overflow -= reduce_w
                if overflow > 0:
                    v_min = max(0, v_min - overflow)
            else:
                reduce_v = min(v_min, overflow)
                v_min -= reduce_v
                overflow -= reduce_v
                if overflow > 0:
                    w_min = max(0, w_min - overflow)
        e_w.setText(minutes_to_hhmm(w_min))
        e_v.setText(minutes_to_hhmm(v_min))

    def _save(self):
        q_from = self.e_from.date()
        q_to = self.e_to.date()
        d_from = date(q_from.year(), q_from.month(), q_from.day())
        d_to = date(q_to.year(), q_to.month(), q_to.day())
        if d_from > d_to:
            d_from, d_to = d_to, d_from
        rows_cfg: list[dict[str, object]] = []
        active_any = False
        for i, cfg in enumerate(self.day_rows):
            chk = cfg["check"]
            e_w = cfg["work"]
            e_v = cfg["free"]
            cmb = cfg["extra"]
            e_reason = cfg["reason"]
            is_active = bool(chk.isChecked())
            w_h = normalize_hhmm(e_w.text(), "00:00")
            v_h = normalize_hhmm(e_v.text(), "00:00")
            if not parse_hhmm_strict(w_h) or not parse_hhmm_strict(v_h):
                QMessageBox.warning(self, "Fout", f"Ongeldige werk/vrij uren bij {WEEKDAYS[i]}. Gebruik hh:mm.")
                return
            if is_active:
                active_any = True
            rows_cfg.append(
                {
                    "weekday": i,
                    "active": is_active,
                    "work_hours": w_h,
                    "free_hours": v_h,
                    "extra_mode": cmb.currentText().strip(),
                    "reason_tpl": e_reason.text().strip(),
                }
            )

        if not active_any:
            QMessageBox.warning(self, "Fout", "Selecteer minimaal één weekdag.")
            return
        self.result_data = {
            "start": d_from,
            "end": d_to,
            "skip_weekend": bool(self.chk_skip_weekend.isChecked()),
            "skip_holiday": bool(self.chk_skip_holiday.isChecked()),
            "rows_cfg": rows_cfg,
        }
        self.accept()


class SchoolRegionDialog(FramelessDialog):
    def __init__(self, parent: QWidget, current_region: str):
        super().__init__(parent)
        self.setWindowTitle("Schoolvakantie regio")
        self.setModal(True)
        self.resize(340, 140)
        self.region = (current_region or "zuid").strip().casefold()
        if self.region not in {"noord", "midden", "zuid"}:
            self.region = "zuid"

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 10)
        root.setSpacing(8)

        form = QFormLayout()
        self.cb_region = QComboBox()
        self.cb_region.addItems(["noord", "midden", "zuid"])
        self.cb_region.setCurrentText(self.region)
        form.addRow("Regio", self.cb_region)
        root.addLayout(form)

        btn_cancel = QPushButton("Annuleren")
        btn_save = QPushButton("Opslaan")
        btn_cancel.clicked.connect(self.reject)
        btn_save.clicked.connect(self._save)
        row = QHBoxLayout()
        row.addStretch(1)
        row.addWidget(btn_cancel)
        row.addWidget(btn_save)
        root.addLayout(row)

    def _save(self):
        self.region = self.cb_region.currentText().strip().casefold()
        self.accept()


class ContractBudgetDialog(FramelessDialog):
    def __init__(
        self,
        parent: QWidget,
        employment_pct: float,
        fulltime_week_min: int,
        contract_year_min: int,
        vacation_stat_min: int,
        vacation_extra_min: int,
        carry_hours_prev_min: int,
        carry_vac_prev_min: int,
        workdays_mask: list[bool],
    ):
        super().__init__(parent)
        self.setWindowTitle("Contract en vakantiebudget")
        self.setModal(True)
        self.resize(540, 400)
        self.employment_pct = float(employment_pct)
        self.fulltime_week = format_hours_int(fulltime_week_min)
        self.contract_year = format_hours_int(contract_year_min)
        self.vacation_stat = format_hours_int(vacation_stat_min)
        self.vacation_extra = format_hours_int(vacation_extra_min)
        self.carry_hours_prev = format_hours_int(carry_hours_prev_min, signed=True)
        self.carry_vac_prev = format_hours_int(carry_vac_prev_min, signed=True)
        self.workdays_mask = [bool(x) for x in (workdays_mask or [True, True, True, True, True, False, False])][:7]
        while len(self.workdays_mask) < 7:
            self.workdays_mask.append(False)

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 10)
        root.setSpacing(8)
        form = QFormLayout()
        self.e_pct = QLineEdit(f"{int(round(self.employment_pct))}")
        self.e_fulltime = QLineEdit(self.fulltime_week)
        self.e_contract_year = QLineEdit(self.contract_year)
        self.e_stat = QLineEdit(self.vacation_stat)
        self.e_extra = QLineEdit(self.vacation_extra)
        self.e_carry_hours = QLineEdit(self.carry_hours_prev)
        self.e_carry_vac = QLineEdit(self.carry_vac_prev)
        self.e_pct.setMaxLength(3)
        self.e_pct.setFixedWidth(70)
        for e in (self.e_fulltime, self.e_contract_year, self.e_stat, self.e_extra):
            e.setPlaceholderText("uren")
            e.setMaxLength(4)
            e.setFixedWidth(90)
        self.e_carry_hours.setPlaceholderText("+/-uren")
        self.e_carry_vac.setPlaceholderText("+/-uren")
        self.e_carry_hours.setMaxLength(5)
        self.e_carry_vac.setMaxLength(5)
        self.e_carry_hours.setFixedWidth(90)
        self.e_carry_vac.setFixedWidth(90)
        form.addRow("Omvang dienstbetrekking (%)", self.e_pct)
        form.addRow("Fulltime uren/week (100%)", self.e_fulltime)
        form.addRow("Contracturen jaar", self.e_contract_year)
        form.addRow("Wettelijke vakantie (jaar)", self.e_stat)
        form.addRow("Bovenwettelijke vakantie (jaar)", self.e_extra)
        form.addRow("Meer/minder uren vorig jaar", self.e_carry_hours)
        form.addRow("Vakantie over/tekort vorig jaar", self.e_carry_vac)
        root.addLayout(form)

        days_box = QGroupBox("Werkdagen per week")
        days_grid = QGridLayout(days_box)
        self.day_checks: list[QCheckBox] = []
        for i, dn in enumerate(WEEKDAYS):
            cb = QCheckBox(dn)
            cb.setChecked(bool(self.workdays_mask[i]))
            self.day_checks.append(cb)
            days_grid.addWidget(cb, i // 4, i % 4)
        root.addWidget(days_box)

        btn_defaults = QPushButton("Herstel 100%")
        btn_defaults.clicked.connect(self._reset_100_defaults)
        btn_cancel = QPushButton("Annuleren")
        btn_save = QPushButton("Opslaan")
        btn_cancel.clicked.connect(self.reject)
        btn_save.clicked.connect(self._save)
        row = QHBoxLayout()
        row.addWidget(btn_defaults)
        row.addStretch(1)
        row.addWidget(btn_cancel)
        row.addWidget(btn_save)
        root.addLayout(row)

    def _reset_100_defaults(self):
        # Herstelt alleen de basisdefaults; carry-over velden blijven bewust intact.
        # Reden: carry-over vertegenwoordigt historische correcties en mag niet per ongeluk verdwijnen.
        # Deze knop is bedoeld als snelle reset van de "normale" contractinstelling voor dit jaar.
        self.e_pct.setText("100")
        self.e_fulltime.setText("38")
        self.e_contract_year.setText("1938")
        self.e_stat.setText("152")
        self.e_extra.setText("56")
        for i, cb in enumerate(self.day_checks):
            cb.setChecked(i < 5)

    def _save(self):
        try:
            pct = float(self.e_pct.text().strip().replace(",", "."))
        except Exception:
            QMessageBox.warning(self, "Fout", "Ongeldige % voor dienstbetrekking.")
            return
        if pct < 0 or pct > 200:
            QMessageBox.warning(self, "Fout", "Dienstbetrekking % moet tussen 0 en 200 liggen.")
            return
        fw = parse_duration_hhmm(self.e_fulltime.text(), default=-1)
        cy = parse_duration_hhmm(self.e_contract_year.text(), default=-1)
        s = parse_duration_hhmm(self.e_stat.text(), default=-1)
        x = parse_duration_hhmm(self.e_extra.text(), default=-1)
        ch = parse_duration_hhmm(self.e_carry_hours.text(), default=0, allow_signed=True)
        cv = parse_duration_hhmm(self.e_carry_vac.text(), default=0, allow_signed=True)
        if fw < 0 or cy < 0 or s < 0 or x < 0:
            QMessageBox.warning(self, "Fout", "Gebruik hele uren (of hh:mm) voor uren.")
            return
        mask = [cb.isChecked() for cb in self.day_checks]
        if sum(1 for x_on in mask if x_on) == 0:
            QMessageBox.warning(self, "Fout", "Selecteer minimaal 1 werkdag.")
            return
        self.employment_pct = pct
        self.fulltime_week = format_duration_hhmm(fw)
        self.contract_year = format_duration_hhmm(cy)
        self.vacation_stat = format_duration_hhmm(s)
        self.vacation_extra = format_duration_hhmm(x)
        self.carry_hours_prev = format_duration_hhmm(ch, signed=True)
        self.carry_vac_prev = format_duration_hhmm(cv, signed=True)
        self.workdays_mask = mask
        self.accept()


class ExtraInfoSettingsDialog(FramelessDialog):
    def __init__(self, parent: QWidget, options: list[str], enabled_options: list[str]):
        super().__init__(parent)
        self.setWindowTitle("Aanvullende informatie opties")
        self.setModal(True)
        self.resize(720, 520)
        self.options = normalize_extra_info_options(options)
        self.enabled_options = normalize_extra_info_enabled(self.options, enabled_options)

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 10)
        root.setSpacing(8)
        root.addWidget(QLabel("Beheer opties als losse regels en zet per optie actief/inactief."))

        self.table = QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(["Optie", "Actief"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setAlternatingRowColors(True)
        root.addWidget(self.table, 1)

        enabled_set = {o.casefold() for o in self.enabled_options}
        for opt in self.options:
            self._add_row(opt, opt.casefold() in enabled_set)

        add_row = QHBoxLayout()
        self.e_new = QLineEdit()
        self.e_new.setPlaceholderText("Nieuwe optie (bijv. training)")
        btn_add = QPushButton("Toevoegen")
        btn_remove = QPushButton("Verwijder geselecteerde")
        btn_add.clicked.connect(self._add_from_input)
        btn_remove.clicked.connect(self._remove_selected)
        add_row.addWidget(self.e_new, 1)
        add_row.addWidget(btn_add)
        add_row.addWidget(btn_remove)
        root.addLayout(add_row)

        btn_cancel = QPushButton("Annuleren")
        btn_save = QPushButton("Opslaan")
        btn_cancel.clicked.connect(self.reject)
        btn_save.clicked.connect(self._save)
        row = QHBoxLayout()
        row.addStretch(1)
        row.addWidget(btn_cancel)
        row.addWidget(btn_save)
        root.addLayout(row)

    def _contains_option(self, option: str) -> bool:
        key = option.casefold()
        for r in range(self.table.rowCount()):
            it = self.table.item(r, 0)
            if it and it.text().strip().casefold() == key:
                return True
        return False

    def _add_row(self, option: str, active: bool):
        r = self.table.rowCount()
        self.table.insertRow(r)

        it_name = QTableWidgetItem(option)
        it_name.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)
        self.table.setItem(r, 0, it_name)

        it_active = QTableWidgetItem("")
        it_active.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsUserCheckable)
        it_active.setCheckState(Qt.Checked if active else Qt.Unchecked)
        self.table.setItem(r, 1, it_active)

    def _add_from_input(self):
        val = self.e_new.text().strip()
        if not val:
            return
        if self._contains_option(val):
            QMessageBox.warning(self, "Bestaat al", f"Optie '{val}' bestaat al.")
            return
        self._add_row(val, True)
        self.e_new.clear()

    def _remove_selected(self):
        r = self.table.currentRow()
        if r < 0:
            return
        self.table.removeRow(r)

    def _save(self):
        options = []
        enabled = []
        seen = set()
        for r in range(self.table.rowCount()):
            it_name = self.table.item(r, 0)
            if not it_name:
                continue
            name = it_name.text().strip()
            if not name:
                continue
            key = name.casefold()
            if key in seen:
                QMessageBox.warning(self, "Dubbele optie", f"Optie '{name}' staat dubbel in de lijst.")
                return
            seen.add(key)
            options.append(name)
            it_active = self.table.item(r, 1)
            if it_active and it_active.checkState() == Qt.Checked:
                enabled.append(name)

        self.options = normalize_extra_info_options(options)
        self.enabled_options = normalize_extra_info_enabled(self.options, enabled)
        self.accept()


# ============================================================================
# TIMER SUBSYSTEEM
# Dit blok bevat de live-metering en de timer-UI.
# De timer tikt per seconde, detecteert idle/call-activiteit en synchroniseert
# periodiek naar opslag zonder dat de gebruiker handmatig hoeft op te slaan.
# ============================================================================
class TimerPanel(QFrame):
    """Timer met oud `Log tijd.py` gedrag (auto work/idle/call detectie).

    Deze component draait op een korte tick en schrijft periodiek weg naar
    opslag. De UI is compact gehouden zodat het venster klein kan blijven,
    terwijl de kernmeters toch live zichtbaar zijn.
    """

    height_changed = Signal(int)

    DEFAULT_IDLE_THRESHOLD_SEC = 60
    PAUSE_CONFIRM_TIMEOUT_SEC = 45
    CALL_PROBE_INTERVAL_SEC = 2.0
    CALL_HOLD_SECONDS = 10.0
    CALL_WARMUP_SECONDS = 5.0
    OUTLOOK_PROBE_INTERVAL_SEC = 60.0

    CALL_APP_PROCESSES = {
        "teams",
        "ms-teams",
        "msteams",
        "teams2",
        "zoom",
        "slack",
        "webex",
    }
    BROWSER_PROCESSES = {
        "chrome",
        "msedge",
        "firefox",
        "brave",
        "opera",
    }
    CALL_TITLE_KEYWORDS = (
        "teams meeting",
        "meeting in microsoft teams",
        "microsoft teams meeting",
        "zoom meeting",
        "google meet",
        "meet.google.com",
        "webex",
        "vergadering",
        "vergadering met",
        "gesprek met",
        "in gesprek",
        "in call",
        "on a call",
    )
    OUTLOOK_CALL_KEYWORDS = (
        "teams",
        "zoom",
        "google meet",
        "meet.google.com",
        "webex",
    )

    class LASTINPUTINFO(ctypes.Structure):
        _fields_ = [("cbSize", wintypes.UINT), ("dwTime", wintypes.DWORD)]

    @staticmethod
    def _make_calendar_icon() -> QIcon:
        pm = QPixmap(18, 18)
        pm.fill(Qt.transparent)
        p = QPainter(pm)
        p.setRenderHint(QPainter.Antialiasing, True)
        body = QRect(1, 2, 16, 15)
        p.setBrush(QColor("#2f4f75"))
        p.setPen(QPen(QColor("#8fb3da"), 1))
        p.drawRoundedRect(body, 3, 3)
        header = QRect(1, 2, 16, 5)
        p.setBrush(QColor("#4e7fba"))
        p.setPen(Qt.NoPen)
        p.drawRoundedRect(header, 3, 3)
        p.setBrush(QColor("#e8f3ff"))
        p.setPen(Qt.NoPen)
        p.drawRect(4, 1, 2, 3)
        p.drawRect(12, 1, 2, 3)
        p.setBrush(QColor("#dce9f7"))
        p.drawRect(4, 9, 3, 2)
        p.drawRect(8, 9, 3, 2)
        p.drawRect(12, 9, 2, 2)
        p.drawRect(4, 12, 3, 2)
        p.drawRect(8, 12, 3, 2)
        p.drawRect(12, 12, 2, 2)
        p.end()
        return QIcon(pm)

    def __init__(self, store: StorageBackend, on_refresh, on_toggle_full, on_drag_move, idle_threshold_sec: int = DEFAULT_IDLE_THRESHOLD_SEC):
        super().__init__()
        self.store = store
        self.on_refresh = on_refresh
        self.on_toggle_full = on_toggle_full
        self.on_drag_move = on_drag_move
        self.running = True
        self.tracking_enabled = True
        self.work_seconds = 0
        self.idle_seconds = 0
        self.call_seconds = 0
        self.current_day = date.today()
        self.save_tick = 0
        self.save_interval = 5
        self.warning_alpha = 1.0
        self.warning_direction = -0.1
        self.warning_active = False
        self._drag_start = None
        self.idle_threshold_sec = self.DEFAULT_IDLE_THRESHOLD_SEC
        self.inactive_glass = False
        self.inactive_glass_opacity = 0.72
        self.min_inactive_glass_opacity = GLASS_OPACITY_MIN
        self.max_inactive_glass_opacity = GLASS_OPACITY_MAX
        self._colors_cache: dict[str, str] = {}
        self._call_detect_cache = False
        self._call_hold_until = 0.0
        self._call_last_probe = 0.0
        self._call_warmup_start = 0.0
        self._outlook_last_probe = 0.0
        self._outlook_meeting_now_cache = False
        self._lock_elapsed_seconds = 0
        self._await_unlock_pause_prompt = False
        self.call_debug_enabled = CALL_DEBUG_ENABLED
        self.call_debug_log_enabled = CALL_DEBUG_LOG_ENABLED or self.call_debug_enabled
        self._call_debug_text = "call-debug: init"
        self._call_debug_last_emit_ts = 0.0
        self._call_debug_last_text = ""
        base_dir = getattr(self.store, "base_dir", os.path.dirname(os.path.abspath(__file__)))
        self._call_debug_log_path = os.path.join(base_dir, "call_detection_debug.log")
        self._idle_episode_active = False
        self._idle_episode_seconds = 0
        self._unconfirmed_pause_seconds = 0
        self.os_session_locked = False
        self.os_system_sleeping = False
        self.include_lockscreen_idle = True
        self.include_sleep_idle = False

        self.setObjectName("timerPanel")
        self.setMinimumWidth(372)
        self.setMaximumWidth(372)
        self.setMinimumHeight(56)
        self.setMaximumHeight(56)

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 10)
        root.setSpacing(8)
        root.setContentsMargins(4, 4, 4, 4)
        root.setSpacing(2)

        row = QHBoxLayout()
        row.setContentsMargins(0, 0, 0, 0)
        row.setSpacing(4)
        self.lbl_line = QLabel("T: 00:00:00  P: 00:00:00  C: 00:00:00")
        self.lbl_line.setStyleSheet("font: 11pt 'Consolas'; font-weight:700;")
        self.lbl_line.setAlignment(Qt.AlignVCenter | Qt.AlignLeft)
        self.lbl_line.setMinimumHeight(42)
        self.lbl_line.installEventFilter(self)
        row.addWidget(self.lbl_line, 1)

        self.btn_open = QPushButton("")
        self.btn_open.setIcon(self._make_calendar_icon())
        self.btn_open.setIconSize(QPixmap(18, 18).size())
        self.btn_open.setToolTip("Planner tonen/verbergen")
        self.btn_open.setFixedSize(44, 42)
        self.btn_open.setCheckable(True)
        self.btn_open.clicked.connect(self.on_toggle_full)
        row.addWidget(self.btn_open)
        root.addLayout(row)

        self.idle_label = QLabel("IDLE binnenkort")
        self.idle_label.setVisible(False)
        self.idle_label.setAlignment(Qt.AlignCenter)
        self.idle_label.setStyleSheet("font: 9pt 'Consolas'; font-weight:700;")
        root.addWidget(self.idle_label)

        self.call_debug_label = QLabel("")
        self.call_debug_label.setVisible(bool(self.call_debug_enabled))
        self.call_debug_label.setWordWrap(True)
        self.call_debug_label.setStyleSheet("font: 7.8pt 'Consolas'; color:#9fb0c5;")
        if self.call_debug_enabled:
            self.call_debug_label.setText("call-debug: waiting for first probe")
        root.addWidget(self.call_debug_label)

        self.pause_confirm_row = QWidget(self)
        pause_row_layout = QHBoxLayout(self.pause_confirm_row)
        pause_row_layout.setContentsMargins(2, 0, 2, 0)
        pause_row_layout.setSpacing(4)
        self.pause_confirm_label = QLabel("Idle gedetecteerd. Als pauze registreren?")
        self.pause_confirm_label.setStyleSheet("font: 8.7pt 'Segoe UI'; font-weight:600;")
        pause_row_layout.addWidget(self.pause_confirm_label, 1)
        self.btn_pause_yes = QPushButton("Ja")
        self.btn_pause_no = QPushButton("Nee")
        self.btn_pause_yes.setFixedHeight(24)
        self.btn_pause_no.setFixedHeight(24)
        self.btn_pause_yes.clicked.connect(self._confirm_pause_yes)
        self.btn_pause_no.clicked.connect(self._confirm_pause_no)
        pause_row_layout.addWidget(self.btn_pause_yes)
        pause_row_layout.addWidget(self.btn_pause_no)
        self.pause_confirm_row.setVisible(False)
        root.addWidget(self.pause_confirm_row)

        self.timer = QTimer(self)
        self.timer.setInterval(1000)
        self.timer.timeout.connect(self._tick)
        self.warn_timer = QTimer(self)
        self.warn_timer.setInterval(100)
        self.warn_timer.timeout.connect(self._animate_warning)
        self.pause_confirm_timer = QTimer(self)
        self.pause_confirm_timer.setSingleShot(True)
        self.pause_confirm_timer.setInterval(self.PAUSE_CONFIRM_TIMEOUT_SEC * 1000)
        self.pause_confirm_timer.timeout.connect(self._confirm_pause_no)

        self.set_idle_threshold(idle_threshold_sec)
        self.load_today()
        self.update_ui()
        self.timer.start()

    def set_tracking_enabled(self, enabled: bool):
        self.tracking_enabled = bool(enabled)
        if not self.tracking_enabled:
            self._hide_idle_warning()
            self._idle_episode_active = False
            self._idle_episode_seconds = 0
            self._confirm_pause_no()
            self.persist()
            self.timer.stop()
        elif self.running and not self.timer.isActive():
            self.timer.start()
        self.update_ui()

    def eventFilter(self, obj, event):
        if obj is self.lbl_line:
            if event.type() == QEvent.MouseButtonPress and event.button() == Qt.LeftButton:
                self._drag_start = event.globalPosition().toPoint()
                return True
            if event.type() == QEvent.MouseMove and self._drag_start is not None:
                delta = event.globalPosition().toPoint() - self._drag_start
                self._drag_start = event.globalPosition().toPoint()
                self.on_drag_move(delta)
                return True
            if event.type() == QEvent.MouseButtonRelease:
                self._drag_start = None
                return True
        return super().eventFilter(obj, event)

    def apply_colors(self, colors: dict[str, str]):
        self._colors_cache = dict(colors or {})
        bg = self._colors_cache.get("timer_bg_dark", "#000000")
        fg = self._colors_cache.get("timer_text_dark", "#7CFC00")
        btn = self._colors_cache.get("timer_btn_dark", "#222222")

        bg_c = QColor(bg)
        fg_c = QColor(fg)
        btn_c = QColor(btn)
        if self.inactive_glass:
            alpha = int(255 * self.inactive_glass_opacity)
            alpha = max(48, min(235, alpha))
            bg_c.setAlpha(alpha)
            fg_c = fg_c.lighter(115)
            fg_c.setAlpha(min(240, alpha + 65))
            btn_c.setAlpha(min(235, alpha + 35))

        bg_css = f"rgba({bg_c.red()}, {bg_c.green()}, {bg_c.blue()}, {bg_c.alpha()})"
        fg_css = f"rgba({fg_c.red()}, {fg_c.green()}, {fg_c.blue()}, {fg_c.alpha()})"
        btn_css = f"rgba({btn_c.red()}, {btn_c.green()}, {btn_c.blue()}, {btn_c.alpha()})"
        self.setStyleSheet(
            f"""
            QFrame#timerPanel {{
                background:qlineargradient(
                    x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(255,255,255,{18 if self.inactive_glass else 8}),
                    stop:1 {bg_css}
                );
                border:1px solid rgba(220,235,255,{130 if self.inactive_glass else 90});
                border-radius:10px;
            }}
            QFrame#timerPanel QLabel {{ color:{fg_css}; }}
            QFrame#timerPanel QPushButton {{
                color:#f8fbff;
                border-radius:7px;
                border-top:1px solid #9cb3cf;
                border-left:1px solid #839ab8;
                border-right:1px solid #32465f;
                border-bottom:2px solid #203248;
                background:{btn_css};
                background:qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #5e7895, stop:1 #355071);
                font: 12pt 'Segoe UI Emoji';
                padding: 0px;
            }}
            QFrame#timerPanel QPushButton:hover {{
                border-top:1px solid #b8ccee;
                background:qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #6f88a3, stop:1 #3f6287);
            }}
            QFrame#timerPanel QPushButton:checked, QFrame#timerPanel QPushButton:pressed {{
                border-top:2px solid #1e2f45;
                border-left:2px solid #22364d;
                border-right:1px solid #7f99b8;
                border-bottom:1px solid #9ab4d3;
                background:qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #2c4666, stop:1 #1f3650);
                padding-top:2px;
                padding-left:1px;
            }}
            """
        )
        self.idle_label.setStyleSheet("color:#111111; background:#ffff66; font: 10pt 'Consolas'; font-weight:700; border-radius:4px;")

    def set_inactive_glass(self, enabled: bool):
        flag = bool(enabled)
        if self.inactive_glass == flag:
            return
        self.inactive_glass = flag
        self.apply_colors(self._colors_cache)

    def set_glass_opacity(self, opacity: float):
        try:
            v = float(opacity)
        except Exception:
            v = 0.72
        v = max(self.min_inactive_glass_opacity, min(self.max_inactive_glass_opacity, v))
        if abs(v - self.inactive_glass_opacity) < 0.0001:
            return
        self.inactive_glass_opacity = v
        if self.inactive_glass:
            self.apply_colors(self._colors_cache)

    def set_glass_limits(self, min_opacity: float, max_opacity: float):
        changed = (
            abs(GLASS_OPACITY_MIN - self.min_inactive_glass_opacity) > 0.0001
            or abs(GLASS_OPACITY_MAX - self.max_inactive_glass_opacity) > 0.0001
        )
        self.min_inactive_glass_opacity = GLASS_OPACITY_MIN
        self.max_inactive_glass_opacity = GLASS_OPACITY_MAX
        if self.inactive_glass_opacity < GLASS_OPACITY_MIN or self.inactive_glass_opacity > GLASS_OPACITY_MAX:
            self.inactive_glass_opacity = max(GLASS_OPACITY_MIN, min(GLASS_OPACITY_MAX, self.inactive_glass_opacity))
            changed = True
        if changed and self.inactive_glass:
            self.apply_colors(self._colors_cache)

    def get_idle_time(self) -> float:
        if sys.platform != "win32":
            return 0.0
        lii = self.LASTINPUTINFO()
        lii.cbSize = ctypes.sizeof(self.LASTINPUTINFO)
        if ctypes.windll.user32.GetLastInputInfo(ctypes.byref(lii)):
            try:
                now64 = int(ctypes.windll.kernel32.GetTickCount64())
                now32 = now64 & 0xFFFFFFFF
                ms = (now32 - int(lii.dwTime)) & 0xFFFFFFFF
            except Exception:
                ms = int(ctypes.windll.kernel32.GetTickCount()) - int(lii.dwTime)
            return ms / 1000.0
        return 0.0

    @staticmethod
    def _normalize_process_name(raw: str) -> str:
        name = (raw or "").strip().lower()
        if name.endswith(".exe"):
            name = name[:-4]
        return name

    @classmethod
    def _is_likely_call_title(cls, title: str) -> bool:
        t = (title or "").strip().casefold()
        if not t:
            return False
        if t in {"microsoft teams", "teams", "zoom workplace", "zoom"}:
            return False
        return any(k in t for k in cls.CALL_TITLE_KEYWORDS)

    @classmethod
    def _has_outlook_call_hint(cls, subject: str, location: str, body: str = "") -> bool:
        txt = f"{subject or ''} {location or ''} {body or ''}".casefold()
        if not txt.strip():
            return False
        return any(k in txt for k in cls.OUTLOOK_CALL_KEYWORDS)

    def _is_outlook_meeting_now(self) -> bool:
        if sys.platform != "win32" or win32_client is None:
            return False
        now_m = time.monotonic()
        if now_m - self._outlook_last_probe < self.OUTLOOK_PROBE_INTERVAL_SEC:
            return self._outlook_meeting_now_cache
        self._outlook_last_probe = now_m
        self._outlook_meeting_now_cache = False

        try:
            outlook = win32_client.Dispatch("Outlook.Application")
            ns = outlook.GetNamespace("MAPI")
            cal = ns.GetDefaultFolder(9)  # olFolderCalendar
            items = cal.Items
            items.IncludeRecurrences = True
            items.Sort("[Start]")
            now_dt = datetime.now()
            max_end = now_dt + timedelta(hours=6)
            count = 0
            for appt in items:
                count += 1
                if count > 300:
                    break
                try:
                    st = getattr(appt, "Start", None)
                    en = getattr(appt, "End", None)
                    if not st or not en:
                        continue
                    if st > max_end:
                        break
                    if st <= now_dt <= en:
                        subject = str(getattr(appt, "Subject", "") or "")
                        location = str(getattr(appt, "Location", "") or "")
                        body = str(getattr(appt, "Body", "") or "")
                        if bool(getattr(appt, "IsOnlineMeeting", False)) or self._has_outlook_call_hint(subject, location, body):
                            self._outlook_meeting_now_cache = True
                            return True
                except Exception:
                    continue
        except Exception:
            self._outlook_meeting_now_cache = False
        return self._outlook_meeting_now_cache

    def _running_process_names(self) -> set[str]:
        if psutil is None:
            return set()
        names: set[str] = set()
        try:
            for proc in psutil.process_iter(["name"]):
                nm = self._normalize_process_name(proc.info.get("name") or "")
                if nm:
                    names.add(nm)
        except Exception:
            return names
        return names

    def _active_audio_session_processes(self) -> set[str]:
        if AudioUtilities is None:
            return set()
        active: set[str] = set()
        try:
            for session in AudioUtilities.GetAllSessions():
                proc = getattr(session, "Process", None)
                if proc is None:
                    continue
                try:
                    name = self._normalize_process_name(proc.name())
                except Exception:
                    continue
                if not name:
                    continue
                state = int(getattr(session, "State", 0))
                # 1=Active, 2=Expired op sommige builds; active sessies tellen mee.
                if state == 1:
                    active.add(name)
        except Exception:
            return active
        return active

    def _foreground_context(self) -> tuple[str, str]:
        title = ""
        proc_name = ""
        if gw is not None:
            try:
                w = gw.getActiveWindow()
                if w is not None:
                    title = str(getattr(w, "title", "") or "")
            except Exception:
                pass
        if sys.platform == "win32" and psutil is not None:
            try:
                hwnd = ctypes.windll.user32.GetForegroundWindow()
                if hwnd:
                    pid = wintypes.DWORD()
                    ctypes.windll.user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
                    if pid.value:
                        proc_name = self._normalize_process_name(psutil.Process(int(pid.value)).name())
            except Exception:
                pass
        return title, proc_name

    def _probe_call_now(self) -> bool:
        running = self._running_process_names()
        audio_active = self._active_audio_session_processes()
        fg_title, fg_proc = self._foreground_context()
        fg_has_call = self._is_likely_call_title(fg_title)
        outlook_meeting_now = self._is_outlook_meeting_now()
        has_call_app = any(p in running for p in self.CALL_APP_PROCESSES)
        has_browser = any(p in running for p in self.BROWSER_PROCESSES)
        has_call_audio = any(p in audio_active for p in self.CALL_APP_PROCESSES)
        has_browser_audio = any(p in audio_active for p in self.BROWSER_PROCESSES)
        titles_have_call = False
        detected = False

        # Directe signalen: call-app met actieve audio of foreground-call context.
        if has_call_audio:
            detected = True
        elif fg_has_call:
            if fg_proc in self.CALL_APP_PROCESSES and has_call_audio:
                detected = True
            elif fg_proc in self.BROWSER_PROCESSES and has_browser_audio:
                detected = True
            elif has_call_audio or has_browser_audio:
                detected = True
            elif outlook_meeting_now and has_call_app:
                detected = True

        if not detected and gw is not None:
            try:
                titles = gw.getAllTitles()
            except Exception:
                titles = []
            for t in titles:
                if self._is_likely_call_title(str(t or "")):
                    titles_have_call = True
                    if has_call_audio:
                        detected = True
                        break
                    if has_browser and has_browser_audio:
                        detected = True
                        break
                    if has_call_app:
                        detected = True
                        break
                    # Minder strikt: als call-achtige titel bestaat, telt dit als
                    # actief call-signaal, ook als procesnaam op corporate build
                    # afwijkt van onze bekende lijst.
                    detected = True
                    break
        if not detected and outlook_meeting_now and titles_have_call and has_call_app:
            detected = True

        if self.call_debug_enabled:
            fg_short = (fg_title or "").strip().replace("\n", " ")
            if len(fg_short) > 42:
                fg_short = fg_short[:42] + "..."
            self._call_debug_text = (
                f"call={int(detected)} audio_app={int(has_call_audio)} audio_browser={int(has_browser_audio)} "
                f"proc_call={int(has_call_app)} proc_browser={int(has_browser)} title={int(fg_has_call)} "
                f"titles={int(titles_have_call)} outlook_now={int(outlook_meeting_now)} fg_proc={fg_proc or '-'} "
                f"fg='{fg_short or '-'}'"
            )
        return detected

    def _append_call_debug_log(self):
        if not self.call_debug_log_enabled:
            return
        now = time.monotonic()
        txt = self._call_debug_text or ""
        if txt == self._call_debug_last_text and (now - self._call_debug_last_emit_ts) < 5.0:
            return
        self._call_debug_last_emit_ts = now
        self._call_debug_last_text = txt
        try:
            os.makedirs(os.path.dirname(self._call_debug_log_path) or ".", exist_ok=True)
            with open(self._call_debug_log_path, "a", encoding="utf-8") as f:
                f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | {txt}\n")
        except Exception:
            pass

    def detect_call(self) -> bool:
        now = time.monotonic()
        if self._call_detect_cache and now < self._call_hold_until:
            return True
        if now - self._call_last_probe < self.CALL_PROBE_INTERVAL_SEC:
            return self._call_detect_cache
        self._call_last_probe = now

        detected = self._probe_call_now()
        if detected:
            self._call_hold_until = now + self.CALL_HOLD_SECONDS
            if not self._call_detect_cache:
                if self._call_warmup_start <= 0.0:
                    self._call_warmup_start = now
                if now - self._call_warmup_start >= self.CALL_WARMUP_SECONDS:
                    self._call_detect_cache = True
                    return True
            else:
                self._call_warmup_start = 0.0
        else:
            self._call_warmup_start = 0.0
        if now >= self._call_hold_until:
            self._call_detect_cache = False
        return self._call_detect_cache

    def load_today(self):
        self._confirm_pause_no()
        self.current_day = date.today()
        row = self.store.get_timer_log(self.current_day)
        self.work_seconds = row["work"]
        self.idle_seconds = row["idle"]
        self.call_seconds = row["call"]

    def pause(self):
        self.running = False
        self.timer.stop()
        self.warn_timer.stop()
        self._hide_idle_warning()
        self._idle_episode_active = False
        self._idle_episode_seconds = 0
        self._confirm_pause_no()
        self.persist()

    def start(self):
        if self.running:
            return
        self.running = True
        if self.tracking_enabled:
            self.timer.start()

    def set_idle_threshold(self, seconds: int):
        try:
            sec = int(seconds)
        except Exception:
            sec = self.DEFAULT_IDLE_THRESHOLD_SEC
        self.idle_threshold_sec = max(15, min(3600, sec))

    def _show_idle_warning(self):
        if self.warning_active:
            return
        self.warning_active = True
        self.idle_label.setVisible(True)
        self._sync_panel_height()
        self.warn_timer.start()

    def _hide_idle_warning(self):
        if not self.warning_active:
            return
        self.warning_active = False
        self.warn_timer.stop()
        self.idle_label.setVisible(False)
        self._sync_panel_height()

    def _sync_panel_height(self):
        extra = 0
        if self.idle_label.isVisible():
            extra += 20
        if self.pause_confirm_row.isVisible():
            extra += 30
        total = 56 + extra
        self.setMinimumHeight(total)
        self.setMaximumHeight(total)
        self.height_changed.emit(total)

    def _queue_pause_confirmation(self, seconds: int):
        secs = max(0, int(seconds))
        if secs <= 0:
            return
        self._unconfirmed_pause_seconds += secs
        self.pause_confirm_label.setText(
            f"{seconds_to_hhmmss(self._unconfirmed_pause_seconds)} idle gedetecteerd. Als pauze registreren?"
        )
        self.pause_confirm_row.setVisible(True)
        self.pause_confirm_timer.start()
        self._sync_panel_height()

    def _confirm_pause_yes(self):
        self.pause_confirm_timer.stop()
        if self._unconfirmed_pause_seconds > 0:
            self.idle_seconds += int(self._unconfirmed_pause_seconds)
            self._unconfirmed_pause_seconds = 0
            self.persist()
        self.pause_confirm_row.setVisible(False)
        self.update_ui()
        self._sync_panel_height()

    def _confirm_pause_no(self):
        self.pause_confirm_timer.stop()
        self._unconfirmed_pause_seconds = 0
        self.pause_confirm_row.setVisible(False)
        self.update_ui()
        self._sync_panel_height()

    def set_os_context(self, locked: bool | None = None, sleeping: bool | None = None):
        was_locked = bool(self.os_session_locked)
        if locked is not None:
            self.os_session_locked = bool(locked)
        if sleeping is not None:
            self.os_system_sleeping = bool(sleeping)

        # Bij lock willen we lockduur apart bijhouden; bij unlock tonen we
        # eventuele pauze-confirmatie pas nadat de sessie weer echt actief is.
        if not was_locked and self.os_session_locked:
            self._lock_elapsed_seconds = 0
        elif was_locked and not self.os_session_locked:
            self._lock_elapsed_seconds = 0
            if self.include_lockscreen_idle and self._await_unlock_pause_prompt and self._idle_episode_seconds > 0:
                self._idle_episode_active = False
                self._queue_pause_confirmation(self._idle_episode_seconds)
                self._idle_episode_seconds = 0
            self._await_unlock_pause_prompt = False

        suppress = (self.os_session_locked and not self.include_lockscreen_idle) or (
            self.os_system_sleeping and not self.include_sleep_idle
        )
        if suppress:
            self._hide_idle_warning()
            self._idle_episode_active = False
            self._idle_episode_seconds = 0
        self.update_ui()

    def set_idle_policy(self, include_lockscreen_idle: bool, include_sleep_idle: bool):
        self.include_lockscreen_idle = bool(include_lockscreen_idle)
        self.include_sleep_idle = bool(include_sleep_idle)
        self.set_os_context()

    def _animate_warning(self):
        self.warning_alpha += self.warning_direction
        if self.warning_alpha <= 0.35 or self.warning_alpha >= 1.0:
            self.warning_direction *= -1
        intensity = int(200 + (55 * self.warning_alpha))
        self.idle_label.setStyleSheet(
            f"color:#111111; background:rgb(255,255,{intensity}); font: 10pt 'Consolas'; font-weight:700; border-radius:4px;"
        )

    def _tick(self):
        if not self.running:
            return
        if not self.tracking_enabled:
            return
        if date.today() != self.current_day:
            self.load_today()

        # Timer telt altijd werktijd zolang tracking actief is.
        self.work_seconds += 1

        # Via policy bepaal je expliciet of lock/sleep als pauze-kandidaat meetelt.
        suppress = (self.os_session_locked and not self.include_lockscreen_idle) or (
            self.os_system_sleeping and not self.include_sleep_idle
        )
        if suppress:
            self._hide_idle_warning()
            self._idle_episode_active = False
            self._idle_episode_seconds = 0
            self.save_tick += 1
            if self.save_tick >= self.save_interval:
                self.persist()
                self.save_tick = 0
            self.update_ui()
            return

        # Tijdens lockscreen negeren we lockscreen-input (zoals muisbeweging op
        # inlogscherm). We tellen lockduur zelf en tonen de pauze-popup pas na unlock.
        if self.os_session_locked and self.include_lockscreen_idle:
            self._hide_idle_warning()
            self._lock_elapsed_seconds += 1
            if self._lock_elapsed_seconds >= self.idle_threshold_sec:
                self._idle_episode_active = True
                self._idle_episode_seconds += 1
                self._await_unlock_pause_prompt = True
            self.save_tick += 1
            if self.save_tick >= self.save_interval:
                self.persist()
                self.save_tick = 0
            self.update_ui()
            return

        call_active = self.detect_call()
        if call_active:
            self._hide_idle_warning()
            self._idle_episode_active = False
            self._idle_episode_seconds = 0
            self.call_seconds += 1
            self.save_tick += 1
            if self.save_tick >= self.save_interval:
                self.persist()
                self.save_tick = 0
            self.update_ui()
            return

        idle_sec = self.get_idle_time()
        warning_sec = max(5, min(30, self.idle_threshold_sec // 6))
        if self.idle_threshold_sec - warning_sec <= idle_sec < self.idle_threshold_sec:
            self._show_idle_warning()
        else:
            self._hide_idle_warning()

        is_idle = idle_sec >= self.idle_threshold_sec
        if is_idle:
            self._idle_episode_active = True
            self._idle_episode_seconds += 1
        elif self._idle_episode_active:
            self._idle_episode_active = False
            if self._idle_episode_seconds > 0:
                self._queue_pause_confirmation(self._idle_episode_seconds)
            self._idle_episode_seconds = 0

        self.save_tick += 1
        if self.save_tick >= self.save_interval:
            self.persist()
            self.save_tick = 0
        self.update_ui()

    def persist(self):
        self.store.save_timer_log(self.current_day, self.work_seconds, self.idle_seconds, self.call_seconds)
        self.on_refresh()

    def update_ui(self):
        suffix_parts = []
        if not self.tracking_enabled:
            suffix_parts.append("uit")
        if self.os_session_locked:
            suffix_parts.append("vergrendeld")
        if self.os_system_sleeping:
            suffix_parts.append("slaapstand")
        suffix = f" ({', '.join(suffix_parts)})" if suffix_parts else ""
        self.lbl_line.setText(
            f"T: {seconds_to_hhmmss(self.work_seconds)}  P: {seconds_to_hhmmss(self.idle_seconds)}  C: {seconds_to_hhmmss(self.call_seconds)}{suffix}"
        )
        if self.call_debug_enabled and self.call_debug_label is not None:
            self.call_debug_label.setText(self._call_debug_text)
        self._append_call_debug_log()


# ============================================================================
# DASHBOARD WIDGETS
# Custom widgets voor visualisatie zonder externe chart-library.
# Elk widget tekent zelf om layout, kleurgebruik en performance consistent te
# houden met de rest van de applicatie.
# ============================================================================
class PieChartWidget(QWidget):
    """Compacte pie chart.

    De chart wordt volledig met QPainter gerenderd voor strakke controle over
    labels, percentages en sizing. Hierdoor vermijden we extra chart-packages
    en houden we het dashboard visueel consistent met de rest van de app.
    """

    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.slices: list[tuple[str, float, str]] = []
        self.center_text = "Geen data"
        self.setMinimumSize(180, 180)

    def set_slices(self, slices: list[tuple[str, float, str]], center_text: str):
        self.slices = [(lbl, max(0.0, float(val)), clr) for lbl, val, clr in (slices or [])]
        self.center_text = center_text or ""
        self.update()

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        r = self.rect().adjusted(10, 10, -10, -10)
        p.fillRect(r, QColor(28, 36, 49, 170))
        p.setPen(QPen(QColor(120, 145, 178), 1))
        p.drawRoundedRect(r, 8, 8)

        side = max(60, min(r.width() - 36, r.height() - 36))
        pie_rect = QRect(
            r.center().x() - (side // 2),
            r.center().y() - (side // 2),
            side,
            side,
        )
        total = sum(v for _lbl, v, _clr in self.slices)
        if total <= 0.0:
            p.setPen(QColor("#dce4f0"))
            p.drawText(r, Qt.AlignCenter, "Geen data")
            return

        start = 90 * 16
        for _lbl, value, clr in self.slices:
            if value <= 0:
                continue
            span = int(round((value / total) * 360 * 16))
            p.setBrush(QColor(clr))
            p.setPen(QPen(QColor("#0f1724"), 1))
            p.drawPie(pie_rect, start, -span)

            # Percentage-label in slice (alleen als segment groot genoeg is).
            pct = (value / total) * 100.0
            if pct >= 7.0:
                mid_deg = (start - (span / 2.0)) / 16.0
                radius = (pie_rect.width() / 2.0) * 0.64
                cx = pie_rect.center().x()
                cy = pie_rect.center().y()
                tx = int(cx + (radius * math.cos(math.radians(mid_deg))))
                ty = int(cy - (radius * math.sin(math.radians(mid_deg))))
                txt_rect = QRect(tx - 20, ty - 10, 40, 20)
                p.setPen(QColor("#f8fbff"))
                p.setFont(QFont("Segoe UI", 8, QFont.Bold))
                p.drawText(txt_rect, Qt.AlignCenter, f"{int(round(pct))}%")
            start -= span

        inner = pie_rect.adjusted(int(pie_rect.width() * 0.23), int(pie_rect.height() * 0.23), -int(pie_rect.width() * 0.23), -int(pie_rect.height() * 0.23))
        p.setBrush(QColor("#1f2835"))
        p.setPen(QPen(QColor("#3a4658"), 1))
        p.drawEllipse(inner)
        p.setPen(QColor("#e6edf7"))
        p.setFont(QFont("Segoe UI", 10, QFont.Bold))
        p.drawText(inner, Qt.AlignCenter, self.center_text)


class DashboardPieTile(QGroupBox):
    def __init__(self, title: str, subtitle: str, parent: QWidget | None = None):
        super().__init__(title, parent)
        lay = QVBoxLayout(self)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.lbl_sub = QLabel(subtitle)
        self.lbl_sub.setWordWrap(True)
        self.chart = PieChartWidget()
        self.chart.setMinimumHeight(170)
        self.chart.setMaximumHeight(240)
        self.lbl_legend = QLabel("")
        self.lbl_legend.setWordWrap(True)
        self.lbl_legend.setStyleSheet("color:#d3deea;")
        self.lbl_note = QLabel("")
        self.lbl_note.setWordWrap(True)
        self.lbl_note.setStyleSheet("color:#b8c7d8;")
        lay.addWidget(self.lbl_sub)
        lay.addWidget(self.chart)
        lay.addWidget(self.lbl_legend)
        lay.addWidget(self.lbl_note)

    def set_data(self, slices: list[tuple[str, float, str]], center_text: str, note: str):
        self.chart.set_slices(slices, center_text)
        total = sum(max(0.0, float(v)) for _lbl, v, _clr in (slices or []))
        lines = []
        for lbl, value, clr in slices or []:
            v = max(0.0, float(value))
            pct = int(round((v / total) * 100)) if total > 0 else 0
            lines.append(f"<span style='color:{clr};'>¦</span> {lbl}: {pct}%")
        self.lbl_legend.setText(" | ".join(lines))
        self.lbl_note.setText(note)


class TrendLineWidget(QWidget):
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.points: list[tuple[str, float]] = []
        self.setMinimumHeight(180)

    def set_points(self, points: list[tuple[str, float]]):
        self.points = list(points or [])
        self.update()

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        r = self.rect().adjusted(10, 10, -10, -10)
        p.fillRect(r, QColor(28, 36, 49, 170))
        p.setPen(QPen(QColor(120, 145, 178), 1))
        p.drawRoundedRect(r, 8, 8)
        if len(self.points) < 2:
            p.setPen(QColor("#dce4f0"))
            p.drawText(r, Qt.AlignCenter, "Te weinig data voor trend")
            return
        left = r.left() + 24
        right = r.right() - 12
        top = r.top() + 16
        bottom = r.bottom() - 22
        p.setPen(QPen(QColor(88, 105, 126), 1))
        p.drawLine(left, bottom, right, bottom)
        p.drawLine(left, top, left, bottom)
        values = [v for _l, v in self.points]
        vmax = max(1.0, max(values))
        n = len(self.points)
        pts = []
        for i, (label, v) in enumerate(self.points):
            x = left + int((i / max(1, n - 1)) * (right - left))
            y = bottom - int((v / vmax) * (bottom - top))
            pts.append(QPoint(x, y))
            if n <= 12 or i % max(1, n // 6) == 0:
                p.setPen(QColor("#9eb3cc"))
                p.setFont(QFont("Segoe UI", 7))
                p.drawText(QRect(x - 20, bottom + 2, 40, 16), Qt.AlignCenter, label)
        p.setPen(QPen(QColor("#60a5fa"), 2))
        for i in range(1, len(pts)):
            p.drawLine(pts[i - 1], pts[i])
        p.setBrush(QColor("#93c5fd"))
        p.setPen(QPen(QColor("#dbeafe"), 1))
        for pt in pts:
            p.drawEllipse(pt, 2, 2)


class CategoryBarWidget(QWidget):
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.values: list[tuple[str, float, str]] = []
        self.setMinimumHeight(180)

    def set_values(self, values: list[tuple[str, float, str]]):
        self.values = [(lbl, max(0.0, float(v)), clr) for lbl, v, clr in (values or [])]
        self.update()

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        r = self.rect().adjusted(10, 10, -10, -10)
        p.fillRect(r, QColor(28, 36, 49, 170))
        p.setPen(QPen(QColor(120, 145, 178), 1))
        p.drawRoundedRect(r, 8, 8)
        if not self.values:
            p.setPen(QColor("#dce4f0"))
            p.drawText(r, Qt.AlignCenter, "Geen categorie-data")
            return
        left = r.left() + 18
        right = r.right() - 12
        top = r.top() + 18
        bottom = r.bottom() - 18
        w = max(120, right - left)
        vmax = max(1.0, max(v for _l, v, _c in self.values))
        total = max(1.0, sum(v for _l, v, _c in self.values))
        row_h = max(18, int((bottom - top) / max(1, len(self.values))))
        for i, (label, value, clr) in enumerate(self.values):
            y = top + (i * row_h)
            p.setPen(QColor("#c8d6e6"))
            p.setFont(QFont("Segoe UI", 8))
            p.drawText(QRect(left, y, 120, row_h - 2), Qt.AlignVCenter | Qt.AlignLeft, label)
            bar_x = left + 120
            bar_w = int(((w - 190) * (value / vmax)))
            p.setBrush(QColor(clr))
            p.setPen(Qt.NoPen)
            p.drawRoundedRect(QRect(bar_x, y + 4, max(2, bar_w), row_h - 8), 4, 4)
            p.setPen(QColor("#e6edf7"))
            pct = int(round((value / total) * 100))
            p.drawText(QRect(bar_x + bar_w + 6, y, 58, row_h), Qt.AlignVCenter | Qt.AlignLeft, f"{pct}%")


class DashboardTrendTile(QGroupBox):
    def __init__(self, title: str, subtitle: str, parent: QWidget | None = None):
        super().__init__(title, parent)
        lay = QVBoxLayout(self)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.lbl_sub = QLabel(subtitle)
        self.lbl_sub.setWordWrap(True)
        self.chart = TrendLineWidget()
        self.lbl_note = QLabel("")
        self.lbl_note.setWordWrap(True)
        self.lbl_note.setStyleSheet("color:#b8c7d8;")
        lay.addWidget(self.lbl_sub)
        lay.addWidget(self.chart, 1)
        lay.addWidget(self.lbl_note)

    def set_data(self, points: list[tuple[str, float]], note: str):
        self.chart.set_points(points)
        self.lbl_note.setText(note)


class DashboardBarTile(QGroupBox):
    def __init__(self, title: str, subtitle: str, parent: QWidget | None = None):
        super().__init__(title, parent)
        lay = QVBoxLayout(self)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.lbl_sub = QLabel(subtitle)
        self.lbl_sub.setWordWrap(True)
        self.chart = CategoryBarWidget()
        self.lbl_note = QLabel("")
        self.lbl_note.setWordWrap(True)
        self.lbl_note.setStyleSheet("color:#b8c7d8;")
        lay.addWidget(self.lbl_sub)
        lay.addWidget(self.chart, 1)
        lay.addWidget(self.lbl_note)

    def set_data(self, values: list[tuple[str, float, str]], note: str):
        self.chart.set_values(values)
        self.lbl_note.setText(note)


# ============================================================================
# MAIN APP ORCHESTRATION
# MainWindow verbindt opslag, kalenderweergaven, dialogs en dashboard.
# De klasse is groot omdat dit project momenteel single-file is; methodes zijn
# thematisch gegroepeerd om toekomstige opsplitsing naar modules eenvoudiger te maken.
# ============================================================================
class MainWindow(QMainWindow):
    """Hoofdvenster.

    Bevat:
    - kalenderweergaven (jaar + maand),
    - timerdrawer,
    - instellingen/menu.
    """
    def __init__(self):
        super().__init__()
        self.year = datetime.now().year
        base_dir = os.path.dirname(os.path.abspath(__file__))
        self.base_dir = base_dir
        self.store = ExcelStore(self.year, base_dir)
        self.mode = "worked"
        self.dark_mode = True
        (
            self.worked_colors,
            self.planned_colors,
            self.extra_info_colors,
            self.idle_threshold_sec,
            self.include_lockscreen_idle,
            self.include_sleep_idle,
            self.extra_info_enabled,
            self.school_region,
            self.inactive_glass_opacity,
            self.min_inactive_glass_opacity,
            self.max_inactive_glass_opacity,
            self.timer_enabled,
        ) = self.load_color_settings()
        self.min_inactive_glass_opacity = GLASS_OPACITY_MIN
        self.max_inactive_glass_opacity = GLASS_OPACITY_MAX
        self.inactive_glass_opacity = max(GLASS_OPACITY_MIN, min(GLASS_OPACITY_MAX, self.inactive_glass_opacity))
        self.ensure_extra_info_colors()
        self.ensure_extra_info_enabled()
        self.school_region = (self.school_region or "zuid").strip().casefold()
        if self.school_region not in {"noord", "midden", "zuid"}:
            self.school_region = "zuid"
        self.store.school_region = self.school_region
        self.timer_host = None
        self._drag_active = False
        self._drag_offset = QPoint()
        self._allow_close = False
        self._slide_anim = None
        self._sliding = False
        self._planner_target_geometry = QRect(120, 80, 2140, 1050)
        self.setWindowTitle(f"Tijdplanner Pro - {self.year}")
        self.setWindowFlags(Qt.Window | Qt.FramelessWindowHint)
        # Start met drie maanden waar mogelijk, maar schaal veilig terug op kleinere schermen.
        self._planner_target_geometry = self._fit_rect_to_screen(self._planner_target_geometry)
        self.resize(self._planner_target_geometry.width(), self._planner_target_geometry.height())
        self.move(self._planner_target_geometry.topLeft())
        self.apply_theme()
        self.setup_ui()
        self.menuBar().installEventFilter(self)
        self.refresh_all()

    def _fit_rect_to_screen(self, rect: QRect, screen=None, margin: int = 18) -> QRect:
        if screen is None:
            screen = QApplication.screenAt(rect.center())
            if screen is None:
                screen = QApplication.primaryScreen()
        avail = screen.availableGeometry() if screen else QRect(0, 0, 2560, 1440)

        max_w = max(360, avail.width() - (margin * 2))
        max_h = max(260, avail.height() - (margin * 2))

        w = min(rect.width(), max_w)
        h = min(rect.height(), max_h)

        x_min = avail.left() + margin
        y_min = avail.top() + margin
        x_max = avail.right() - margin - w + 1
        y_max = avail.bottom() - margin - h + 1

        x = max(x_min, min(rect.x(), x_max))
        y = max(y_min, min(rect.y(), y_max))
        return QRect(x, y, w, h)

    def apply_theme(self):
        self.setStyleSheet(
            """
            QMainWindow { background: #1c212b; }
            QMainWindow[modeTheme="worked"] { background: #1c212b; }
            QMainWindow[modeTheme="planned"] { background: #1c212b; }
            QMenuBar { background: #121720; color: #e8edf7; padding: 5px; font: 10pt "Segoe UI"; }
            QMenuBar::item:selected { background: #355071; color:#ffffff; border-radius: 4px; }
            QMainWindow[modeTheme="worked"] QMenuBar::item:selected { background: #355071; }
            QMenu { background: #1d2532; color: #e8edf7; border: 1px solid #344156; }
            QMenu::item { padding: 6px 20px; border-radius: 4px; }
            QMenu::item:selected { background: #355071; color: #ffffff; }
            QMainWindow[modeTheme="worked"] QMenu::item:selected { background: #355071; }
            QDialog, QMessageBox { background: #1c2430; color: #e6edf3; }
            QToolTip { color: #e6edf3; background-color: #10151d; border: 1px solid #3b4a60; padding: 6px; }
            QToolBar { background: #232c3a; border: 0; spacing: 8px; padding: 6px; }
            QMainWindow[modeTheme="worked"] QToolBar { background: #232c3a; }
            QPushButton { background: #3f6ea6; color: #fff; border: 0; border-radius: 8px; padding: 8px 14px; font: 10pt "Segoe UI Semibold"; }
            QPushButton:hover { background: #4e7fba; }
            QPushButton:pressed { background: #325987; }
            QMainWindow[modeTheme="worked"] QPushButton { background: #3f6ea6; }
            QMainWindow[modeTheme="worked"] QPushButton:hover { background: #4e7fba; }
            QMainWindow[modeTheme="worked"] QPushButton:pressed { background: #325987; }
            QLabel#modeLeft, QLabel#modeRight { color:#9fb0c5; font: 10pt "Segoe UI Semibold"; padding: 0 4px; }
            QLabel#modeLeft[active="true"] { color:#9ecbff; }
            QLabel#modeRight[active="true"] { color:#9ecbff; }
            QCheckBox#modeSwitch { spacing:0px; }
            QCheckBox#modeSwitch::indicator {
                width:56px; height:28px; border-radius:14px;
                border:1px solid #32465f;
                background:qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #4d6f96, stop:1 #2f4f75);
            }
            QCheckBox#modeSwitch::indicator:unchecked {
                border-top:1px solid #a8c6e7;
                border-left:1px solid #9bb8d7;
                border-right:1px solid #324a62;
                border-bottom:2px solid #213344;
            }
            QCheckBox#modeSwitch::indicator:checked {
                border-top:1px solid #a8c6e7;
                border-left:1px solid #9bb8d7;
                border-right:1px solid #324a62;
                border-bottom:2px solid #213344;
                background:qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #4d6f96, stop:1 #2f4f75);
            }
            QFrame#timerPanel QPushButton[active="true"] { background:#355071; border:1px solid #4c739f; }
            QMainWindow[modeTheme="worked"] QFrame#timerPanel QPushButton[active="true"] { background:#355071; border:1px solid #4c739f; }
            QTabWidget::pane { border: 1px solid #3a4658; background: #222b38; border-radius: 8px; }
            QMainWindow[modeTheme="worked"] QTabWidget::pane { border: 1px solid #3a4658; background: #222b38; }
            QTabBar::tab { background: #2a3443; color: #cfd8e6; padding: 10px 18px; margin-right: 4px; border-top-left-radius: 8px; border-top-right-radius: 8px; border: 1px solid #3a4658; border-bottom: 0; font: 10pt "Segoe UI Semibold"; }
            QMainWindow[modeTheme="worked"] QTabBar::tab { background: #2a3443; color: #cfd8e6; border: 1px solid #3a4658; border-bottom: 0; }
            QTabBar::tab:selected { background: #355071; color: #f6fbff; border: 1px solid #5b84b5; border-bottom: 0; font: 10pt "Segoe UI Black"; }
            QMainWindow[modeTheme="worked"] QTabBar::tab:selected { background: #355071; color: #f6fbff; border: 1px solid #5b84b5; border-bottom: 0; }
            QTabBar#workedMonthTabBar::tab, QTabBar#plannedMonthTabBar::tab { background: #283240; margin-right: 3px; border: 1px solid #3a4658; border-bottom: 0; }
            QMainWindow[modeTheme="worked"] QTabBar#workedMonthTabBar::tab, QMainWindow[modeTheme="worked"] QTabBar#plannedMonthTabBar::tab { background: #283240; border: 1px solid #3a4658; border-bottom: 0; }
            QTabBar#workedMonthTabBar::tab:selected, QTabBar#plannedMonthTabBar::tab:selected { background: #355071; color: #f3f8ff; border: 1px solid #5b84b5; border-bottom: 0; font: 10pt "Segoe UI Black"; }
            QMainWindow[modeTheme="worked"] QTabBar#workedMonthTabBar::tab:selected, QMainWindow[modeTheme="worked"] QTabBar#plannedMonthTabBar::tab:selected { background: #355071; color: #f3f8ff; border: 1px solid #5b84b5; border-bottom: 0; }
            QScrollArea#calendarScroll { background: #222b38; border: 0; }
            QScrollArea#calendarScroll[modeTheme="worked"] { background: #222b38; }
            QScrollArea#calendarScroll[modeTheme="planned"] { background: #222b38; }
            QWidget#calendarHost { background: #222b38; }
            QWidget#calendarHost[modeTheme="worked"] { background: #222b38; }
            QWidget#calendarHost[modeTheme="planned"] { background: #222b38; }
            QScrollArea#dashDetailScroll { background: #222b38; border: 1px solid #3a4658; border-radius: 8px; }
            QWidget#dashDetailHost { background: #263141; }
            QGroupBox { border: 1px solid #3b4759; border-radius: 10px; margin-top: 18px; padding-top: 12px; background: #263141; font: 10pt "Segoe UI Semibold"; color: #e0e7f2; }
            QMainWindow[modeTheme="worked"] QGroupBox { border: 1px solid #3b4759; background: #263141; color: #e0e7f2; }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; top: -2px; padding: 2px 8px; }
            QTableWidget { background: #1f2835; alternate-background-color: #222e3d; border: 1px solid #3a4658; gridline-color: #3a4658; color: #e6edf7; font: 9pt "Segoe UI"; }
            QMainWindow[modeTheme="worked"] QTableWidget { background: #1f2835; alternate-background-color: #222e3d; border: 1px solid #3a4658; gridline-color: #3a4658; color: #e6edf7; }
            QTableWidget#dashTableDetail::item { background: transparent; }
            QTableWidget::item:selected { background: #355071; color: #ffffff; }
            QHeaderView::section { background: #2c3849; color: #dce4f0; padding: 6px; border: 1px solid #425066; font: 9pt "Segoe UI Semibold"; }
            QMainWindow[modeTheme="worked"] QHeaderView::section { background: #2c3849; color: #dce4f0; border: 1px solid #425066; }
            QStatusBar { background: #232c3a; color: #d8e1ef; border-top: 1px solid #3a4658; font: 9pt "Segoe UI"; }
            QMainWindow[modeTheme="worked"] QStatusBar { background: #232c3a; border-top: 1px solid #3a4658; }
            QLabel { color: #dce4f0; }
            QMainWindow[modeTheme="worked"] QLabel { color: #dce4f0; }
            QCheckBox { color: #dce4f0; spacing: 6px; }
            QMainWindow[modeTheme="worked"] QCheckBox { color: #dce4f0; }
            QLineEdit, QComboBox, QDateEdit { border: 1px solid #45556d; border-radius: 6px; background: #1f2835; color: #eaf1fb; padding: 6px; font: 9pt "Segoe UI"; }
            QMainWindow[modeTheme="worked"] QLineEdit, QMainWindow[modeTheme="worked"] QComboBox, QMainWindow[modeTheme="worked"] QDateEdit { border: 1px solid #45556d; background: #1f2835; color: #eaf1fb; }
            QComboBox::drop-down { border: 0; width: 24px; }
            QComboBox::down-arrow { width: 10px; height: 10px; }
            QComboBox QAbstractItemView { background: #1f2835; color: #eaf1fb; border: 1px solid #45556d; selection-background-color: #355071; selection-color: #ffffff; outline: 0; }
            QMainWindow[modeTheme="worked"] QComboBox QAbstractItemView { background: #1f2835; color: #eaf1fb; border: 1px solid #45556d; selection-background-color: #355071; }
            QDateEdit::drop-down { border: 0; width: 24px; }
            QDateEdit::down-arrow { width: 10px; height: 10px; }
            QCalendarWidget QWidget { background: #1f2835; color: #eaf1fb; }
            QCalendarWidget QToolButton { color: #eaf1fb; background: #2a3443; border: 1px solid #45556d; border-radius: 4px; }
            QCalendarWidget QToolButton:hover { background: #355071; }
            QCalendarWidget QMenu { background: #1f2835; color: #eaf1fb; }
            QScrollBar:vertical { background: #1a2230; width: 12px; margin: 2px; }
            QMainWindow[modeTheme="worked"] QScrollBar:vertical { background: #1a2230; }
            QScrollBar::handle:vertical { background: #40516b; border-radius: 5px; min-height: 28px; }
            QMainWindow[modeTheme="worked"] QScrollBar::handle:vertical { background: #40516b; }
            QScrollBar:horizontal { background: #1a2230; height: 12px; margin: 2px; }
            QMainWindow[modeTheme="worked"] QScrollBar:horizontal { background: #1a2230; }
            QScrollBar::handle:horizontal { background: #40516b; border-radius: 5px; min-width: 28px; }
            QMainWindow[modeTheme="worked"] QScrollBar::handle:horizontal { background: #40516b; }
            """
        )

    def setup_ui(self):
        self.build_menu()
        self.add_window_controls()
        self.build_toolbar()
        self.build_statusbar()

        root = QWidget()
        self.setCentralWidget(root)
        content_lay = QVBoxLayout(root)
        content_lay.setContentsMargins(6, 6, 6, 6)

        self.tabs = QTabWidget()
        self.tabs.tabBar().setExpanding(False)
        content_lay.addWidget(self.tabs)
        self.tab_worked = QWidget()
        self.tab_planned = QWidget()
        self.tab_dashboard = QWidget()
        self.tabs.addTab(self.tab_worked, "Uren")
        self.tabs.addTab(self.tab_planned, "Planning")
        self.tabs.addTab(self.tab_dashboard, "Dashboard")

        worked_layout = QVBoxLayout(self.tab_worked)
        self.worked_month_tabbar = QTabBar()
        self.worked_month_tabbar.setObjectName("workedMonthTabBar")
        self.worked_month_tabbar.setDrawBase(False)
        self.worked_month_tabbar.setExpanding(False)
        self.worked_month_tabbar.setUsesScrollButtons(True)
        self.worked_month_tabbar.setElideMode(Qt.ElideNone)
        self.worked_month_tabbar.addTab("Jaaroverzicht")
        for m in MONTHS_NL:
            self.worked_month_tabbar.addTab(m)
        self.worked_month_tabbar.setCurrentIndex(0)
        self.worked_month_tabbar.setMaximumHeight(36)
        self.worked_month_tabbar.installEventFilter(self)
        self.worked_month_tabbar.currentChanged.connect(self.on_worked_period_tab_changed)
        worked_layout.addWidget(self.worked_month_tabbar)
        self.worked_year_board = CalendarBoard(
            self.store,
            self.year,
            "worked",
            list(range(1, 13)),
            3,
            focus_mode=False,
            dark_mode=self.dark_mode,
            colors=self.worked_colors,
            extra_info_colors=self.extra_info_colors,
        )
        self.worked_year_board.installEventFilter(self)
        self.worked_year_board.scroll.installEventFilter(self)
        self.worked_year_board.host.installEventFilter(self)
        self.worked_year_board.day_double_clicked.connect(self.edit_day)
        self.worked_year_board.day_range_selected.connect(self.open_bulk_planning)
        worked_layout.addWidget(self.worked_year_board, 1)
        self.worked_board = CalendarBoard(
            self.store,
            self.year,
            "worked",
            [max(1, date.today().month)],
            1,
            focus_mode=True,
            dark_mode=self.dark_mode,
            colors=self.worked_colors,
            extra_info_colors=self.extra_info_colors,
        )
        self.worked_board.installEventFilter(self)
        self.worked_board.scroll.installEventFilter(self)
        self.worked_board.host.installEventFilter(self)
        self.worked_board.day_double_clicked.connect(self.edit_day)
        self.worked_board.day_range_selected.connect(self.open_bulk_planning)
        self.worked_board.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        worked_layout.addWidget(self.worked_board, 1)
        self.worked_board.setVisible(False)

        planned_layout = QVBoxLayout(self.tab_planned)
        self.planned_month_tabbar = QTabBar()
        self.planned_month_tabbar.setObjectName("plannedMonthTabBar")
        self.planned_month_tabbar.setDrawBase(False)
        self.planned_month_tabbar.setExpanding(False)
        self.planned_month_tabbar.setUsesScrollButtons(True)
        self.planned_month_tabbar.setElideMode(Qt.ElideNone)
        self.planned_month_tabbar.addTab("Jaaroverzicht")
        for m in MONTHS_NL:
            self.planned_month_tabbar.addTab(m)
        self.planned_month_tabbar.setCurrentIndex(0)
        self.planned_month_tabbar.setMaximumHeight(36)
        self.planned_month_tabbar.installEventFilter(self)
        self.planned_month_tabbar.currentChanged.connect(self.on_planned_period_tab_changed)
        planned_layout.addWidget(self.planned_month_tabbar)
        self.planned_year_board = CalendarBoard(
            self.store,
            self.year,
            "planned",
            list(range(1, 13)),
            3,
            focus_mode=False,
            dark_mode=self.dark_mode,
            colors=self.planned_colors,
            extra_info_colors=self.extra_info_colors,
        )
        self.planned_year_board.installEventFilter(self)
        self.planned_year_board.scroll.installEventFilter(self)
        self.planned_year_board.host.installEventFilter(self)
        self.planned_year_board.day_double_clicked.connect(self.edit_day)
        self.planned_year_board.day_range_selected.connect(self.open_bulk_planning)
        planned_layout.addWidget(self.planned_year_board, 1)
        self.planned_board = CalendarBoard(
            self.store,
            self.year,
            "planned",
            [max(1, date.today().month)],
            1,
            focus_mode=True,
            dark_mode=self.dark_mode,
            colors=self.planned_colors,
            extra_info_colors=self.extra_info_colors,
        )
        self.planned_board.installEventFilter(self)
        self.planned_board.scroll.installEventFilter(self)
        self.planned_board.host.installEventFilter(self)
        self.planned_board.day_double_clicked.connect(self.edit_day)
        self.planned_board.day_range_selected.connect(self.open_bulk_planning)
        self.planned_board.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        planned_layout.addWidget(self.planned_board, 1)
        self.planned_board.setVisible(False)

        dash_layout = QVBoxLayout(self.tab_dashboard)
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("Snelfilter"))
        self.dash_period = QComboBox()
        self.dash_period.addItems(["Huidige maand", "Laatste 30 dagen", "Dit jaar", "Aangepast"])
        filter_row.addWidget(self.dash_period)
        filter_row.addWidget(QLabel("Begin"))
        self.dash_from = QLineEdit()
        self.dash_from.setPlaceholderText("dd-mm-jjjj")
        filter_row.addWidget(self.dash_from)
        filter_row.addWidget(QLabel("Eind"))
        self.dash_to = QLineEdit()
        self.dash_to.setPlaceholderText("dd-mm-jjjj")
        filter_row.addWidget(self.dash_to)
        self.dash_include_weekend = QCheckBox("Weekend meenemen")
        self.dash_include_weekend.setChecked(True)
        self.dash_include_weekend.setStyleSheet("color:#dce4f0;")
        filter_row.addWidget(self.dash_include_weekend)
        self.btn_dash_refresh = QPushButton("Vernieuwen")
        filter_row.addWidget(self.btn_dash_refresh)
        filter_row.addStretch(1)
        dash_layout.addLayout(filter_row)

        kpi_box = QGroupBox("Kerncijfers")
        kpi_box.setMinimumHeight(130)
        kpi_box.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        kpi_grid = QGridLayout(kpi_box)
        kpi_grid.setContentsMargins(12, 12, 12, 10)
        kpi_grid.setHorizontalSpacing(18)
        kpi_grid.setVerticalSpacing(8)
        self.lbl_dash_work = QLabel("Totaal gewerkt: 00:00")
        self.lbl_dash_planned = QLabel("Totaal gepland: 00:00")
        self.lbl_dash_idle = QLabel("Totaal pauze: 00:00")
        self.lbl_dash_call = QLabel("Totaal call: 00:00")
        self.lbl_dash_ratio = QLabel("Productief ratio: 0%")
        self.lbl_dash_days = QLabel("Actieve dagen: 0")
        self.lbl_dash_avg = QLabel("Gemiddeld/dag: 00:00")
        self.lbl_dash_variance = QLabel("Afwijking gepland vs gewerkt: 00:00")
        self.lbl_dash_budget = QLabel("Contract: 00:00 | Nog te plannen: 00:00 | Vrij over: 00:00")
        # Layout volgens dashboard best-practice: KPI's compact bovenin, charts daaronder.
        kpi_grid.addWidget(self.lbl_dash_work, 0, 0)
        kpi_grid.addWidget(self.lbl_dash_planned, 0, 1)
        kpi_grid.addWidget(self.lbl_dash_idle, 0, 2)
        kpi_grid.addWidget(self.lbl_dash_call, 0, 3)
        kpi_grid.addWidget(self.lbl_dash_ratio, 1, 0)
        kpi_grid.addWidget(self.lbl_dash_days, 1, 1)
        kpi_grid.addWidget(self.lbl_dash_avg, 1, 2)
        kpi_grid.addWidget(self.lbl_dash_variance, 1, 3)
        kpi_grid.addWidget(self.lbl_dash_budget, 2, 0, 1, 4)
        for c in range(4):
            kpi_grid.setColumnStretch(c, 1)

        dash_body = QHBoxLayout()
        left_host = QWidget()
        left_lay = QVBoxLayout(left_host)
        left_lay.setContentsMargins(0, 0, 0, 0)
        left_lay.setSpacing(8)

        tile_grid = QGridLayout()
        tile_grid.setHorizontalSpacing(10)
        tile_grid.setVerticalSpacing(10)
        self.tile_mix = DashboardPieTile("Urenmix", "Verdeling van werk-, pauze- en calltijd binnen de gekozen periode.")
        self.tile_mix2 = DashboardPieTile("Planningbalans", "Inzicht in boven planning, onder planning en dagen op schema.")
        self.tile_cat = DashboardBarTile("Plan vs Werk", "Afwijking ten opzichte van planning in de gekozen periode.")
        self.tile_mix.setMinimumHeight(250)
        self.tile_mix2.setMinimumHeight(250)
        self.tile_cat.setMinimumHeight(520)
        self.tile_mix.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.tile_mix2.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.tile_cat.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        tile_grid.addWidget(self.tile_mix, 0, 0)
        tile_grid.addWidget(self.tile_mix2, 1, 0)
        tile_grid.addWidget(self.tile_cat, 0, 1, 2, 1)
        tile_grid.setColumnStretch(0, 1)
        tile_grid.setColumnStretch(1, 1)
        left_lay.addWidget(kpi_box)
        left_lay.addLayout(tile_grid, 1)
        left_host.setLayout(left_lay)

        self.health_box = QGroupBox("Data Health")
        health_lay = QGridLayout(self.health_box)
        health_lay.setContentsMargins(10, 10, 10, 10)
        health_lay.setHorizontalSpacing(10)
        health_lay.setVerticalSpacing(6)
        self.lbl_health_save = QLabel("Laatste save: onbekend")
        self.lbl_health_backup = QLabel("Laatste backup: onbekend")
        self.lbl_health_snapshots = QLabel("Snapshots: 0")
        self.lbl_health_status = QLabel("Status: onbekend")
        health_lay.addWidget(self.lbl_health_save, 0, 0)
        health_lay.addWidget(self.lbl_health_backup, 0, 1)
        health_lay.addWidget(self.lbl_health_snapshots, 1, 0)
        health_lay.addWidget(self.lbl_health_status, 1, 1)

        self.dash_table = QTableWidget(0, 6)
        self.dash_table.setObjectName("dashTableDetail")
        self.dash_table.setHorizontalHeaderLabels(["Datum", "Gepland", "Gewerkt", "Afwijking", "Pauze", "Call"])
        self.dash_table.verticalHeader().setVisible(False)
        self.dash_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.dash_table.setSelectionMode(QTableWidget.NoSelection)
        self.dash_table.setAlternatingRowColors(True)
        self.dash_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        for c in (1, 2, 3, 4, 5):
            self.dash_table.horizontalHeader().setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self.dash_table.setMinimumHeight(440)
        self.dash_table.setStyleSheet(
            "QTableWidget { background:#1f2835; alternate-background-color:#222e3d; color:#e6edf7; } "
            "QTableWidget::item { background: transparent; } "
            "QTableWidget QWidget { background:#1f2835; color:#e6edf7; }"
        )
        right_host = QWidget()
        right_host.setObjectName("dashDetailHost")
        right_lay = QVBoxLayout(right_host)
        right_lay.setContentsMargins(6, 6, 6, 6)
        right_lay.setSpacing(8)
        right_lay.addWidget(self.health_box)
        right_lay.addWidget(QLabel("Detaildata (scrollbaar)"))
        right_lay.addWidget(self.dash_table, 1)
        self.dash_detail_scroll = QScrollArea()
        self.dash_detail_scroll.setObjectName("dashDetailScroll")
        self.dash_detail_scroll.setWidgetResizable(True)
        self.dash_detail_scroll.setWidget(right_host)
        self.dash_detail_scroll.setMinimumWidth(420)
        self.dash_detail_scroll.viewport().setStyleSheet("background:#222b38;")
        self.dash_table.viewport().setStyleSheet("background:#1f2835;")

        dash_body.addWidget(left_host, 3)
        dash_body.addWidget(self.dash_detail_scroll, 2)
        dash_layout.addLayout(dash_body, 1)

        self.dash_period.currentIndexChanged.connect(self.on_dashboard_period_changed)
        self.dash_from.editingFinished.connect(self.refresh_dashboard)
        self.dash_to.editingFinished.connect(self.refresh_dashboard)
        self.dash_include_weekend.toggled.connect(lambda _: self.refresh_dashboard())
        self.btn_dash_refresh.clicked.connect(self.refresh_dashboard)
        self.on_dashboard_period_changed(self.dash_period.currentIndex())
        self.refresh_dashboard()
        self._apply_current_month_accent()
        self.tabs.setCurrentIndex(0)
        self.tabs.currentChanged.connect(self.on_tab_change)

    def slide_in_from(self, timer_geo: QRect):
        target = QRect(self._planner_target_geometry)
        if target.width() < 600 or target.height() < 400:
            target = QRect(120, 80, 2140, 1050)
        target = self._fit_rect_to_screen(target, QApplication.screenAt(timer_geo.center()))

        start = QRect(timer_geo.x(), timer_geo.y(), target.width(), target.height())
        self.setGeometry(start)
        self.setWindowOpacity(0.0)
        self.show()

        geo_anim = QPropertyAnimation(self, b"geometry")
        geo_anim.setDuration(320)
        geo_anim.setStartValue(start)
        geo_anim.setEndValue(target)
        geo_anim.setEasingCurve(QEasingCurve.OutCubic)

        opacity_anim = QPropertyAnimation(self, b"windowOpacity")
        opacity_anim.setDuration(220)
        opacity_anim.setStartValue(0.0)
        opacity_anim.setEndValue(1.0)
        opacity_anim.setEasingCurve(QEasingCurve.OutQuad)

        group = QParallelAnimationGroup(self)
        group.addAnimation(geo_anim)
        group.addAnimation(opacity_anim)
        self._slide_anim = group
        self._sliding = True

        def done():
            self._sliding = False
            self.setGeometry(target)
            self.setWindowOpacity(1.0)
            self._planner_target_geometry = QRect(target)

        group.finished.connect(done)
        group.start()

    def slide_out_to(self, timer_geo: QRect):
        start = self.geometry()
        if start.width() > 300 and start.height() > 200:
            self._planner_target_geometry = QRect(start)

        end = QRect(timer_geo.x(), timer_geo.y(), start.width(), start.height())

        geo_anim = QPropertyAnimation(self, b"geometry")
        geo_anim.setDuration(280)
        geo_anim.setStartValue(start)
        geo_anim.setEndValue(end)
        geo_anim.setEasingCurve(QEasingCurve.InOutCubic)

        opacity_anim = QPropertyAnimation(self, b"windowOpacity")
        opacity_anim.setDuration(210)
        opacity_anim.setStartValue(1.0)
        opacity_anim.setEndValue(0.0)
        opacity_anim.setEasingCurve(QEasingCurve.InQuad)

        group = QParallelAnimationGroup(self)
        group.addAnimation(geo_anim)
        group.addAnimation(opacity_anim)
        self._slide_anim = group
        self._sliding = True

        def done():
            self._sliding = False
            self.hide()
            self._planner_target_geometry = self._fit_rect_to_screen(self._planner_target_geometry)
            self.setGeometry(self._planner_target_geometry)
            self.setWindowOpacity(1.0)

        group.finished.connect(done)
        group.start()

    def build_menu(self):
        m = self.menuBar()
        menu_file = m.addMenu("Bestand")
        act_save = QAction("Opslaan", self)
        act_export = QAction("Exporteer kopie", self)
        act_exit = QAction("Afsluiten", self)
        act_save.triggered.connect(self.save)
        act_export.triggered.connect(self.export_copy)
        act_exit.triggered.connect(self.quit_app)
        menu_file.addAction(act_save)
        menu_file.addAction(act_export)
        menu_file.addSeparator()
        menu_file.addAction(act_exit)

        menu_view = m.addMenu("Beeld")
        menu_view.addAction("Uren", lambda: self.tabs.setCurrentIndex(0))
        menu_view.addAction("Planning", lambda: self.tabs.setCurrentIndex(1))
        menu_view.addAction("Dashboard", lambda: self.tabs.setCurrentIndex(2))
        menu_view.addSeparator()
        menu_view.addAction("Ga naar huidige maand", self.goto_today)

        menu_plan = m.addMenu("Planning")
        menu_plan.addAction("Werkpatroon bewerken", self.open_work_pattern_editor)
        menu_plan.addAction("Dienstverband en jaarbudget", self.open_contract_budget_settings)
        menu_plan.addAction("Bulk plannen", lambda: self.open_bulk_planning())
        menu_settings = m.addMenu("Instellingen")
        menu_settings.addAction("Kleuren", self.open_color_settings)
        menu_settings.addAction("Aanvullende info opties", self.open_extra_info_settings)
        menu_settings.addAction("Schoolvakantie regio", self.open_school_region_settings)
        menu_settings.addAction("Timer idle detectie", self.open_timer_settings)
        self.act_timer_enabled = QAction("Timer uren meten", self)
        self.act_timer_enabled.setCheckable(True)
        self.act_timer_enabled.setChecked(bool(self.timer_enabled))
        self.act_timer_enabled.triggered.connect(self.on_timer_enabled_toggled)
        menu_settings.addAction(self.act_timer_enabled)
        menu_settings.addAction("Doorzichtigheid", self.open_glass_settings)
        menu_help = m.addMenu("Help")
        menu_help.addAction("Over", self.about)

    def add_window_controls(self):
        host = QWidget()
        row = QHBoxLayout(host)
        row.setContentsMargins(0, 0, 4, 0)
        row.setSpacing(4)

        btn_min = QPushButton("-")
        btn_close = QPushButton("×")
        for b in (btn_min, btn_close):
            b.setFixedSize(28, 24)
            b.setStyleSheet(
                "QPushButton { background:#2a3443; border:1px solid #3a4658; border-radius:4px; color:#dce4f0; padding:0px; }"
                "QPushButton:hover { background:#3a475b; }"
            )
        btn_close.setStyleSheet(
            "QPushButton { background:#4a2a30; border:1px solid #6a3d44; border-radius:4px; color:#ffdce0; padding:0px; }"
            "QPushButton:hover { background:#6a303a; }"
        )

        btn_min.clicked.connect(self.showMinimized)
        btn_close.clicked.connect(self.close)

        row.addWidget(btn_min)
        row.addWidget(btn_close)
        self.menuBar().setCornerWidget(host, Qt.TopRightCorner)

    def build_toolbar(self):
        tb = QToolBar("Main")
        tb.setMovable(False)
        self.addToolBar(Qt.TopToolBarArea, tb)

        self.btn_pattern = QPushButton("Werkpatroon toepassen")
        self.btn_pattern.clicked.connect(self.open_work_pattern_editor)
        self.btn_pattern.setVisible(False)
        tb.addWidget(self.btn_pattern)
        self.btn_pattern_edit = QPushButton("Werkpatroon")
        self.btn_pattern_edit.clicked.connect(self.open_work_pattern_editor)
        self.btn_pattern_edit.setVisible(False)
        tb.addWidget(self.btn_pattern_edit)
        spacer = QWidget()
        spacer.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        tb.addWidget(spacer)
        self._refresh_mode_toggle_style()

    def build_statusbar(self):
        sb = QStatusBar()
        self.setStatusBar(sb)
        self.lbl_status = QLabel("Gereed")
        self.lbl_saldo = QLabel("")
        sb.addPermanentWidget(self.lbl_status, 1)
        sb.addPermanentWidget(self.lbl_saldo)

    def refresh_all(self):
        # We verversen bewust alleen de zichtbare kalendersectie voor performance.
        # Beide jaar- en maandvarianten telkens opnieuw renderen is merkbaar zwaarder bij grote tabellen
        # en geeft geen functionele winst zolang de gebruiker die sectie niet ziet.
        self.worked_year_board.set_dark_mode(self.dark_mode)
        self.planned_year_board.set_dark_mode(self.dark_mode)
        self.worked_board.set_dark_mode(self.dark_mode)
        self.planned_board.set_dark_mode(self.dark_mode)
        self.worked_year_board.set_colors(self.worked_colors)
        self.planned_year_board.set_colors(self.planned_colors)
        self.worked_board.set_colors(self.worked_colors)
        self.planned_board.set_colors(self.planned_colors)
        self.worked_year_board.set_extra_info_colors(self.extra_info_colors)
        self.planned_year_board.set_extra_info_colors(self.extra_info_colors)
        self.worked_board.set_extra_info_colors(self.extra_info_colors)
        self.planned_board.set_extra_info_colors(self.extra_info_colors)
        self.worked_year_board.set_mode("worked")
        self.planned_year_board.set_mode("planned")
        self.worked_board.set_mode("worked")
        self.planned_board.set_mode("planned")
        # Performance: refresh alleen de zichtbare kalendersectie; andere secties verversen bij activatie.
        if hasattr(self, "tabs"):
            tab_idx = self.tabs.currentIndex()
            if tab_idx == 0:
                if self.worked_month_tabbar.currentIndex() == 0:
                    self.worked_year_board.refresh()
                else:
                    self.worked_board.refresh()
            elif tab_idx == 1:
                if self.planned_month_tabbar.currentIndex() == 0:
                    self.planned_year_board.refresh()
                else:
                    self.planned_board.refresh()
        else:
            self.worked_year_board.refresh()
            self.worked_board.refresh()
        if self.timer_host:
            self.timer_host.panel.apply_colors(self.active_colors())
        self.lbl_saldo.setText(self.store.get_saldo_text())
        self._refresh_mode_toggle_style()
        self._apply_mode_theme()
        if hasattr(self, "tabs") and self.tabs.currentIndex() == 2:
            self.refresh_dashboard()

    def active_colors(self) -> dict[str, str]:
        return self.planned_colors if self.mode == "planned" else self.worked_colors

    def _refresh_mode_toggle_style(self):
        return

    def _apply_current_month_accent(self):
        today_month = date.today().month
        normal = QColor("#d6deea")
        accent = QColor("#56d6ff")
        for tabbar in (self.worked_month_tabbar, self.planned_month_tabbar):
            tabbar.setTabTextColor(0, normal)
            for i in range(1, tabbar.count()):
                base = MONTHS_NL[i - 1]
                tabbar.setTabText(i, base)
                tabbar.setTabTextColor(i, normal)
            current_idx = today_month  # index 0 is Jaaroverzicht
            if 1 <= current_idx < tabbar.count():
                tabbar.setTabTextColor(current_idx, accent)
            tabbar.update()

    def _apply_mode_theme(self):
        mode_theme = "worked" if self.mode == "worked" else "planned"
        self.setProperty("modeTheme", mode_theme)
        for w in (
            self,
            self.worked_year_board.scroll,
            self.worked_year_board.host,
            self.worked_board.scroll,
            self.worked_board.host,
            self.planned_year_board.scroll,
            self.planned_year_board.host,
            self.planned_board.scroll,
            self.planned_board.host,
        ):
            w.setProperty("modeTheme", mode_theme)
            w.update()

    def set_mode(self, mode: str):
        if mode not in ("worked", "planned"):
            return
        self.tabs.setCurrentIndex(1 if mode == "planned" else 0)

    def toggle_mode(self):
        self.set_mode("planned" if self.mode == "worked" else "worked")

    def on_worked_period_tab_changed(self, idx: int):
        is_year = idx == 0
        self.worked_year_board.setVisible(is_year)
        self.worked_board.setVisible(not is_year)
        if is_year:
            self.worked_year_board.refresh()
            self.lbl_status.setText("Uren: Jaaroverzicht")
            return
        month_idx = idx - 1
        self.worked_board.set_month(month_idx + 1)
        self.worked_board.set_mode("worked")
        self.worked_board.refresh()
        self.lbl_status.setText(f"Uren maand: {MONTHS_NL[month_idx]}")

    def on_planned_period_tab_changed(self, idx: int):
        is_year = idx == 0
        self.planned_year_board.setVisible(is_year)
        self.planned_board.setVisible(not is_year)
        if is_year:
            self.planned_year_board.refresh()
            self.lbl_status.setText("Planning: Jaaroverzicht")
            return
        month_idx = idx - 1
        self.planned_board.set_month(month_idx + 1)
        self.planned_board.set_mode("planned")
        self.planned_board.refresh()
        self.lbl_status.setText(f"Planning maand: {MONTHS_NL[month_idx]}")

    def on_tab_change(self, idx: int):
        if idx == 0:
            self.mode = "worked"
            self.btn_pattern.setVisible(False)
            self.btn_pattern_edit.setVisible(False)
            self.on_worked_period_tab_changed(self.worked_month_tabbar.currentIndex())
        elif idx == 1:
            self.mode = "planned"
            self.btn_pattern.setVisible(True)
            self.btn_pattern_edit.setVisible(True)
            self.on_planned_period_tab_changed(self.planned_month_tabbar.currentIndex())
        else:
            self.mode = "planned"
            self.btn_pattern.setVisible(False)
            self.btn_pattern_edit.setVisible(False)
            self.refresh_dashboard()
        self._refresh_mode_toggle_style()
        self._apply_mode_theme()
        self.lbl_status.setText("Tab: Uren" if idx == 0 else ("Tab: Planning" if idx == 1 else "Tab: Dashboard"))

    def _seconds_to_hhmm(self, seconds: int) -> str:
        return minutes_to_hhmm(int(max(0, seconds)) // 60)

    def _dashboard_health_snapshot(self) -> dict[str, str]:
        main_path = self.store.path
        bak_path = f"{main_path}.bak"
        backups_dir = os.path.join(os.path.dirname(main_path), "_backups")

        def _fmt_ts(path: str) -> str:
            try:
                ts = os.path.getmtime(path)
                return datetime.fromtimestamp(ts).strftime("%d-%m-%Y %H:%M")
            except Exception:
                return "onbekend"

        snapshots = []
        if os.path.isdir(backups_dir):
            try:
                base = os.path.splitext(os.path.basename(main_path))[0].lower() + "_"
                snapshots = [p for p in os.listdir(backups_dir) if p.lower().startswith(base) and p.lower().endswith(".xlsx")]
            except Exception:
                snapshots = []
        snap_count = len(snapshots)
        status = "Gezond"
        if not os.path.exists(main_path):
            status = "Fout: hoofdbestand ontbreekt"
        elif not os.path.exists(bak_path):
            status = "Waarschuwing: .bak ontbreekt"
        elif snap_count == 0:
            status = "Waarschuwing: nog geen snapshots"

        return {
            "save": _fmt_ts(main_path),
            "backup": _fmt_ts(bak_path),
            "snapshots": str(snap_count),
            "status": status,
        }

    def on_dashboard_period_changed(self, _idx: int):
        today = date.today()
        mode = self.dash_period.currentText() if hasattr(self, "dash_period") else "Huidige maand"
        if mode == "Laatste 30 dagen":
            start, end = today - timedelta(days=29), today
        elif mode == "Dit jaar":
            start, end = date(self.year, 1, 1), date(self.year, 12, 31)
        elif mode == "Aangepast":
            return
        else:
            start = date(today.year, today.month, 1)
            end = date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
        if hasattr(self, "dash_from"):
            self.dash_from.setText(start.strftime("%d-%m-%Y"))
        if hasattr(self, "dash_to"):
            self.dash_to.setText(end.strftime("%d-%m-%Y"))
        self.refresh_dashboard()

    def _dashboard_date_range(self) -> tuple[date, date]:
        today = date.today()
        start_txt = self.dash_from.text().strip() if hasattr(self, "dash_from") else ""
        end_txt = self.dash_to.text().strip() if hasattr(self, "dash_to") else ""
        start = parse_nl_date(start_txt) if start_txt else None
        end = parse_nl_date(end_txt) if end_txt else None
        if start and end:
            if start <= end:
                return start, end
            return end, start
        if start and not end:
            return start, today
        if end and not start:
            return date(end.year, 1, 1), end
        start = date(today.year, today.month, 1)
        end = date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
        return start, end

    def refresh_dashboard(self):
        if not hasattr(self, "tile_mix"):
            return
        # Dashboard aggregeert per kalenderdag. Gewerkte uren komen primair uit
        # day.worked (inclusief handmatige correcties), terwijl idle/call uit
        # data_log komen. Zo blijven handmatige uren zichtbaar in alle KPI's.
        start, end = self._dashboard_date_range()
        include_weekend = bool(self.dash_include_weekend.isChecked())

        rows: list[tuple[date, int, int, int]] = []
        for dt in daterange(start, end):
            if dt < start or dt > end:
                continue
            if (not include_weekend) and dt.weekday() >= 5:
                continue
            day = self.store.get_day(dt)
            log = self.store.get_timer_log(dt)
            planned_total_m = hhmm_to_minutes(day.w) + hhmm_to_minutes(day.v) + hhmm_to_minutes(day.z)
            worked_s = hhmm_to_minutes(day.worked) * 60
            idle_s = int(log.get("idle", 0))
            call_s = int(log.get("call", 0))
            # Neem alle relevante dagen mee binnen de actieve dashboardfilter:
            # gepland (werk/vrij), gewerkt, pauze of call.
            if worked_s <= 0 and idle_s <= 0 and call_s <= 0 and planned_total_m <= 0:
                continue
            rows.append((dt, int(worked_s), idle_s, call_s))
        rows.sort(key=lambda x: x[0])

        rows_with_plan: list[tuple[date, int, int, int, int, int]] = []
        for dt, w_s, i_s, c_s in rows:
            planned_min = hhmm_to_minutes(self.store.get_day(dt).w)
            worked_min = int(max(0, w_s) // 60)
            delta_min = worked_min - planned_min
            rows_with_plan.append((dt, planned_min, worked_min, delta_min, i_s, c_s))

        total_work = sum(r[1] for r in rows)
        total_idle = sum(r[2] for r in rows)
        total_call = sum(r[3] for r in rows)
        planned_sel = self.store.planned_work_minutes_between(start, end)
        total_all = max(1, total_work + total_idle + total_call)
        active_days = len(rows)
        avg_work = int(total_work / active_days) if active_days else 0
        prod_ratio = int(round((total_work / total_all) * 100))
        total_variance_min = sum(r[3] for r in rows_with_plan)
        over_plan_min = sum(max(0, r[3]) for r in rows_with_plan)
        under_plan_min = sum(max(0, -r[3]) for r in rows_with_plan)
        on_target_days = sum(1 for r in rows_with_plan if r[3] == 0)
        budget = self.store.get_budget_overview()
        vacation_left_total = budget["vacation_stat_left"] + budget["vacation_extra_left"]

        self.lbl_dash_work.setText(f"Totaal gewerkt: {self._seconds_to_hhmm(total_work)}")
        self.lbl_dash_planned.setText(f"Totaal gepland: {minutes_to_hhmm(planned_sel)}")
        self.lbl_dash_idle.setText(f"Totaal pauze: {self._seconds_to_hhmm(total_idle)}")
        self.lbl_dash_call.setText(f"Totaal call: {self._seconds_to_hhmm(total_call)}")
        self.lbl_dash_ratio.setText(f"Productief ratio: {prod_ratio}%")
        self.lbl_dash_days.setText(f"Actieve dagen: {active_days}")
        self.lbl_dash_avg.setText(f"Gemiddeld/dag: {self._seconds_to_hhmm(avg_work)}")
        sign = "+" if total_variance_min >= 0 else "-"
        self.lbl_dash_variance.setText(
            f"Afwijking gepland vs gewerkt: {sign}{minutes_to_hhmm(abs(total_variance_min))}"
        )
        self.lbl_dash_budget.setText(
            f"Contract/week: {format_hours_int(budget['contract_week'])}u | "
            f"Nog te plannen jaar: {format_hours_int(budget['to_plan_year'])}u | "
            f"Vrij over: {format_hours_int(vacation_left_total)}u"
        )

        health = self._dashboard_health_snapshot()
        self.lbl_health_save.setText(f"Laatste save: {health['save']}")
        self.lbl_health_backup.setText(f"Laatste backup: {health['backup']}")
        self.lbl_health_snapshots.setText(f"Snapshots: {health['snapshots']}")
        self.lbl_health_status.setText(f"Status: {health['status']}")
        if health["status"].startswith("Gezond"):
            self.lbl_health_status.setStyleSheet("color:#86d29b;")
        elif health["status"].startswith("Waarschuwing"):
            self.lbl_health_status.setStyleSheet("color:#f3c27a;")
        else:
            self.lbl_health_status.setStyleSheet("color:#ff9492;")

        self.tile_mix.set_data(
            [("Werk", total_work, "#3f6ea6"), ("Pauze", total_idle, "#b38e52"), ("Call", total_call, "#b4535a")],
            f"{prod_ratio}%",
            "Aandeel per categorie in de hele selectie.",
        )
        on_target_min = sum(min(r[1], r[2]) for r in rows_with_plan)
        self.tile_mix2.set_data(
            [
                ("Op schema", on_target_min, "#4e7fba"),
                ("Boven planning", over_plan_min, "#4e9d6d"),
                ("Onder planning", under_plan_min, "#b86a62"),
            ],
            f"{on_target_days} dagen",
            "Balans tussen planning en realisatie over de gekozen periode.",
        )
        self.tile_cat.set_data(
            [
                ("Boven planning", over_plan_min, "#4e9d6d"),
                ("Onder planning", under_plan_min, "#b86a62"),
                ("Op schema (dagen)", on_target_days, "#4e7fba"),
            ],
            f"Totale afwijking: {sign}{minutes_to_hhmm(abs(total_variance_min))}",
        )

        self.dash_table.setRowCount(0)
        # Detailset toont alle dagen in de selectie zodat niets wegvalt.
        # Sorteer op datum (nieuwste bovenaan) voor snelle dagelijkse controle.
        for dt, planned_m, worked_m, delta_m, i_s, c_s in sorted(rows_with_plan, key=lambda x: x[0], reverse=True):
            r = self.dash_table.rowCount()
            self.dash_table.insertRow(r)
            self.dash_table.setItem(r, 0, QTableWidgetItem(dt.strftime("%d-%m-%Y")))
            self.dash_table.setItem(r, 1, QTableWidgetItem(minutes_to_hhmm(planned_m)))
            self.dash_table.setItem(r, 2, QTableWidgetItem(minutes_to_hhmm(worked_m)))
            d_sign = "+" if delta_m >= 0 else "-"
            self.dash_table.setItem(r, 3, QTableWidgetItem(f"{d_sign}{minutes_to_hhmm(abs(delta_m))}"))
            self.dash_table.setItem(r, 4, QTableWidgetItem(self._seconds_to_hhmm(i_s)))
            self.dash_table.setItem(r, 5, QTableWidgetItem(self._seconds_to_hhmm(c_s)))

    def edit_day(self, dt: date):
        existing_reason = self.store.day_reason(dt)
        allow_tvt = self.store.get_saldo_minutes() > 0
        dlg = DayEditDialog(
            self,
            dt,
            self.store.get_day(dt),
            self.store.get_day_limit(dt),
            existing_reason,
            self.store.get_extra_info(dt),
            self.extra_info_enabled,
            self.extra_info_colors,
            free_budget_text=self.store.get_vacation_budget_text(),
            allow_tvt=allow_tvt,
        )
        if dlg.exec() != QDialog.Accepted or not dlg.data_out:
            return
        self.store.set_day(dt, dlg.data_out, dlg.reason_out, dlg.extra_info_out)
        self.refresh_all()
        self.lbl_status.setText(f"Dag opgeslagen: {dt.strftime('%d-%m-%Y')}")

    def open_work_pattern_editor(self):
        dlg = WorkPatternDialog(self, self.store, self.year, apply_callback=self.apply_pattern_year)
        if dlg.exec() == QDialog.Accepted:
            self.store.load_patterns()
            self.refresh_all()
            self.lbl_status.setText("Werkpatroon opgeslagen")

    def open_bulk_planning(self, start_date: date | None = None, end_date: date | None = None):
        if not isinstance(start_date, date):
            start_date = None
        if not isinstance(end_date, date):
            end_date = None
        allow_tvt = self.store.get_saldo_minutes() > 0
        weekday_limits = [
            hhmm_to_minutes(normalize_hhmm(self.store.weekday_pattern.get(i, {}).get("M", "08:00"), "08:00"))
            for i in range(7)
        ]
        dlg = BulkPlanningDialog(
            self,
            self.year,
            self.extra_info_enabled,
            allow_tvt=allow_tvt,
            start_date=start_date,
            end_date=end_date,
            weekday_limits=weekday_limits,
        )
        if dlg.exec() != QDialog.Accepted or not dlg.result_data:
            return
        data = dlg.result_data
        d_from: date = data["start"]
        d_to: date = data["end"]
        skip_weekend = bool(data["skip_weekend"])
        skip_holiday = bool(data["skip_holiday"])
        rows_cfg = {int(r.get("weekday", -1)): r for r in (data.get("rows_cfg") or [])}

        changed = 0
        skipped = 0
        clamped = 0

        for dt in daterange(d_from, d_to):
            if dt.year != self.year:
                skipped += 1
                continue
            cfg = rows_cfg.get(dt.weekday())
            if not cfg or not bool(cfg.get("active")):
                continue
            if skip_weekend and dt.weekday() >= 5:
                continue
            if skip_holiday and dt in self.store.nl_holidays:
                continue

            cur = self.store.get_day(dt)
            day_limit = self.store.get_day_limit(dt)
            if day_limit <= 0:
                skipped += 1
                continue

            target_w = hhmm_to_minutes(str(cfg.get("work_hours", "00:00")))
            target_v = hhmm_to_minutes(str(cfg.get("free_hours", "00:00")))

            # Dagmax uit werkpatroon is leidend. Bij overschrijding reduceren we eerst werk,
            # zodat ingepland vrij zo veel mogelijk behouden blijft.
            effective_limit = max(0, day_limit - hhmm_to_minutes(cur.z))
            overflow = max(0, (target_w + target_v) - effective_limit)
            if overflow > 0:
                cut_w = min(target_w, overflow)
                target_w -= cut_w
                overflow -= cut_w
                if overflow > 0:
                    target_v = max(0, target_v - overflow)
                clamped += 1

            wd_name = WEEKDAY_NAMES_NL[dt.weekday()]
            reason_tpl = str(cfg.get("reason_tpl", "") or "").strip()
            reason = reason_tpl.replace("{datum}", dt.strftime("%d-%m-%Y")).replace("{weekdag}", wd_name) if reason_tpl else ""
            out = DayData(
                w=minutes_to_hhmm(max(0, target_w)),
                v=minutes_to_hhmm(max(0, target_v)),
                z=cur.z,
                worked=cur.worked,
            )
            current_extra = self.store.get_extra_info(dt)
            extra_mode = str(cfg.get("extra_mode", "(ongewijzigd)")).strip()
            if extra_mode == "(ongewijzigd)":
                new_extra = current_extra
            elif extra_mode == "(leegmaken)":
                new_extra = ""
            else:
                new_extra = extra_mode
            self.store.set_day(dt, out, reason, new_extra)
            changed += 1

        self.refresh_all()
        self.lbl_status.setText(
            f"Bulk planning: {changed} dagen aangepast, {skipped} overgeslagen, {clamped} afgekapt op dagmax"
        )

    def open_extra_info_settings(self):
        dlg = ExtraInfoSettingsDialog(self, self.store.extra_info_options, self.extra_info_enabled)
        if dlg.exec() != QDialog.Accepted:
            return
        self.store.save_extra_info_options(dlg.options, dlg.enabled_options)
        self.extra_info_enabled = list(dlg.enabled_options)
        self.ensure_extra_info_enabled()
        self.ensure_extra_info_colors()
        self.save_color_settings()
        self.refresh_all()
        self.lbl_status.setText("Aanvullende info opties opgeslagen")

    def open_contract_budget_settings(self):
        dlg = ContractBudgetDialog(
            self,
            self.store.employment_pct,
            self.store.fulltime_week_minutes,
            self.store.contract_year_minutes,
            self.store.vacation_stat_minutes,
            self.store.vacation_extra_minutes,
            self.store.carry_hours_prev_minutes,
            self.store.carry_vac_prev_minutes,
            self.store.workdays_mask,
        )
        if dlg.exec() != QDialog.Accepted:
            return
        self.store.save_contract_budget_settings(
            dlg.employment_pct,
            dlg.fulltime_week,
            dlg.contract_year,
            dlg.vacation_stat,
            dlg.vacation_extra,
            dlg.carry_hours_prev,
            dlg.carry_vac_prev,
            dlg.workdays_mask,
        )
        self.refresh_all()
        self.lbl_status.setText("Contract- en vakantiebudget opgeslagen")

    def open_school_region_settings(self):
        dlg = SchoolRegionDialog(self, self.school_region)
        if dlg.exec() != QDialog.Accepted:
            return
        self.school_region = dlg.region
        self.store.school_region = self.school_region
        self.store.load_school_holidays()
        if not self.store.school_vakanties:
            added = self.store.seed_school_holidays_from_api(region_filter=None)
            if added:
                self.store.safe_save()
                self.store.load_school_holidays()
        self.save_color_settings()
        self.refresh_all()
        self.lbl_status.setText(f"Schoolvakantie regio: {self.school_region}")

    def open_timer_settings(self):
        dlg = TimerSettingsDialog(
            self,
            self.idle_threshold_sec,
            self.include_lockscreen_idle,
            self.include_sleep_idle,
        )
        if dlg.exec() != QDialog.Accepted:
            return
        self.idle_threshold_sec = dlg.idle_threshold_sec
        self.include_lockscreen_idle = bool(dlg.include_lockscreen_idle)
        self.include_sleep_idle = bool(dlg.include_sleep_idle)
        if self.timer_host and hasattr(self.timer_host, "panel"):
            self.timer_host.panel.set_idle_threshold(self.idle_threshold_sec)
            self.timer_host.panel.set_idle_policy(self.include_lockscreen_idle, self.include_sleep_idle)
        self.save_color_settings()
        self.lbl_status.setText(
            f"Pauze-detectie: {self.idle_threshold_sec // 60} min | "
            f"lockscreen {'aan' if self.include_lockscreen_idle else 'uit'} | "
            f"sleep {'aan' if self.include_sleep_idle else 'uit'}"
        )

    def on_timer_enabled_toggled(self, checked: bool):
        self.timer_enabled = bool(checked)
        if self.timer_host and hasattr(self.timer_host, "panel"):
            self.timer_host.panel.set_tracking_enabled(self.timer_enabled)
        if not self.timer_enabled:
            if self.timer_host:
                self.timer_host.disable_timer_ui()
                self.timer_host = None
            self.showNormal()
            self.raise_()
            self.activateWindow()
        else:
            if self.timer_host is None:
                timer = TimerWindow(self)
                self.timer_host = timer
                self.hide()
                timer.show()
        self.save_color_settings()
        self.lbl_status.setText("Timer meten: aan" if self.timer_enabled else "Timer meten: uit")

    def open_glass_settings(self):
        dlg = GlassSettingsDialog(self, self.inactive_glass_opacity)
        if dlg.exec() != QDialog.Accepted:
            return
        self.inactive_glass_opacity = dlg.inactive_glass_opacity
        if self.timer_host and hasattr(self.timer_host, "panel"):
            self.timer_host.panel.set_glass_limits(self.min_inactive_glass_opacity, self.max_inactive_glass_opacity)
            self.timer_host.panel.set_glass_opacity(self.inactive_glass_opacity)
            self.timer_host.update_focus_glass()
        self.save_color_settings()
        self.lbl_status.setText(
            f"Doorzichtigheid: {int(round(self.inactive_glass_opacity * 100)):02d}% (vast min {int(GLASS_OPACITY_MIN * 100)}% / max {int(GLASS_OPACITY_MAX * 100)}%)"
        )

    def save(self):
        self.store.safe_save()
        self.lbl_status.setText("Opgeslagen")
        self.lbl_saldo.setText(self.store.get_saldo_text())

    def export_copy(self):
        dst = self.store.export_copy()
        QMessageBox.information(self, "Export", f"Export gemaakt:\n{dst}")
        self.lbl_status.setText("Export gereed")

    def goto_today(self):
        t = date.today()
        if t.year != self.year:
            return
        self.worked_month_tabbar.setCurrentIndex(t.month)
        self.planned_month_tabbar.setCurrentIndex(t.month)
        if self.tabs.currentIndex() == 2:
            self.tabs.setCurrentIndex(0)

    def apply_pattern_year(self, start_date: date | None = None, end_date: date | None = None):
        ws = self.store.wb["Planning"]
        rows_by_date = {}
        updated_dates = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = normalize_to_date(row[0] if row else None)
            if d:
                rows_by_date[d] = normalize_hhmm((row[1] if len(row) > 1 else "00:00") or "00:00")

        for m in range(1, 13):
            days = calendar.monthrange(self.year, m)[1]
            for d in range(1, days + 1):
                dt = date(self.year, m, d)
                if start_date and dt < start_date:
                    continue
                if end_date and dt > end_date:
                    continue
                if dt in self.store.nl_holidays:
                    continue
                pat = self.store.weekday_pattern[dt.weekday()]
                w = normalize_hhmm(pat["W"])
                if hhmm_to_minutes(w) > self.store.get_day_limit(dt):
                    continue
                rows_by_date[dt] = w
                updated_dates.add(dt)
                key = dt.strftime("%Y-%m-%d")
                self.store.planned_data[key] = {"W": w, "V": "00:00", "Z": "00:00"}
                self.store.vakantie_dagen.pop(key, None)

        sorted_rows = sorted(rows_by_date.items(), key=lambda x: x[0])
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        for i, (dt, vals) in enumerate(sorted_rows, start=2):
            ws.cell(i, 1, dt).number_format = "DD-MM-YYYY"
            ws.cell(i, 2, vals)

        ws_vrij = self.store.wb["vrije dagen"]
        rows_to_delete = []
        for row in ws_vrij.iter_rows(min_row=2):
            r_dt = normalize_to_date(row[0].value)
            if r_dt in updated_dates:
                rows_to_delete.append(row[0].row)
        for r in sorted(rows_to_delete, reverse=True):
            ws_vrij.delete_rows(r, 1)

        self.store.load_free_days()
        self.store.safe_save()
        self.refresh_all()
        self.lbl_status.setText("Werkpatroon toegepast")
        if start_date and end_date:
            QMessageBox.information(self, "Klaar", f"Werkpatroon toegepast van {start_date.strftime('%d-%m-%Y')} t/m {end_date.strftime('%d-%m-%Y')}.")
        else:
            QMessageBox.information(self, "Klaar", "Werkpatroon toegepast op het jaar.")

    def colors_path(self) -> str:
        return os.path.join(self.base_dir, "calendar_colors.json")

    def color_settings_candidates(self) -> list[str]:
        """Zoekvolgorde voor kleurinstellingen.

        Primair lezen we naast het script. Als fallback proberen we de lokale
        Downloads-map, zodat bestaande gebruikersinstellingen niet kwijt lijken
        wanneer het script vanaf een andere map wordt gestart.
        """
        primary = self.colors_path()
        out = [primary]
        try:
            downloads = os.path.join(os.path.expanduser("~"), "Downloads", "calendar_colors.json")
            if downloads not in out:
                out.append(downloads)
        except Exception:
            pass
        return out

    def ensure_extra_info_enabled(self):
        self.extra_info_enabled = normalize_extra_info_enabled(self.store.extra_info_options, self.extra_info_enabled)

    def ensure_extra_info_colors(self):
        normalized = {}
        for k, v in self.extra_info_colors.items():
            if isinstance(k, str) and isinstance(v, str) and v.startswith("#"):
                normalized[k.strip()] = v
        self.extra_info_colors = normalized
        for opt in self.store.extra_info_options:
            if opt not in self.extra_info_colors:
                self.extra_info_colors[opt] = default_extra_info_color(opt)

    def load_color_settings(self) -> tuple[dict[str, str], dict[str, str], dict[str, str], int, bool, bool, list[str], str, float, float, float, bool]:
        worked_cfg = dict(WORKED_COLOR_DEFAULTS)
        planned_cfg = dict(PLANNED_COLOR_DEFAULTS)
        extra_cfg: dict[str, str] = {}
        idle_threshold_sec = 60
        include_lockscreen_idle = True
        include_sleep_idle = False
        extra_enabled: list[str] = []
        school_region = "zuid"
        inactive_glass_opacity = 0.72
        min_inactive_glass_opacity = GLASS_OPACITY_MIN
        max_inactive_glass_opacity = GLASS_OPACITY_MAX
        timer_enabled = True
        p = ""
        for candidate in self.color_settings_candidates():
            if os.path.exists(candidate):
                p = candidate
                break
        if p:
            try:
                with open(p, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    if isinstance(data.get("worked_colors"), dict) or isinstance(data.get("planned_colors"), dict):
                        source_worked = data.get("worked_colors") if isinstance(data.get("worked_colors"), dict) else {}
                        source_planned = data.get("planned_colors") if isinstance(data.get("planned_colors"), dict) else {}
                        # Als 1 van beide ontbreekt, mirroren we de aanwezige set.
                        if not source_worked and source_planned:
                            source_worked = dict(source_planned)
                        if not source_planned and source_worked:
                            source_planned = dict(source_worked)
                        for k, v in source_worked.items():
                            if k in worked_cfg and isinstance(v, str) and v.startswith("#"):
                                worked_cfg[k] = v
                        for k, v in source_planned.items():
                            if k in planned_cfg and isinstance(v, str) and v.startswith("#"):
                                planned_cfg[k] = v
                    else:
                        source_colors = data.get("colors") if isinstance(data.get("colors"), dict) else data
                        for k, v in source_colors.items():
                            if isinstance(v, str) and v.startswith("#") and k in planned_cfg:
                                planned_cfg[k] = v
                                worked_cfg[k] = v
                    source_extra = data.get("extra_info_colors") if isinstance(data.get("extra_info_colors"), dict) else {}
                    for k, v in source_extra.items():
                        if isinstance(k, str) and isinstance(v, str) and v.startswith("#"):
                            extra_cfg[k.strip()] = v
                    if isinstance(data.get("idle_threshold_sec"), int):
                        idle_threshold_sec = max(15, min(3600, int(data.get("idle_threshold_sec"))))
                    if isinstance(data.get("include_lockscreen_idle"), bool):
                        include_lockscreen_idle = bool(data.get("include_lockscreen_idle"))
                    if isinstance(data.get("include_sleep_idle"), bool):
                        include_sleep_idle = bool(data.get("include_sleep_idle"))
                    src_enabled = data.get("extra_info_enabled")
                    if isinstance(src_enabled, list):
                        extra_enabled = [str(x).strip() for x in src_enabled if str(x).strip()]
                    sr = str(data.get("school_region", "zuid")).strip().casefold()
                    if sr in {"noord", "midden", "zuid"}:
                        school_region = sr
                    if isinstance(data.get("inactive_glass_opacity"), (int, float)):
                        inactive_glass_opacity = max(GLASS_OPACITY_MIN, min(GLASS_OPACITY_MAX, float(data.get("inactive_glass_opacity"))))
                    if isinstance(data.get("timer_enabled"), bool):
                        timer_enabled = bool(data.get("timer_enabled"))
            except:
                pass
        return (
            worked_cfg,
            planned_cfg,
            extra_cfg,
            idle_threshold_sec,
            include_lockscreen_idle,
            include_sleep_idle,
            extra_enabled,
            school_region,
            inactive_glass_opacity,
            min_inactive_glass_opacity,
            max_inactive_glass_opacity,
            timer_enabled,
        )

    def save_color_settings(self):
        try:
            with open(self.colors_path(), "w", encoding="utf-8") as f:
                json.dump(
                    {
                        "worked_colors": self.worked_colors,
                        "planned_colors": self.planned_colors,
                        "extra_info_colors": self.extra_info_colors,
                        "idle_threshold_sec": int(self.idle_threshold_sec),
                        "include_lockscreen_idle": bool(self.include_lockscreen_idle),
                        "include_sleep_idle": bool(self.include_sleep_idle),
                        "extra_info_enabled": list(self.extra_info_enabled),
                        "school_region": self.school_region,
                        "inactive_glass_opacity": float(self.inactive_glass_opacity),
                        "timer_enabled": bool(self.timer_enabled),
                    },
                    f,
                    indent=2,
                )
        except Exception as exc:
            QMessageBox.warning(self, "Fout", f"Kleuren opslaan mislukt:\n{exc}")

    def open_color_settings(self):
        self.ensure_extra_info_colors()
        dlg = ColorSettingsDialog(
            self,
            self.planned_colors,
            self.store.extra_info_options,
            self.extra_info_colors,
        )
        if dlg.exec() == QDialog.Accepted:
            self.planned_colors = dict(dlg.colors)
            self.worked_colors = dict(self.planned_colors)
            self.extra_info_colors = dict(dlg.extra_info_colors)
            self.ensure_extra_info_colors()
            self.save_color_settings()
            self.refresh_all()
            self.lbl_status.setText("Kleuren bijgewerkt")

    def reset_color_settings(self):
        self.worked_colors = dict(WORKED_COLOR_DEFAULTS)
        self.planned_colors = dict(PLANNED_COLOR_DEFAULTS)
        self.extra_info_colors = {}
        self.ensure_extra_info_colors()
        self.save_color_settings()
        self.refresh_all()
        self.lbl_status.setText("Kleuren hersteld")

    def about(self):
        QMessageBox.information(
            self,
            "Over",
            (
                "Tijdplanner Pro (PySide6)\n"
                f"Versie: {APP_VERSION}\n"
                f"Auteur: {APP_AUTHOR}\n"
                f"Laatste update: {APP_LAST_UPDATE}\n\n"
                "Roadmap: migratie van Excel-opslag naar SQLite."
            ),
        )

    def toggle_maximize_restore(self):
        if self.isMaximized():
            self.showNormal()
        else:
            self.showMaximized()

    def eventFilter(self, obj, event):
        if hasattr(self, "worked_month_tabbar") and event.type() == QEvent.Wheel:
            # Wheel->maandwissel mag alleen op de maand-tabbars zelf gebeuren.
            # In jaar/maandweergaven wil de gebruiker normaal verticaal kunnen scrollen door content;
            # daarom kapen we het wheel-event buiten de tabbar niet meer af.
            if obj in {self.worked_month_tabbar, self.planned_month_tabbar}:
                tabbar = self.worked_month_tabbar if obj == self.worked_month_tabbar else self.planned_month_tabbar
                delta = event.angleDelta().y()
                idx = tabbar.currentIndex()
                if delta < 0 and idx < tabbar.count() - 1:
                    tabbar.setCurrentIndex(idx + 1)
                elif delta > 0 and idx > 0:
                    tabbar.setCurrentIndex(idx - 1)
                return True

        if obj == self.menuBar():
            menu_popup_open = any(m.isVisible() for m in self.menuBar().findChildren(QMenu))
            if event.type() == QEvent.MouseButtonPress and event.button() == Qt.LeftButton:
                # Alleen slepen vanaf lege menubalkruimte; menu-items blijven klikbaar.
                if (
                    not menu_popup_open
                    and self.menuBar().activeAction() is None
                    and self.menuBar().actionAt(event.position().toPoint()) is None
                    and not self.isMaximized()
                    and not self.isFullScreen()
                ):
                    self._drag_active = True
                    self._drag_offset = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
                    return True
            elif event.type() == QEvent.MouseMove and self._drag_active:
                if menu_popup_open:
                    self._drag_active = False
                elif not self.isMaximized() and not self.isFullScreen():
                    self.move(event.globalPosition().toPoint() - self._drag_offset)
                    return True
            elif event.type() == QEvent.MouseButtonRelease and self._drag_active:
                self._drag_active = False
                return True
        return super().eventFilter(obj, event)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if self.isVisible() and not self._sliding and self.width() > 300 and self.height() > 200:
            self._planner_target_geometry = QRect(self.geometry())

    def moveEvent(self, event):
        super().moveEvent(event)
        if self.isVisible() and not self._sliding and self.width() > 300 and self.height() > 200:
            self._planner_target_geometry = QRect(self.geometry())

    def closeEvent(self, event):
        if self._allow_close:
            event.accept()
            return
        self.hide()
        event.ignore()

    def quit_app(self):
        if self.timer_host:
            self.timer_host.quit_all()
            return
        self._allow_close = True
        self.close()


# ============================================================================
# TIMER SHELL + SYSTEM TRAY
# TimerWindow host het compacte paneel en regelt tray-interacties.
# Het venster is los van MainWindow zodat gebruikers snel kunnen wisselen
# tussen een minimalistische timer-flow en volledige planner-bediening.
# ============================================================================
class WindowsSessionEventFilter(QAbstractNativeEventFilter):
    """Native Windows events voor lock/unlock en suspend/resume."""

    WM_WTSSESSION_CHANGE = 0x02B1
    WM_POWERBROADCAST = 0x0218
    WTS_SESSION_LOCK = 0x0007
    WTS_SESSION_UNLOCK = 0x0008
    PBT_APMSUSPEND = 0x0004
    PBT_APMRESUMEAUTOMATIC = 0x0012
    PBT_APMRESUMESUSPEND = 0x0007

    def __init__(self, on_lock=None, on_unlock=None, on_suspend=None, on_resume=None):
        super().__init__()
        self.on_lock = on_lock
        self.on_unlock = on_unlock
        self.on_suspend = on_suspend
        self.on_resume = on_resume

    def nativeEventFilter(self, event_type, message):
        if sys.platform != "win32":
            return False, 0
        et = bytes(event_type).decode(errors="ignore") if isinstance(event_type, (bytes, bytearray)) else str(event_type)
        if et not in {"windows_generic_MSG", "windows_dispatcher_MSG"}:
            return False, 0
        try:
            msg = wintypes.MSG.from_address(int(message))
        except Exception:
            return False, 0

        if msg.message == self.WM_WTSSESSION_CHANGE:
            if msg.wParam == self.WTS_SESSION_LOCK and callable(self.on_lock):
                self.on_lock()
            elif msg.wParam == self.WTS_SESSION_UNLOCK and callable(self.on_unlock):
                self.on_unlock()
        elif msg.message == self.WM_POWERBROADCAST:
            if msg.wParam == self.PBT_APMSUSPEND and callable(self.on_suspend):
                self.on_suspend()
            elif msg.wParam in {self.PBT_APMRESUMEAUTOMATIC, self.PBT_APMRESUMESUSPEND} and callable(self.on_resume):
                self.on_resume()
        return False, 0


class TimerWindow(QWidget):
    """Los timer-venster dat de planner in/uit laat schuiven.

    Dit venster is de lichte shell rond TimerPanel en system tray gedrag.
    De planner blijft separaat zodat de gebruiker kan kiezen tussen compacte
    timer-focus en een volledig plan-/dashboardvenster.
    """

    def __init__(self, planner: MainWindow):
        super().__init__()
        self.planner = planner
        self._force_quit = False
        self._native_filter = None
        self._wts_registered = False
        self.setObjectName("timerWindow")
        self.setWindowFlags(Qt.Window | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setWindowTitle("Tijdplanner Timer")
        self.setStyleSheet("QWidget#timerWindow { background: transparent; }")

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 10)
        root.setSpacing(8)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        self.panel = TimerPanel(
            self.planner.store,
            on_refresh=self.planner.refresh_all,
            on_toggle_full=self.toggle_planner,
            on_drag_move=self.move_window_by,
            idle_threshold_sec=self.planner.idle_threshold_sec,
        )
        self.panel.set_glass_limits(self.planner.min_inactive_glass_opacity, self.planner.max_inactive_glass_opacity)
        self.panel.set_glass_opacity(self.planner.inactive_glass_opacity)
        self.panel.apply_colors(self.planner.active_colors())
        self.panel.set_idle_policy(self.planner.include_lockscreen_idle, self.planner.include_sleep_idle)
        self.panel.set_tracking_enabled(self.planner.timer_enabled)
        self.panel.height_changed.connect(self.on_panel_height_changed)
        self.panel.timer.timeout.connect(self.update_tray_visual)
        root.addWidget(self.panel)
        self.on_panel_height_changed(self.panel.maximumHeight())
        self.move(self._initial_timer_position())
        self.setup_tray()
        app = QApplication.instance()
        if app is not None and hasattr(app, "applicationStateChanged"):
            app.applicationStateChanged.connect(lambda _state: self.update_focus_glass())
        self.update_focus_glass()
        self._register_system_event_hooks()

    def on_panel_height_changed(self, height: int):
        self.setFixedSize(self.panel.maximumWidth(), max(56, height))

    def move_window_by(self, delta: QPoint):
        self.move(self.pos() + delta)

    def _initial_timer_position(self) -> QPoint:
        # Start linksboven met marge zodat de timer zichtbaar maar niet "in de hoek geplakt" staat.
        screen = QApplication.primaryScreen()
        avail = screen.availableGeometry() if screen else QRect(0, 0, 2560, 1440)
        margin_x = 18
        margin_y = 26
        x = max(avail.left(), avail.left() + margin_x)
        y = max(avail.top(), avail.top() + margin_y)
        return QPoint(x, y)

    def update_focus_glass(self):
        app = QApplication.instance()
        app_active = bool(app is not None and app.applicationState() == Qt.ApplicationActive)
        self.panel.set_inactive_glass(not app_active)
        self.setWindowOpacity(1.0 if app_active else self.planner.inactive_glass_opacity)
        if self.planner.isVisible():
            planner_opacity = min(0.98, self.planner.inactive_glass_opacity + 0.08)
            self.planner.setWindowOpacity(1.0 if app_active else planner_opacity)

    def _register_system_event_hooks(self):
        if sys.platform != "win32":
            return
        app = QApplication.instance()
        if app is None:
            return
        if self._native_filter is None:
            self._native_filter = WindowsSessionEventFilter(
                on_lock=self._on_session_lock,
                on_unlock=self._on_session_unlock,
                on_suspend=self._on_system_suspend,
                on_resume=self._on_system_resume,
            )
            app.installNativeEventFilter(self._native_filter)
        try:
            hwnd = int(self.winId())
            rc = ctypes.windll.wtsapi32.WTSRegisterSessionNotification(wintypes.HWND(hwnd), 0)
            self._wts_registered = bool(rc)
        except Exception:
            self._wts_registered = False

    def _unregister_system_event_hooks(self):
        app = QApplication.instance()
        if sys.platform == "win32" and self._wts_registered:
            try:
                hwnd = int(self.winId())
                ctypes.windll.wtsapi32.WTSUnRegisterSessionNotification(wintypes.HWND(hwnd))
            except Exception:
                pass
            self._wts_registered = False
        if app is not None and self._native_filter is not None:
            try:
                app.removeNativeEventFilter(self._native_filter)
            except Exception:
                pass
            self._native_filter = None

    def _on_session_lock(self):
        self.panel.set_os_context(locked=True)
        if hasattr(self.planner, "lbl_status"):
            self.planner.lbl_status.setText("Sessie vergrendeld: idle-detectie loopt door")

    def _on_session_unlock(self):
        self.panel.set_os_context(locked=False)
        if hasattr(self.planner, "lbl_status"):
            self.planner.lbl_status.setText("Sessie ontgrendeld: idle-detectie actief")

    def _on_system_suspend(self):
        self.panel.set_os_context(sleeping=True)
        if hasattr(self.planner, "lbl_status"):
            self.planner.lbl_status.setText("Systeem in slaapstand: timerdetectie tijdelijk onderdrukt")

    def _on_system_resume(self):
        self.panel.set_os_context(sleeping=False)
        if hasattr(self.planner, "lbl_status"):
            self.planner.lbl_status.setText("Systeem actief: timerdetectie hervat")

    def _planner_target_near_timer(self, timer_geo: QRect) -> QRect:
        target = QRect(self.planner._planner_target_geometry)

        screen = QApplication.screenAt(timer_geo.center())
        if screen is None:
            screen = QApplication.primaryScreen()
        avail = screen.availableGeometry() if screen else QRect(0, 0, 2560, 1440)

        gap = 8
        margin = 18
        max_w = max(520, avail.width() - (margin * 2))
        max_h = max(420, avail.height() - (margin * 2))
        desired_w = max(1200, target.width())
        desired_h = max(760, target.height())
        width = min(desired_w, max_w)
        height = min(desired_h, max_h)

        def clamp(v: int, lo: int, hi: int) -> int:
            if hi < lo:
                return lo
            return max(lo, min(v, hi))

        candidates = []

        x = timer_geo.right() + 1 + gap
        y = clamp(timer_geo.y() - 12, avail.top(), avail.bottom() + 1 - height)
        candidates.append(QRect(x, y, width, height))  # rechts

        x = timer_geo.x() - gap - width
        y = clamp(timer_geo.y() - 12, avail.top(), avail.bottom() + 1 - height)
        candidates.append(QRect(x, y, width, height))  # links

        x = clamp(timer_geo.x() - 40, avail.left(), avail.right() + 1 - width)
        y = timer_geo.bottom() + 1 + gap
        candidates.append(QRect(x, y, width, height))  # beneden

        x = clamp(timer_geo.x() - 40, avail.left(), avail.right() + 1 - width)
        y = timer_geo.y() - gap - height
        candidates.append(QRect(x, y, width, height))  # boven

        timer_box = QRect(timer_geo)
        timer_box.adjust(-gap, -gap, gap, gap)

        for c in candidates:
            if avail.contains(c) and not c.intersects(timer_box):
                return c

        best = candidates[0]
        best_score = -1
        for c in candidates:
            i = c.intersected(avail)
            visible_area = max(0, i.width()) * max(0, i.height())
            overlap = c.intersected(timer_box)
            overlap_area = max(0, overlap.width()) * max(0, overlap.height())
            score = visible_area - (overlap_area * 8)
            if score > best_score:
                best_score = score
                best = c

        final_rect = QRect(
            clamp(best.x(), avail.left(), avail.right() + 1 - width),
            clamp(best.y(), avail.top(), avail.bottom() + 1 - height),
            width,
            height,
        )

        if final_rect.intersects(timer_box):
            shifts = [
                QPoint((timer_box.right() + gap + 1) - final_rect.left(), 0),
                QPoint((timer_box.left() - gap - 1) - final_rect.right(), 0),
                QPoint(0, (timer_box.bottom() + gap + 1) - final_rect.top()),
                QPoint(0, (timer_box.top() - gap - 1) - final_rect.bottom()),
            ]
            fixed = None
            for d in shifts:
                c = final_rect.translated(d)
                c = QRect(
                    clamp(c.x(), avail.left(), avail.right() + 1 - width),
                    clamp(c.y(), avail.top(), avail.bottom() + 1 - height),
                    width,
                    height,
                )
                if not c.intersects(timer_box):
                    fixed = c
                    break
            if fixed is not None:
                final_rect = fixed

        return self.planner._fit_rect_to_screen(final_rect, screen, margin)

    def toggle_planner(self):
        geo = self.geometry()
        if self.planner.isVisible():
            self.planner.slide_out_to(geo)
            self.panel.btn_open.setChecked(False)
            self.panel.btn_open.setToolTip("Planner tonen")
        else:
            self.planner._planner_target_geometry = self._planner_target_near_timer(geo)
            self.planner.slide_in_from(geo)
            self.panel.btn_open.setChecked(True)
            self.panel.btn_open.setToolTip("Planner verbergen")

    def setup_tray(self):
        if not QSystemTrayIcon.isSystemTrayAvailable():
            self.tray = None
            return
        self.tray = QSystemTrayIcon(self)
        self.tray.setIcon(self.make_tray_icon(self.panel.running, self.panel.work_seconds, self.panel.idle_seconds))
        self.tray.setToolTip("Tijdplanner Pro")
        menu = QMenu(self)
        act_open = QAction("Open timer", self)
        act_toggle = QAction("Toon/verberg planner", self)
        self.act_start = QAction("Start timer", self)
        self.act_pause = QAction("Pauze timer", self)
        act_exit = QAction("Afsluiten", self)
        act_open.triggered.connect(self.show_from_tray)
        act_toggle.triggered.connect(self.toggle_planner)
        self.act_start.triggered.connect(self.tray_start_timer)
        self.act_pause.triggered.connect(self.tray_pause_timer)
        act_exit.triggered.connect(self.quit_all)
        menu.addAction(act_open)
        menu.addAction(act_toggle)
        menu.addSeparator()
        menu.addAction(self.act_start)
        menu.addAction(self.act_pause)
        menu.addSeparator()
        menu.addAction(act_exit)
        self.tray.setContextMenu(menu)
        self.tray.activated.connect(self.on_tray_activated)
        self.tray.show()
        self.update_tray_visual()

    def make_tray_icon(self, running: bool, work_seconds: int, idle_seconds: int) -> QIcon:
        return build_status_icon(running=running, work_seconds=work_seconds, idle_seconds=idle_seconds)

    def update_tray_visual(self):
        running = bool(self.panel.running)
        w = int(self.panel.work_seconds)
        i = int(self.panel.idle_seconds)
        c = int(self.panel.call_seconds)

        icon = self.make_tray_icon(running, w, i)
        self.setWindowIcon(icon)
        self.planner.setWindowIcon(icon)
        app = QApplication.instance()
        if app is not None:
            app.setWindowIcon(icon)

        if not self.tray:
            return
        self.tray.setIcon(icon)
        state = "actief" if running else "gepauzeerd"
        self.tray.setToolTip(
            f"Tijdplanner Pro ({state})\n"
            f"Werk: {seconds_to_hhmmss(w)}\n"
            f"Pauze: {seconds_to_hhmmss(i)}\n"
            f"Call: {seconds_to_hhmmss(c)}"
        )
        if hasattr(self, "act_start"):
            self.act_start.setEnabled(not running)
        if hasattr(self, "act_pause"):
            self.act_pause.setEnabled(running)

    def tray_start_timer(self):
        self.panel.start()
        self.update_tray_visual()

    def tray_pause_timer(self):
        self.panel.pause()
        self.update_tray_visual()

    def on_tray_activated(self, reason):
        if reason == QSystemTrayIcon.DoubleClick:
            self.show_from_tray()

    def show_from_tray(self):
        self.showNormal()
        self.raise_()
        self.activateWindow()

    def quit_all(self):
        self._force_quit = True
        self.panel.pause()
        self._unregister_system_event_hooks()
        if self.tray:
            self.tray.hide()
        self.planner._allow_close = True
        self.planner.close()
        self.close()

    def disable_timer_ui(self):
        self._force_quit = True
        self.panel.pause()
        self._unregister_system_event_hooks()
        if self.tray:
            self.tray.hide()
            self.tray.deleteLater()
            self.tray = None
        self.hide()
        self.close()

    def closeEvent(self, event):
        if self._force_quit:
            self._unregister_system_event_hooks()
            event.accept()
            return
        if self.tray and self.tray.isVisible():
            self.hide()
            event.ignore()
            return
        event.accept()


# ============================================================================
# ENTRYPOINT
# De startroute kiest afhankelijk van timer-instelling tussen compacte timerstart
# of directe plannerstart. Hiermee blijft de opstartervaring voorspelbaar voor
# beide gebruiksscenario's.
# ============================================================================
def main():
    """Applicatie-entrypoint met timer-aan/uit route.

    Start altijd MainWindow als centrale statehouder. Als timer actief is, wordt
    eerst alleen het timer-venster getoond; anders start de planner direct zichtbaar.
    """
    set_windows_app_user_model_id()
    app = QApplication(sys.argv)
    app.setApplicationName("Tijdplanner Pro")
    app.setWindowIcon(build_status_icon(running=True, work_seconds=0, idle_seconds=0))
    if not acquire_single_instance_lock():
        lock = QLockFile(os.path.join(tempfile.gettempdir(), "tijdplanner_pro.instance.lock"))
        pid, _host, _app = lock.getLockInfo()
        txt = "Tijdplanner Pro draait al."
        if int(pid or 0) > 0:
            txt += f"\nActieve sessie PID: {pid}"
        txt += "\n\nWil je de bestaande sessie afsluiten en deze sessie starten?"
        reply = QMessageBox.question(
            None,
            "Al actief",
            txt,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        if not try_takeover_existing_instance():
            QMessageBox.warning(
                None,
                "Overnemen mislukt",
                "Bestaande sessie kon niet afgesloten worden.\nSluit die eerst handmatig.",
            )
            return
    planner = MainWindow()
    if planner.timer_enabled:
        planner.hide()
        timer = TimerWindow(planner)
        planner.timer_host = timer
        timer.show()
    else:
        planner.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
